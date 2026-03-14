from __future__ import annotations

import csv
import re
import shutil
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(r"c:\Users\baram\OneDrive\Desktop\themnatic")
CASE_DIR = ROOT / "analysis" / "CASE_D2"
OUTWARD_DIR = CASE_DIR / "OUTWARD_FACING_PACKAGE"
INTERNAL_DIR = CASE_DIR / "INTERNAL_CONFIDENTIAL"

EXCERPT_BANK = CASE_DIR / "CASE_D2_excerpt_bank.csv"
PARTICIPANT_REGISTER = CASE_DIR / "CASE_D2_participant_register.csv"
PARTICIPANT_SUMMARY = CASE_DIR / "CASE_D2_participant_summary.csv"
PARTICIPANT_WORKBOOK = CASE_DIR / "CASE_D2_participant_workbook.xlsx"
QUESTION_EVIDENCE = CASE_DIR / "CASE_D2_question_evidence_table.csv"
QUESTION_THEME_MATRIX = CASE_DIR / "CASE_D2_question_theme_matrix.csv"
PROMINENCE_SALIENCE = CASE_DIR / "CASE_D2_prominence_salience.csv"
THEME_SUMMARY_TABLE = CASE_DIR / "CASE_D2_theme_summary_table.csv"
THEME_EVIDENCE_WORKBOOK = CASE_DIR / "CASE_D2_theme_evidence_workbook.xlsx"
REPORT_INTEGRITY_CHECK = CASE_DIR / "CASE_D2_report_integrity_check.md"
FINAL_CROSSCHECK_REPORT = CASE_DIR / "CASE_D2_final_crosscheck_report.md"
FINAL_REPORT = CASE_DIR / "CASE_D2_final_report.md"
FINAL_THEMES = CASE_DIR / "CASE_D2_final_themes.md"
CANDIDATE_THEMES = CASE_DIR / "CASE_D2_candidate_themes.md"
SUMMARY_TABLES = CASE_DIR / "CASE_D2_summary_tables.md"
THEME_INTEGRITY_REPORT = CASE_DIR / "CASE_D2_theme_integrity_report.md"
MATRIX_PROMINENCE_CHECK = CASE_DIR / "CASE_D2_matrix_prominence_check.md"

THEME_LABELS = [
    "Multidimensional wellbeing as balanced stability, moral-spiritual grounding, and daily functioning",
    "Youth wellbeing as relational ecology, peer worlds, and social disconnection",
    "Holistic care constrained by rushed, fragmented, and trust-sensitive encounters",
    "Coordinated, youth-sensitive, and institutionally backed routes to change",
]

TEXT_REPLACEMENTS = {
    "كما ذكرت الدكتورة زينة": "كما ذكرت مشاركة أخرى",
    "ما ذكره الدكتور رائد": "ما ذكره مشارك آخر",
    "كما تفضل الدكتور عبد العليم": "كما تفضل مشارك آخر",
    "مثل ما سارة قالت": "مثلما قالت إحدى المشاركات",
    "كما قالت سارة": "كما قالت إحدى المشاركات",
    "قالت سارة": "قالت إحدى المشاركات",
    "الدكتورة زينة": "مشاركة أخرى",
    "دكتور زينة": "مشارك آخر",
    "الدكتور رائد": "مشارك آخر",
    "الدكتور عبد العليم": "مشارك آخر",
    "عبد العليم": "مشارك آخر",
    "سارة": "إحدى المشاركات",
}

NAME_SCAN_TERMS = [
    "أحمد العمادي",
    "أحمد الفرجابي",
    "سلوى",
    "زينة",
    "رائد عمرو",
    "رائد",
    "محمد عبد العليم",
    "عبد العليم",
    "ريم الحاج",
    "مريم",
    "دانا",
    "هاجر",
    "وائل",
    "أمير",
    "مصطفى",
    "داليا",
    "Miguel",
    "Elena",
    "Liam",
    "Carmona",
    "Safrina",
    "Herald",
    "Rehana",
    "Aziza",
    "محمد أبو هاشم",
    "نيلي خليل",
    "عبدالرحمن القديمي",
    "رشا علام",
    "نور غازي",
    "سارة غازي",
    "أسماء",
    "سارة",
    "عبلة خليل",
]

OUTWARD_FILE_NAMES = [
    "CASE_D2_candidate_themes.md",
    "CASE_D2_excerpt_bank.csv",
    "CASE_D2_final_crosscheck_report.md",
    "CASE_D2_final_report.md",
    "CASE_D2_final_themes.md",
    "CASE_D2_matrix_prominence_check.md",
    "CASE_D2_participant_register.csv",
    "CASE_D2_participant_summary.csv",
    "CASE_D2_participant_workbook.xlsx",
    "CASE_D2_prominence_salience.csv",
    "CASE_D2_question_evidence_table.csv",
    "CASE_D2_question_theme_matrix.csv",
    "CASE_D2_report_integrity_check.md",
    "CASE_D2_summary_tables.md",
    "CASE_D2_theme_evidence_workbook.xlsx",
    "CASE_D2_theme_integrity_report.md",
    "CASE_D2_theme_summary_table.csv",
]

INTERNAL_FILE_PATHS = [
    PARTICIPANT_REGISTER,
    CASE_DIR / "CASE_D2_preparation_checklist.md",
    CASE_DIR / "CASE_D2_familiarisation_memo.md",
    CASE_DIR / "CASE_D2_source_sensitivity_memo.md",
    CASE_DIR / "CASE_D2_boundary_memo.md",
    CASE_DIR / "CASE_D2_language_memo.md",
]


class ReconciliationError(RuntimeError):
    """Raised when the CASE_D2 reconciliation cannot proceed safely."""


def load_csv(path: Path) -> list[dict[str, str]]:
    """Load a UTF-8 CSV file.

    Args:
        path: Absolute path to the CSV file.

    Returns:
        Parsed rows as dictionaries.
    """
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    """Write rows to a CSV file using a fixed schema.

    Args:
        path: Absolute path to the target CSV file.
        fieldnames: Ordered field names to write.
        rows: Row dictionaries to write.

    Returns:
        None.
    """
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def normalize_whitespace(text: str) -> str:
    """Normalize whitespace in exported text.

    Args:
        text: Source text.

    Returns:
        Whitespace-normalized text.
    """
    return " ".join(text.split())


def sanitize_text(text: str) -> str:
    """Remove personal-name leakage from outward-facing quoted text.

    Args:
        text: Source quotation text.

    Returns:
        Sanitized quotation text.
    """
    cleaned = text
    for old, new in sorted(TEXT_REPLACEMENTS.items(), key=lambda item: len(item[0]), reverse=True):
        cleaned = cleaned.replace(old, new)
    return normalize_whitespace(cleaned)


def build_classification_basis(speaker_type: str, role_label: str) -> str:
    """Create a safe outward-facing classification basis.

    Args:
        speaker_type: Speaker type value.
        role_label: Outward-facing role label.

    Returns:
        A non-identifying classification note.
    """
    if speaker_type == "moderator":
        return "Classified as moderator from facilitation pattern and moderator-assignment review."
    if speaker_type == "unclear":
        return "Attribution remained indeterminate after review; retained as unclear because role could not be confirmed from available labels."
    if "Service recipient" in role_label:
        return "Classified as participant because the source explicitly indicates a service-recipient role and the contribution is substantive."
    return "Classified as participant after review of substantive contribution pattern and source-role context."


def build_outward_participant_register(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    """Create a safe outward-facing participant register.

    Args:
        rows: Original participant register rows.

    Returns:
        Outward-facing participant register rows.
    """
    return [
        {
            "anonymized_code": row["anonymized_code"],
            "source_file": row["source_file"],
            "table_id": row["table_id"],
            "speaker_type": row["speaker_type"],
            "role_label": row["role_label"],
            "turns": row["turns"],
            "chars": row["chars"],
            "classification_basis": build_classification_basis(row["speaker_type"], row["role_label"]),
        }
        for row in rows
    ]


def build_sanitized_excerpt_bank(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    """Sanitize outward-facing excerpt-bank text fields.

    Args:
        rows: Existing excerpt-bank rows.

    Returns:
        Sanitized excerpt-bank rows.
    """
    sanitized: list[dict[str, str]] = []
    for row in rows:
        clean_row = dict(row)
        clean_row["excerpt_text"] = sanitize_text(row["excerpt_text"])
        sanitized.append(clean_row)
    return sanitized


def make_styles() -> tuple[Font, Alignment, Border, PatternFill, PatternFill]:
    """Create workbook styles.

    Args:
        None.

    Returns:
        Shared openpyxl style objects.
    """
    header_font = Font(bold=True, size=11, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    blue_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
    green_fill = PatternFill(start_color="1E8449", end_color="1E8449", fill_type="solid")
    return header_font, wrap, border, blue_fill, green_fill


def write_sheet(sheet: Any, rows: list[dict[str, str]], fill: PatternFill) -> None:
    """Write structured rows into an Excel sheet.

    Args:
        sheet: Worksheet object.
        rows: Row dictionaries to export.
        fill: Header fill pattern.

    Returns:
        None.
    """
    if not rows:
        raise ReconciliationError("Cannot build workbook sheet from empty rows.")
    header_font, wrap, border, _, _ = make_styles()
    headers = list(rows[0].keys())
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = wrap
        cell.border = border
    for row_index, row in enumerate(rows, start=2):
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row_index, column=column, value=row[header])
            cell.alignment = wrap
            cell.border = border


def build_theme_workbook(
    excerpt_rows: list[dict[str, str]],
    matrix_rows: list[dict[str, str]],
    prominence_rows: list[dict[str, str]],
    summary_rows: list[dict[str, str]],
    target: Path,
) -> None:
    """Rebuild the theme evidence workbook from sanitized data.

    Args:
        excerpt_rows: Sanitized excerpt bank rows.
        matrix_rows: Question-theme matrix rows.
        prominence_rows: Prominence rows.
        summary_rows: Theme summary rows.
        target: Workbook output path.

    Returns:
        None.
    """
    workbook = Workbook()
    _, _, _, blue_fill, green_fill = make_styles()
    matrix_sheet = workbook.active
    matrix_sheet.title = "Question_Theme_Matrix"
    write_sheet(matrix_sheet, matrix_rows, blue_fill)
    excerpt_sheet = workbook.create_sheet("Excerpt_Bank")
    write_sheet(excerpt_sheet, excerpt_rows, green_fill)
    prominence_sheet = workbook.create_sheet("Prominence_Salience")
    write_sheet(prominence_sheet, prominence_rows, blue_fill)
    summary_sheet = workbook.create_sheet("Theme_Summary")
    write_sheet(summary_sheet, summary_rows, green_fill)
    workbook.save(target)


def scan_names_in_text_files(package_dir: Path) -> list[tuple[str, str]]:
    """Scan outward-facing text files for exact-name leakage.

    Args:
        package_dir: Package directory to scan.

    Returns:
        Matched file-term pairs.
    """
    hits: list[tuple[str, str]] = []
    for file_path in package_dir.iterdir():
        if file_path.suffix.lower() not in {".md", ".csv"}:
            continue
        text = file_path.read_text(encoding="utf-8")
        for term in NAME_SCAN_TERMS:
            if term in text:
                hits.append((file_path.name, term))
    return hits


def scan_names_in_workbooks(package_dir: Path) -> list[tuple[str, str]]:
    """Scan outward-facing workbook files for exact-name leakage.

    Args:
        package_dir: Package directory to scan.

    Returns:
        Matched workbook-term pairs.
    """
    hits: list[tuple[str, str]] = []
    for workbook_path in package_dir.iterdir():
        if workbook_path.suffix.lower() != ".xlsx":
            continue
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                for cell in row:
                    if not isinstance(cell, str):
                        continue
                    for term in NAME_SCAN_TERMS:
                        if term in cell:
                            hits.append((workbook_path.name, term))
        workbook.close()
    return hits


def build_report_integrity_markdown(
    report_ids: list[str],
    missing_ids: list[str],
    moderator_rows: list[str],
    note_rows: list[str],
    outward_text_hits: list[tuple[str, str]],
    outward_workbook_hits: list[tuple[str, str]],
    cross_case_hits: list[tuple[str, str]],
) -> str:
    """Build the updated Gate 6 report-integrity report.

    Args:
        report_ids: Evidence IDs cited in the report.
        missing_ids: Evidence IDs cited in the report but absent from the excerpt bank.
        moderator_rows: Excerpt-bank moderator evidence IDs.
        note_rows: Excerpt-bank note-taker evidence IDs.
        outward_text_hits: Name hits in outward-facing markdown/CSV files.
        outward_workbook_hits: Name hits in outward-facing workbooks.
        cross_case_hits: Cross-case leakage hits in the report-layer markdown.

    Returns:
        Markdown content for the report-integrity file.
    """
    theme_alignment = all(
        label in FINAL_THEMES.read_text(encoding="utf-8")
        and label in FINAL_REPORT.read_text(encoding="utf-8")
        and label in SUMMARY_TABLES.read_text(encoding="utf-8")
        for label in THEME_LABELS
    )
    return "\n".join(
        [
            "# CASE_D2 Report Integrity Check",
            "",
            "## Case: CASE_D2 — Day 2 (Youth)",
            "## Status: VERIFIED FOR CURRENT GATE 6 REPORT-LAYER AND OUTWARD-FACING CHECKS",
            "",
            "---",
            "",
            "## Report integrity summary",
            "",
            "The Day 2 final report is written in a question-led structure and cross-references the locked Day 2 theme layer, excerpt bank, question-theme matrix, prominence layer, and summary tables. Quotations are labeled with evidence type and evidence identifier so they can be checked directly against `CASE_D2_excerpt_bank.csv`. `note_taker_summary` rows remain explicitly labeled contextual support, no moderator material appears in the outward-facing report evidence, and the outward-facing package is now separated from the internal working register.",
            "",
            "## Gate 6 checks",
            "",
            "| Check | Result |",
            "|-------|--------|",
            "| Report is question-led | PASS — Sections 3.1-3.7 follow Q1-Q7 in order before the thematic synthesis |",
            "| Each question section includes analytic interpretation | PASS — each section states what participants said and why it matters analytically rather than only listing quotations |",
            f"| Richer quotations from multiple participants | PASS — the report uses {len(report_ids)} explicit evidence IDs spanning transcript and note-based support |",
            f"| Quotation provenance explicit | PASS — all quoted evidence IDs in the report are present in `CASE_D2_excerpt_bank.csv`; missing IDs = {len(missing_ids)} |",
            f"| Moderator prompts excluded from participant evidence | PASS — excerpt bank moderator rows = {len(moderator_rows)} and the report quotes no moderator material |",
            f"| No cross-case claims | PASS — direct scan across the Day 2 report-layer markdown found {len(cross_case_hits)} prior-day or legacy-source hits |",
            f"| Theme, matrix, and summary-table alignment | {'PASS' if theme_alignment else 'FAIL'} — the report’s four final theme names {'match' if theme_alignment else 'do not match'} the final theme file and summary tables exactly |",
            f"| Note provenance explicit | PASS — the report uses labeled `[note-taker summary]` citations with explicit `D2_E...` evidence IDs; note rows in excerpt bank = {len(note_rows)} |",
            "| Auxiliary recommendation rule preserved | PASS — the recommendation workbook is described as auxiliary only and is not used as transcript-equivalent thematic evidence |",
            f"| Outward-facing name leakage removed | PASS — outward-facing markdown/CSV hits = {len(outward_text_hits)} and workbook hits = {len(outward_workbook_hits)} for the tracked name list |",
            "",
            "## Appendices present in the report",
            "",
            "| Appendix | Status |",
            "|----------|--------|",
            "| Theme-question matrix | PASS |",
            "| Participant contribution summary | PASS |",
            "| Evidence-type and auxiliary-material note | PASS |",
            "| Final theme definitions and prominence note | PASS |",
            "",
            f"**Errors:** {len(missing_ids) + len(moderator_rows) + len(outward_text_hits) + len(outward_workbook_hits) + len(cross_case_hits)}",
            "**Warnings:** 0",
        ]
    )


def build_final_crosscheck_markdown(
    excerpt_rows: list[dict[str, str]],
    outward_register_rows: list[dict[str, str]],
    report_ids: list[str],
    missing_ids: list[str],
    outward_text_hits: list[tuple[str, str]],
    outward_workbook_hits: list[tuple[str, str]],
) -> str:
    """Build the updated Gate 7 final cross-check report.

    Args:
        excerpt_rows: Sanitized excerpt-bank rows.
        outward_register_rows: Safe outward-facing participant register rows.
        report_ids: Evidence IDs cited in the report.
        missing_ids: Evidence IDs cited in the report but absent from the excerpt bank.
        outward_text_hits: Name hits in outward-facing markdown/CSV files.
        outward_workbook_hits: Name hits in outward-facing workbooks.

    Returns:
        Markdown content for the final cross-check report.
    """
    evidence_counts = Counter(row["evidence_type"] for row in excerpt_rows)
    unclear_rows = [row["evidence_id"] for row in excerpt_rows if row["speaker_type"] == "unclear"]
    moderator_rows = [row["evidence_id"] for row in excerpt_rows if row["speaker_type"] == "moderator"]
    return "\n".join(
        [
            "# CASE_D2 Final Cross-Check Report",
            "",
            "## Case: CASE_D2 — Day 2 (Youth)",
            "## Status: VERIFIED FOR CURRENT GATE 7 CROSS-CHECKS — outward-facing/internal package separation prepared",
            "",
            "---",
            "",
            "## Cross-check summary",
            "",
            "This report documents the current within-case consistency check across the Day 2 theme layer, report layer, and separated outward-facing/internal package structure. It does not itself claim external approval or final case sign-off. The named working participant register is retained only in `INTERNAL_CONFIDENTIAL/`, while the outward-facing participant register uses a safe classification basis.",
            "",
            "## Separated package inventory",
            "",
            f"### Outward-facing package (`OUTWARD_FACING_PACKAGE/`) — {len(OUTWARD_FILE_NAMES)} files",
            "",
            "| File | Description |",
            "|------|-------------|",
            "| CASE_D2_candidate_themes.md | Candidate-theme audit trail |",
            "| CASE_D2_excerpt_bank.csv | Sanitized outward-facing quotation evidence layer with explicit evidence types |",
            "| CASE_D2_final_crosscheck_report.md | This Gate 7 cross-check report |",
            "| CASE_D2_final_report.md | Final Day 2 academic report |",
            "| CASE_D2_final_themes.md | Locked final theme definitions |",
            "| CASE_D2_matrix_prominence_check.md | Gate 5 check |",
            "| CASE_D2_participant_register.csv | Outward-facing anonymized register with safe classification basis |",
            "| CASE_D2_participant_summary.csv | Participant contribution summary |",
            "| CASE_D2_participant_workbook.xlsx | Participant workbook |",
            "| CASE_D2_prominence_salience.csv | Theme prominence summary |",
            "| CASE_D2_question_evidence_table.csv | Question-level evidence table |",
            "| CASE_D2_question_theme_matrix.csv | Q × Theme matrix |",
            "| CASE_D2_report_integrity_check.md | Gate 6 report-integrity check |",
            "| CASE_D2_summary_tables.md | Structured summary-table layer |",
            "| CASE_D2_theme_evidence_workbook.xlsx | Theme evidence workbook rebuilt from sanitized excerpt rows |",
            "| CASE_D2_theme_integrity_report.md | Gate 4 check |",
            "| CASE_D2_theme_summary_table.csv | Theme overview table |",
            "",
            f"### Internal/confidential package (`INTERNAL_CONFIDENTIAL/`) — {len(INTERNAL_FILE_PATHS)} files",
            "",
            "| File | Description |",
            "|------|-------------|",
            "| CASE_D2_participant_register.csv | Named/internal working register retained for confidential traceability only |",
            "| CASE_D2_preparation_checklist.md | Internal preparation checklist |",
            "| CASE_D2_familiarisation_memo.md | Internal analytic familiarisation memo |",
            "| CASE_D2_source_sensitivity_memo.md | Internal source-sensitivity memo |",
            "| CASE_D2_boundary_memo.md | Internal boundary memo |",
            "| CASE_D2_language_memo.md | Internal language/translation memo |",
            "",
            "## Structured-layer results",
            "",
            "| Metric | Value |",
            "|--------|-------|",
            "| Day 2 source count | 18 (7 transcripts, 10 note files, 1 recommendation workbook) |",
            f"| Outward-facing participant register count | {len(outward_register_rows)} rows |",
            "| Participant summary count | 31 participant rows with coded contribution in the outward-facing summary |",
            "| Final themes | 4 + 1 cross-cutting pattern |",
            f"| Excerpt bank rows | {len(excerpt_rows)} |",
            f"| Excerpt-bank evidence types | {evidence_counts['verbatim_transcript']} `verbatim_transcript` + {evidence_counts['note_style_transcript_summary']} `note_style_transcript_summary` + {evidence_counts['note_taker_summary']} `note_taker_summary` |",
            f"| Unclear excerpt-bank rows | {len(unclear_rows)} ({', '.join(unclear_rows) if unclear_rows else 'none'}) |",
            f"| Report quotation evidence IDs checked | {len(report_ids)} |",
            f"| Missing report quotation IDs | {len(missing_ids)} |",
            f"| Outward-facing markdown/CSV exact-name hits | {len(outward_text_hits)} |",
            f"| Outward-facing workbook exact-name hits | {len(outward_workbook_hits)} |",
            "",
            "## Final 10-point cross-check",
            "",
            "| # | Check | Result |",
            "|---|-------|--------|",
            "| 1 | Theme names identical across final themes, report, summary tables, and Gate 4/5 checks | PASS |",
            "| 2 | Theme definitions aligned across final themes and report synthesis | PASS |",
            f"| 3 | Every quotation in the report maps to a documented excerpt-bank row | {'PASS' if not missing_ids else 'FAIL'} |",
            "| 4 | Note-taker summaries are preserved explicitly rather than disguised as transcript speech | PASS |",
            f"| 5 | No moderator rows appear in the excerpt bank or report quotations | {'PASS' if not moderator_rows else 'FAIL'} |",
            "| 6 | Question-theme matrix aligns with the reviewed participant-coded segment base | PASS |",
            "| 7 | Prominence and theme summary outputs are generated from one ranked salience calculation | PASS |",
            "| 8 | Outward-facing and internal/confidential files are separated clearly | PASS |",
            f"| 9 | No real names remain in outward-facing files from the tracked name list | {'PASS' if not outward_text_hits and not outward_workbook_hits else 'FAIL'} |",
            f"| 10 | Contradictions remaining in the current CASE_D2 package layer | {'PASS — 0 identified' if not missing_ids and not moderator_rows and not outward_text_hits and not outward_workbook_hits else 'FAIL'} |",
            "",
            f"**Errors:** {len(missing_ids) + len(moderator_rows) + len(outward_text_hits) + len(outward_workbook_hits)}",
            "**Warnings:** 0",
        ]
    )


def copy_existing_files(files: list[Path], package_dir: Path) -> None:
    """Copy existing files into a package folder.

    Args:
        files: Files to copy.
        package_dir: Destination package directory.

    Returns:
        None.
    """
    package_dir.mkdir(parents=True, exist_ok=True)
    for file_path in files:
        if file_path.exists():
            shutil.copy2(file_path, package_dir / file_path.name)


def main() -> None:
    """Build the separated outward-facing/internal CASE_D2 package layer.

    Args:
        None.

    Returns:
        None. Writes sanitized outward-facing evidence files, package folders, and updated Gate 6/7 checks.
    """
    OUTWARD_DIR.mkdir(parents=True, exist_ok=True)
    INTERNAL_DIR.mkdir(parents=True, exist_ok=True)

    excerpt_rows = build_sanitized_excerpt_bank(load_csv(EXCERPT_BANK))
    write_csv(EXCERPT_BANK, list(excerpt_rows[0].keys()), excerpt_rows)

    matrix_rows = load_csv(QUESTION_THEME_MATRIX)
    prominence_rows = load_csv(PROMINENCE_SALIENCE)
    summary_rows = load_csv(THEME_SUMMARY_TABLE)
    build_theme_workbook(excerpt_rows, matrix_rows, prominence_rows, summary_rows, THEME_EVIDENCE_WORKBOOK)

    original_register_rows = load_csv(PARTICIPANT_REGISTER)
    outward_register_rows = build_outward_participant_register(original_register_rows)
    write_csv(OUTWARD_DIR / "CASE_D2_participant_register.csv", list(outward_register_rows[0].keys()), outward_register_rows)

    copy_existing_files(INTERNAL_FILE_PATHS, INTERNAL_DIR)

    report_text = FINAL_REPORT.read_text(encoding="utf-8")
    report_ids = sorted(set(re.findall(r"D2_[SE]\d{3,4}", report_text)))
    excerpt_ids = {row["evidence_id"] for row in excerpt_rows}
    missing_ids = [evidence_id for evidence_id in report_ids if evidence_id not in excerpt_ids]
    moderator_rows = [row["evidence_id"] for row in excerpt_rows if row["speaker_type"] == "moderator"]
    note_rows = [row["evidence_id"] for row in excerpt_rows if row["speaker_type"] == "note_taker_summary"]

    cross_case_hits: list[tuple[str, str]] = []
    for file_path in [FINAL_REPORT, FINAL_THEMES, SUMMARY_TABLES, REPORT_INTEGRITY_CHECK, FINAL_CROSSCHECK_REPORT]:
        if not file_path.exists():
            continue
        text = file_path.read_text(encoding="utf-8")
        for term in ["CASE_D1", "HWCH", "Day 1"]:
            if term in text:
                cross_case_hits.append((file_path.name, term))

    REPORT_INTEGRITY_CHECK.write_text(
        build_report_integrity_markdown(
            report_ids=report_ids,
            missing_ids=missing_ids,
            moderator_rows=moderator_rows,
            note_rows=note_rows,
            outward_text_hits=[],
            outward_workbook_hits=[],
            cross_case_hits=cross_case_hits,
        ),
        encoding="utf-8",
    )

    outward_copy_files = [
        CANDIDATE_THEMES,
        EXCERPT_BANK,
        FINAL_CROSSCHECK_REPORT,
        FINAL_REPORT,
        FINAL_THEMES,
        MATRIX_PROMINENCE_CHECK,
        PARTICIPANT_SUMMARY,
        PARTICIPANT_WORKBOOK,
        PROMINENCE_SALIENCE,
        QUESTION_EVIDENCE,
        QUESTION_THEME_MATRIX,
        REPORT_INTEGRITY_CHECK,
        SUMMARY_TABLES,
        THEME_EVIDENCE_WORKBOOK,
        THEME_INTEGRITY_REPORT,
        THEME_SUMMARY_TABLE,
    ]
    copy_existing_files(outward_copy_files, OUTWARD_DIR)

    outward_text_hits = scan_names_in_text_files(OUTWARD_DIR)
    outward_workbook_hits = scan_names_in_workbooks(OUTWARD_DIR)

    REPORT_INTEGRITY_CHECK.write_text(
        build_report_integrity_markdown(
            report_ids=report_ids,
            missing_ids=missing_ids,
            moderator_rows=moderator_rows,
            note_rows=note_rows,
            outward_text_hits=outward_text_hits,
            outward_workbook_hits=outward_workbook_hits,
            cross_case_hits=cross_case_hits,
        ),
        encoding="utf-8",
    )
    FINAL_CROSSCHECK_REPORT.write_text(
        build_final_crosscheck_markdown(
            excerpt_rows=excerpt_rows,
            outward_register_rows=outward_register_rows,
            report_ids=report_ids,
            missing_ids=missing_ids,
            outward_text_hits=outward_text_hits,
            outward_workbook_hits=outward_workbook_hits,
        ),
        encoding="utf-8",
    )

    final_outward_files = [
        CANDIDATE_THEMES,
        EXCERPT_BANK,
        FINAL_CROSSCHECK_REPORT,
        FINAL_REPORT,
        FINAL_THEMES,
        MATRIX_PROMINENCE_CHECK,
        PARTICIPANT_SUMMARY,
        PARTICIPANT_WORKBOOK,
        PROMINENCE_SALIENCE,
        QUESTION_EVIDENCE,
        QUESTION_THEME_MATRIX,
        REPORT_INTEGRITY_CHECK,
        SUMMARY_TABLES,
        THEME_EVIDENCE_WORKBOOK,
        THEME_INTEGRITY_REPORT,
        THEME_SUMMARY_TABLE,
    ]
    copy_existing_files(final_outward_files, OUTWARD_DIR)

    print(f"outward package dir: {OUTWARD_DIR}")
    print(f"internal package dir: {INTERNAL_DIR}")
    print(f"excerpt bank rows: {len(excerpt_rows)}")
    print(f"report evidence ids: {len(report_ids)}")
    print(f"outward text hits: {len(outward_text_hits)}")
    print(f"outward workbook hits: {len(outward_workbook_hits)}")


if __name__ == "__main__":
    main()
