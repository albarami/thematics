from __future__ import annotations

import csv
import shutil
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(r"c:\Users\baram\OneDrive\Desktop\themnatic")
CASE_DIR = ROOT / "analysis" / "CASE_D2"
WORKING_DIR = CASE_DIR / "_working"
CANDIDATE_SEGMENTS = WORKING_DIR / "CASE_D2_segment_candidates.csv"
CODED_SEGMENTS = CASE_DIR / "CASE_D2_coded_segments.csv"
PARTICIPANT_SUMMARY = CASE_DIR / "CASE_D2_participant_summary.csv"
PARTICIPANT_WORKBOOK = CASE_DIR / "CASE_D2_participant_workbook.xlsx"
QUESTION_EVIDENCE = CASE_DIR / "CASE_D2_question_evidence_table.csv"
QUESTION_ORDER = ["Q1", "Q2", "Q3", "Q4", "Q5", "Q6", "Q7"]
TOP_CODE_LIMIT = 5
TOP_CODE_EXCLUSIONS = {
    "general_response",
    "moderator_context",
    "professional_identity",
    "service_recipient_voice",
    "student_youth_voice",
    "pillar_spiritual",
    "pillar_emotional",
    "pillar_intellectual",
    "pillar_physical",
    "pillar_social",
}
CODE_ORDER = [
    "contentment_acceptance",
    "balance_multidimensional",
    "spiritual_moral_anchor",
    "inner_peace_stability",
    "support_as_wellbeing_core",
    "functionality_self_management",
    "family_support_ecology",
    "peer_belonging_isolation",
    "youth_transition_pressure",
    "digital_social_media_pressure",
    "youth_voice_visibility",
    "intergenerational_disconnect",
    "consultation_time_pressure",
    "quantity_over_quality",
    "stigma_help_seeking_barrier",
    "confidentiality_trust_barrier",
    "language_cultural_mismatch",
    "fragmented_service_pathway",
    "access_delay_waiting",
    "environment_design_stress",
    "communication_training_need",
    "interdisciplinary_integration",
    "awareness_outreach",
    "youth_specific_service_design",
    "family_inclusive_care",
    "healthcare_worker_wellbeing",
    "environmental_design_improvement",
    "practical_recommendation",
    "top_down_system_change",
    "whole_person_care",
    "empathy_reassurance",
    "patient_provider_relationship",
    "self_disclosure_hidden_distress",
    "service_recipient_comparison",
    "pillar_spiritual",
    "pillar_emotional",
    "pillar_intellectual",
    "pillar_physical",
    "pillar_social",
    "professional_identity",
    "service_recipient_voice",
    "student_youth_voice",
    "moderator_context",
    "general_response",
]
ROLE_CODE_PATTERNS = {
    "professional_identity": [
        "consultant",
        "executive",
        "psychiatrist",
        "adviser",
        "specialist",
        "psychologist",
        "clinician",
        "academic",
        "nurse",
        "midwife",
        "therapist",
        "leader",
        "طبيب",
        "دكتور",
        "استشاري",
        "أستاذ",
        "أخصائي",
        "ممرضة",
        "مستشار",
    ],
    "service_recipient_voice": ["service recipient", "beneficiary", "patient", "مستفيد", "متلقي خدمة", "مراجع"],
    "student_youth_voice": ["student", "طالبة", "طالب", "university youth"],
}
TEXT_CODE_PATTERNS = {
    "contentment_acceptance": ["قناعة", "رضا", "راضي", "contentment", "satisfied", "okay with what you are", "contentment with", "at peace"],
    "balance_multidimensional": ["توازن", "balance", "equilibrium", "holistic", "شمول", "ركائز", "عدة أقسام", "multidimensional", "متشابكة", "all aspects"],
    "spiritual_moral_anchor": ["الدين", "الإيمان", "spiritual", "faith", "prayer", "الروحية", "القضاء والقدر", "القرآن", "رب العالمين", "اللهم"],
    "inner_peace_stability": ["طمأنينة", "peace of mind", "peace", "calm", "steady", "stable", "stability", "استقرار", "سلام داخلي", "راحة البال"],
    "support_as_wellbeing_core": ["support system", "people around you", "connected to the people", "connections", "relationships", "support", "دعم", "علاقات", "حولك", "أناس طيبين"],
    "functionality_self_management": ["function", "cope", "coping", "regulate", "get up in the morning", "able to do", "daily life", "daily activities", "maintain", "depend on yourself", "productive", "don't give up", "drive to continue", "قادر يشتغل", "قابل للعمل", "يعتمد على نفسه", "يهتم بنفسه", "الاستمرارية"],
    "family_support_ecology": ["family", "parents", "home", "household", "أهل", "الوالدين", "الأسرة", "البيت", "زوج", "caregiver"],
    "peer_belonging_isolation": ["friends", "friendship", "belong", "alone", "lonely", "isolat", "العزلة", "منعزل", "أصدقاء", "social circle"],
    "youth_transition_pressure": ["exam", "study pressure", "performance", "transition", "adolesc", "اختبارات", "امتحان", "ضغط الدراسة", "مرحلة انتقالية"],
    "digital_social_media_pressure": ["social media", "online", "phones", "phone", "technology", "التواصل الاجتماعي", "هاتف", "آيباد", "شاشات", "إلكترونيات"],
    "youth_voice_visibility": ["كمتلقي", "كشاب", "youth perspective", "student voice", "service recipient", "من منظور الشباب", "as a service recipient"],
    "intergenerational_disconnect": ["generation gap", "weak communication", "between young people and adults", "no one listens", "التواصل بين الأجيال", "لا يسمع", "family distance"],
    "consultation_time_pressure": ["30 minutes", "20 minutes", "20 minute", "5 minutes", "rushed", "no time", "limited consultation", "ضيق الوقت", "appointment time", "spend enough time", "30 دقيقة", "نص ساعة", "5 دقايق", "5 دقائق"],
    "quantity_over_quality": ["quantity", "numbers", "targets", "kpi", "throughput", "more patients", "أعداد", "كمية", "إنتاجية"],
    "stigma_help_seeking_barrier": ["stigma", "shame", "embarrass", "taboo", "نظرة مجتمعية", "وصمة", "حرج", "mental health taboo"],
    "confidentiality_trust_barrier": ["confidential", "privacy", "trust", "سرية", "خصوصية", "ملف", "اطلع", "إفشاء", "exposure", "ثقتها"],
    "language_cultural_mismatch": ["language", "culture", "أردو", "فرنساوي", "هندوسي", "ثقاف", "english", "arabic"],
    "fragmented_service_pathway": ["referral", "تحويل", "disconnected", "fragmented", "silo", "pathway", "coordination", "navigate"],
    "access_delay_waiting": ["waiting", "wait", "queue", "delayed", "delay", "appointment delays", "two months", "one-year", "backlog", "انتظار", "تأخير", "طابور", "شهرين", "ثلاث شهور", "3 شهور", "ساعة ونص"],
    "environment_design_stress": ["waiting room", "environment", "design", "colors", "music", "موسيقى", "ألوان", "clinic design"],
    "communication_training_need": ["communication training", "staff training", "training", "تدريب", "skills training"],
    "interdisciplinary_integration": ["interdisciplinary", "collaboration", "collaborat", "team", "integration", "coordinat", "across specialties", "بين التخصصات"],
    "awareness_outreach": ["awareness", "campaign", "outreach", "emails", "توعية", "increase understanding"],
    "family_inclusive_care": ["involve family", "parents in care", "caregiver role", "family accompaniment", "caregiver involvement", "أولياء الأمور", "الوالدين في العلاج", "دمج أولياء الأمور"],
    "healthcare_worker_wellbeing": ["staff wellbeing", "staff well-being", "burnout", "employee", "provider support", "workforce", "for staff", "healthcare workers", "احتراق", "رفاه العاملين"],
    "top_down_system_change": ["leadership", "management", "ministry", "policy", "policies", "admin", "law", "decision makers", "mandatory", "required", "requirement", "صناع القرار", "قانون", "قرار", "مؤسسية", "الإدارة العليا"],
    "whole_person_care": ["whole person", "holistic care", "not just the disease", "full person", "شمولية", "ليس فقط"],
    "empathy_reassurance": ["reassure", "calm", "هدوء", "لمسة إنسانية", "طمأنة", "احتواء", "empathy"],
    "patient_provider_relationship": ["rapport", "trust with doctor", "communication with patient", "patient-doctor", "doctor helped", "تواصل مع المريض", "ثقة المريض"],
    "self_disclosure_hidden_distress": ["hidden", "not obvious", "surface", "from outside", "physical symptoms", "physical problem", "don't know how to express", "منغلقة", "كالعلبة المغلقة", "لا يفصح", "ما يبين", "الشق الداخلي", "داخليا", "hide", "looks fine"],
    "service_recipient_comparison": ["another doctor", "one doctor", "comparison of encounters", "good and bad", "مقارنة"],
    "pillar_spiritual": ["ركيزة روحية", "روحية", "spiritual pillar", "الروحية"],
    "pillar_emotional": ["ركيزة عاطفية", "عاطفية", "emotional pillar", "العاطفية"],
    "pillar_intellectual": ["ركيزة فكرية", "فكرية", "intellectual pillar", "الفكرية"],
    "pillar_physical": ["ركيزة بدنية", "بدنية", "جسدية", "physical pillar", "البدنية"],
    "pillar_social": ["ركيزة اجتماعية", "اجتماعية", "social pillar", "الاجتماعية"],
}
SEMANTIC_CODES = {
    code
    for code in CODE_ORDER
    if code
    not in {
        "pillar_spiritual",
        "pillar_emotional",
        "pillar_intellectual",
        "pillar_physical",
        "pillar_social",
        "professional_identity",
        "service_recipient_voice",
        "student_youth_voice",
        "moderator_context",
        "general_response",
    }
}
PROPOSAL_MARKERS = ["suggest", "recommend", "proposal", "should", "need to", "أقترح", "اقتراح", "ينبغي", "يجب", "مطلوب", "solution"]
YOUTH_SERVICE_MARKERS = ["youth", "adolescent", "adolescence", "student", "الشباب", "مراهق"]
SERVICE_DESIGN_MARKERS = ["clinic", "service", "pathway", "space", "center", "عيادة", "خدمة", "مسار", "مركز"]


def _load_rows(csv_path: Path) -> list[dict[str, str]]:
    with csv_path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _question_sort_key(question_id: str) -> tuple[int, str]:
    try:
        return (QUESTION_ORDER.index(question_id), question_id)
    except ValueError:
        return (len(QUESTION_ORDER), question_id)


def _normalise_codes(raw_codes: str) -> list[str]:
    return [code.strip() for code in raw_codes.split(";") if code.strip()]


def _format_top_codes(counter: Counter[str]) -> str:
    filtered_counter = Counter({code: count for code, count in counter.items() if code not in TOP_CODE_EXCLUSIONS})
    ranked_source = filtered_counter if filtered_counter else counter
    ranked = sorted(ranked_source.items(), key=lambda item: (-item[1], item[0]))
    return ";".join(code for code, _ in ranked[:TOP_CODE_LIMIT])


def _contains_any(text: str, markers: list[str]) -> bool:
    return any(marker in text for marker in markers)


def _ordered_codes(codes: set[str]) -> list[str]:
    order_map = {code: index for index, code in enumerate(CODE_ORDER)}
    return sorted(codes, key=lambda code: (order_map.get(code, len(order_map)), code))


def _derive_role_codes(role_label: str) -> set[str]:
    lowered_role = role_label.lower()
    derived_codes: set[str] = set()
    for code, markers in ROLE_CODE_PATTERNS.items():
        if _contains_any(lowered_role, markers):
            derived_codes.add(code)
    return derived_codes


def _derive_text_codes(row: dict[str, str]) -> set[str]:
    lowered_text = row["segment_text"].lower()
    lowered_role = row["role_label"].lower()
    derived_codes: set[str] = set()
    for code, markers in TEXT_CODE_PATTERNS.items():
        if _contains_any(lowered_text, markers):
            derived_codes.add(code)
    if _contains_any(lowered_text, YOUTH_SERVICE_MARKERS) and _contains_any(lowered_text, SERVICE_DESIGN_MARKERS):
        derived_codes.add("youth_specific_service_design")
    if row["question_id"] == "Q7" and _contains_any(lowered_text, PROPOSAL_MARKERS):
        derived_codes.add("practical_recommendation")
    if row["question_id"] in {"Q6", "Q7"} and _contains_any(lowered_text, ["school nurse", "school nurses", "counsellor", "counselor", "social workers"]) and _contains_any(lowered_text, ["school", "education", "healthcare", "combining", "combine", "collaborat", "together"]):
        derived_codes.add("interdisciplinary_integration")
    if row["question_id"] in {"Q4", "Q5", "Q6"} and _contains_any(lowered_text, ["refer", "referral", "يحول", "تحويل"]):
        if _contains_any(lowered_text, ["department", "clinic", "عيادة", "قسم", "فريق", "فرق"]):
            derived_codes.add("interdisciplinary_integration")
    if row["question_id"] in {"Q6", "Q7"} and _contains_any(lowered_text, ["trained more", "trained", "training", "teach", "teaching", "education"]):
        if _contains_any(lowered_text, ["school nurse", "school nurses", "parents", "mother", "kids", "children"]):
            derived_codes.add("communication_training_need")
    if row["question_id"] in {"Q6", "Q7"} and "environment" in lowered_text and _contains_any(lowered_text, PROPOSAL_MARKERS):
        derived_codes.add("environmental_design_improvement")
    if row["question_id"] in {"Q6", "Q7"} and "design" in lowered_text and _contains_any(lowered_text, PROPOSAL_MARKERS):
        derived_codes.add("environmental_design_improvement")
    if row["question_id"] in {"Q6", "Q7"} and _contains_any(lowered_text, ["ألوان", "موسيقى", "space", "waiting room"]):
        derived_codes.add("environmental_design_improvement")
    if ("مريض" in lowered_text or "patient" in lowered_text) and _contains_any(lowered_text, ["ثقة", "تواصل", "rapport", "doctor", "طبيب", "communication with patient", "trust with doctor", "patient-doctor"]):
        derived_codes.add("patient_provider_relationship")
    if "privacy" in lowered_text or "خصوصية" in lowered_text or "confidential" in lowered_text:
        derived_codes.add("confidentiality_trust_barrier")
    if ("mental health" in lowered_text or "الصحة النفسية" in lowered_text) and _contains_any(lowered_text, ["حرج", "وصمة", "نظرة", "taboo", "stigma", "resistance", "رفض"]):
        derived_codes.add("stigma_help_seeking_barrier")
    if ("service recipient" in lowered_role or "متلقي خدمة" in lowered_role or "student" in lowered_role or "طالبة" in lowered_role) and _contains_any(lowered_text, ["كمتلقي", "كشاب", "my perspective", "من منظوري"]):
        derived_codes.add("youth_voice_visibility")
    if row["question_id"] in {"Q6", "Q7"} and _contains_any(lowered_text, ["family", "parents", "caregiver", "الأسرة", "الوالدين"]) and _contains_any(lowered_text, PROPOSAL_MARKERS):
        derived_codes.add("family_inclusive_care")
    if row["question_id"] in {"Q6", "Q7"} and _contains_any(lowered_text, ["أولياء الأمور", "الوالدين", "parents"]):
        if _contains_any(lowered_text, ["العلاج", "treatment", "care"]):
            derived_codes.add("family_inclusive_care")
    if row["question_id"] in {"Q6", "Q7"} and _contains_any(lowered_text, ["teach the parents", "teaching the parents", "parenting the parents", "parents", "mother"]):
        if _contains_any(lowered_text, ["teach", "teaching", "education", "opportunity"]):
            derived_codes.add("family_inclusive_care")
    if row["question_id"] in {"Q6", "Q7"} and _contains_any(lowered_text, ["adding a counsellor", "adding a counselor", "requirement for every centre", "requirement for every center", "every centre", "every center"]):
        derived_codes.add("practical_recommendation")
        derived_codes.add("top_down_system_change")
    if row["question_id"] in {"Q4", "Q5", "Q6", "Q7"} and _contains_any(lowered_text, ["psychologist", "psychiatrist", "break rooms", "gym", "gyms"]):
        if _contains_any(lowered_text, ["for staff", "workers", "employee", "staff"]):
            derived_codes.add("healthcare_worker_wellbeing")
    if row["question_id"] in {"Q4", "Q5", "Q6", "Q7"} and _contains_any(lowered_text, ["recommend", "should have", "dedicated", "mandatory", "requirement"]):
        if _contains_any(lowered_text, ["psychologist", "psychiatrist", "counsellor", "counselor"]):
            derived_codes.add("practical_recommendation")
            derived_codes.add("top_down_system_change")
    if row["question_id"] == "Q4" and _contains_any(lowered_text, ["illness", "crisis", "disease"]):
        if _contains_any(lowered_text, ["well-being", "wellbeing", "prevention", "not just"]):
            derived_codes.add("whole_person_care")
    if row["question_id"] in {"Q4", "Q5", "Q6", "Q7"} and _contains_any(lowered_text, ["suggestion box", "no follow-up", "لا يوجد متابعة", "ما في متابعة"]):
        derived_codes.add("top_down_system_change")
    if row["question_id"] in {"Q6", "Q7"} and _contains_any(lowered_text, ["hospital", "المستشفى"]):
        if _contains_any(lowered_text, ["plan", "تخطط", "should plan", "لازم"]):
            derived_codes.add("top_down_system_change")
            derived_codes.add("practical_recommendation")
    if row["question_id"] == "Q5" and _contains_any(lowered_text, ["emotion", "stress", "emotional", "نفسية", "العاطفية"]):
        if _contains_any(lowered_text, ["physical symptoms", "physical problem", "بدني", "جسدي", "ما يعرف كيف يعبر", "don't know how to express"]):
            derived_codes.add("self_disclosure_hidden_distress")
    return derived_codes


def _recode_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    recoded_rows: list[dict[str, str]] = []
    for row in rows:
        updated_row = dict(row)
        if row["speaker_type"] == "moderator":
            updated_row["codes"] = "moderator_context"
            recoded_rows.append(updated_row)
            continue

        derived_codes = _derive_role_codes(row["role_label"])
        derived_codes.update(_derive_text_codes(row))
        if not derived_codes.intersection(SEMANTIC_CODES):
            derived_codes.add("general_response")
        updated_row["codes"] = ";".join(_ordered_codes(derived_codes))
        recoded_rows.append(updated_row)
    return recoded_rows


def _build_participant_summary(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    grouped: dict[tuple[str, str, str, str], list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        if row["speaker_type"] != "participant":
            continue
        key = (
            row["speaker_code"],
            row["source_file"],
            row["table_id"],
            row["speaker_type"],
        )
        grouped[key].append(row)

    summary_rows: list[dict[str, str]] = []
    for key in sorted(grouped):
        speaker_code, source_file, table_id, speaker_type = key
        group_rows = grouped[key]
        question_ids = sorted({row["question_id"] for row in group_rows if row["question_id"]}, key=_question_sort_key)
        code_counter: Counter[str] = Counter()
        for row in group_rows:
            code_counter.update(_normalise_codes(row["codes"]))
        total_chars = sum(len(row["segment_text"]) for row in group_rows)
        summary_rows.append(
            {
                "anonymized_code": speaker_code,
                "source_file": source_file,
                "table_id": table_id,
                "speaker_type": speaker_type,
                "segment_count": str(len(group_rows)),
                "total_chars": str(total_chars),
                "questions_covered": ";".join(question_ids),
                "top_codes": _format_top_codes(code_counter),
            }
        )
    return summary_rows


def _build_question_evidence(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        grouped[row["question_id"]].append(row)

    evidence_rows: list[dict[str, str]] = []
    for question_id in sorted(grouped, key=_question_sort_key):
        group_rows = grouped[question_id]
        code_counter: Counter[str] = Counter()
        participant_speakers: set[str] = set()
        source_files = sorted({row["source_file"] for row in group_rows if row["source_file"]})
        participant_segments = 0
        moderator_segments = 0
        unclear_segments = 0

        for row in group_rows:
            code_counter.update(_normalise_codes(row["codes"]))
            speaker_type = row["speaker_type"]
            if speaker_type == "participant":
                participant_segments += 1
                if row["speaker_code"]:
                    participant_speakers.add(row["speaker_code"])
            elif speaker_type == "moderator":
                moderator_segments += 1
            else:
                unclear_segments += 1

        evidence_rows.append(
            {
                "question_id": question_id,
                "participant_segments": str(participant_segments),
                "moderator_segments": str(moderator_segments),
                "unclear_segments": str(unclear_segments),
                "unique_participant_speakers": str(len(participant_speakers)),
                "source_files": ";".join(source_files),
                "top_codes": _format_top_codes(code_counter),
            }
        )
    return evidence_rows


def _make_workbook_styles() -> tuple[Font, Alignment, Border, PatternFill]:
    header_font = Font(bold=True, size=11, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    blue_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
    return header_font, wrap, border, blue_fill


def _build_participant_workbook(summary_rows: list[dict[str, str]], target: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Participant_Summary"
    header_font, wrap, border, blue_fill = _make_workbook_styles()
    headers = [
        "Anonymized Code",
        "Source",
        "Table",
        "Speaker Type",
        "Segments",
        "Chars",
        "Questions",
        "Top Codes",
    ]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = wrap
        cell.border = border
    for row_index, row in enumerate(summary_rows, start=2):
        values = [
            row["anonymized_code"],
            row["source_file"],
            row["table_id"],
            row["speaker_type"],
            int(row["segment_count"]),
            int(row["total_chars"]),
            row["questions_covered"],
            row["top_codes"],
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.alignment = wrap
            cell.border = border
    workbook.save(target)


def _write_csv(csv_path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    """Promote the cleaned Day 2 segment base and build linked summary tables."""
    shutil.copyfile(CANDIDATE_SEGMENTS, CODED_SEGMENTS)
    coded_rows = _recode_rows(_load_rows(CODED_SEGMENTS))
    _write_csv(
        CODED_SEGMENTS,
        [
            "segment_id",
            "source_file",
            "table_id",
            "speaker_code",
            "speaker_type",
            "role_label",
            "attribution_status",
            "question_id",
            "segment_text",
            "codes",
            "language",
        ],
        coded_rows,
    )

    participant_summary_rows = _build_participant_summary(coded_rows)
    question_evidence_rows = _build_question_evidence(coded_rows)

    _write_csv(
        PARTICIPANT_SUMMARY,
        [
            "anonymized_code",
            "source_file",
            "table_id",
            "speaker_type",
            "segment_count",
            "total_chars",
            "questions_covered",
            "top_codes",
        ],
        participant_summary_rows,
    )
    _write_csv(
        QUESTION_EVIDENCE,
        [
            "question_id",
            "participant_segments",
            "moderator_segments",
            "unclear_segments",
            "unique_participant_speakers",
            "source_files",
            "top_codes",
        ],
        question_evidence_rows,
    )
    _build_participant_workbook(participant_summary_rows, PARTICIPANT_WORKBOOK)

    print(f"Promoted {CANDIDATE_SEGMENTS.name} -> {CODED_SEGMENTS.name}")
    print(f"Participant summary rows: {len(participant_summary_rows)}")
    print(f"Question evidence rows: {len(question_evidence_rows)}")
    print(f"Participant workbook: {PARTICIPANT_WORKBOOK.name}")


if __name__ == "__main__":
    main()
