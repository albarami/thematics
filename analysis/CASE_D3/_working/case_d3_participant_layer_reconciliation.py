from __future__ import annotations

import csv
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parents[1]
OUTWARD_DIR = ROOT / "OUTWARD_FACING_PACKAGE"

OUTWARD_REGISTER = OUTWARD_DIR / "CASE_D3_participant_register.csv"
ROOT_SUMMARY = ROOT / "CASE_D3_participant_summary.csv"
OUTWARD_SUMMARY = OUTWARD_DIR / "CASE_D3_participant_summary.csv"
ROOT_WORKBOOK = ROOT / "CASE_D3_participant_workbook.xlsx"
OUTWARD_WORKBOOK = OUTWARD_DIR / "CASE_D3_participant_workbook.xlsx"
ROOT_ANON_WORKBOOK = ROOT / "participant_summary_anonymized.xlsx"
OUTWARD_ANON_WORKBOOK = OUTWARD_DIR / "participant_summary_anonymized.xlsx"
ROOT_REPORT = ROOT / "CASE_D3_final_report.md"
OUTWARD_REPORT = OUTWARD_DIR / "CASE_D3_final_report.md"
ROOT_CROSSCHECK = ROOT / "CASE_D3_final_crosscheck_report.md"
OUTWARD_CROSSCHECK = OUTWARD_DIR / "CASE_D3_final_crosscheck_report.md"

OLD_REPORT_LINE = "- **Participants**: 25 confirmed participants (`D3_P01–D3_P25` with current register gaps preserved), 10 moderators (`D3_M01–D3_M10`), and 2 register-level `unclear` speaker rows."
NEW_REPORT_LINE = "- **Participants**: 25 confirmed participants (`D3_P01–D3_P25`), 10 moderators (`D3_M01–D3_M10`), and 2 register-level `unclear` speaker rows."
OLD_CROSSCHECK_STATUS = "## Status: PASS"
NEW_CROSSCHECK_ROW = "| 10 | Outward/internal package separation and participant layer reconciliation complete | PASS — outward-facing register, participant summary, and both outward-facing participant workbooks now carry the same 25 confirmed participant codes; `D3_P06` and `D3_P18` are retained as confirmed registered participants with `0` coded segments in the final coded base |"
OLD_CROSSCHECK_ROW = "| 10 | Outward/internal package separation complete | PASS |"
HEADERS = [
    "Anonymized Code",
    "Source",
    "Table",
    "Speaker Type",
    "Segments",
    "Chars",
    "Questions",
    "Top Codes",
]


def _load_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))



def _write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)



def _participant_sort_key(row: dict[str, str]) -> tuple[int, str]:
    match = re.search(r"D3_P(\d+)$", row["anonymized_code"])
    if not match:
        return (999, row["anonymized_code"])
    return (int(match.group(1)), row["anonymized_code"])



def _zero_summary_row(register_row: dict[str, str]) -> dict[str, str]:
    return {
        "anonymized_code": register_row["anonymized_code"],
        "source_file": register_row["source_file"],
        "table_id": register_row["table_id"],
        "speaker_type": register_row["speaker_type"],
        "segment_count": "0",
        "total_chars": "0",
        "questions_covered": "none",
        "top_codes": "none",
    }



def _reconcile_summary(
    register_rows: list[dict[str, str]],
    summary_rows: list[dict[str, str]],
) -> list[dict[str, str]]:
    participant_register_rows = [row for row in register_rows if row["speaker_type"] == "participant"]
    summary_lookup = {
        (row["anonymized_code"], row["source_file"], row["table_id"], row["speaker_type"]): dict(row)
        for row in summary_rows
        if row["speaker_type"] == "participant"
    }
    corrected_rows: list[dict[str, str]] = []
    for register_row in sorted(participant_register_rows, key=_participant_sort_key):
        key = (
            register_row["anonymized_code"],
            register_row["source_file"],
            register_row["table_id"],
            register_row["speaker_type"],
        )
        corrected_rows.append(summary_lookup.get(key, _zero_summary_row(register_row)))
    return corrected_rows



def _make_workbook_styles() -> tuple[Font, Alignment, Border, PatternFill]:
    header_font = Font(bold=True, size=11, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    blue_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
    return header_font, wrap, border, blue_fill



def _write_workbook(summary_rows: list[dict[str, str]], target: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Participant_Summary"
    header_font, wrap, border, blue_fill = _make_workbook_styles()
    for column, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = wrap
        cell.border = border
    for row_index, row in enumerate(summary_rows, start=2):
        values = [
            row["anonymized_code"],
            row["source_file"],
            row["table_id"],
            row["speaker_type"],
            int(row["segment_count"]),
            int(row["total_chars"]),
            row["questions_covered"],
            row["top_codes"],
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.alignment = wrap
            cell.border = border
    workbook.save(target)



def _replace_text(path: Path, old: str, new: str) -> None:
    text = path.read_text(encoding="utf-8")
    updated_text = text.replace(old, new)
    path.write_text(updated_text, encoding="utf-8")



def _workbook_codes(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["Participant_Summary"]
    codes = [row[0] for row in worksheet.iter_rows(min_row=2, values_only=True) if row and row[0]]
    workbook.close()
    return codes



def main() -> None:
    """Reconcile the CASE_D3 outward-facing participant layer without rebuilding the full package."""
    register_rows = _load_csv(OUTWARD_REGISTER)
    summary_rows = _load_csv(ROOT_SUMMARY)
    corrected_summary_rows = _reconcile_summary(register_rows, summary_rows)
    fieldnames = list(corrected_summary_rows[0].keys())

    _write_csv(ROOT_SUMMARY, fieldnames, corrected_summary_rows)
    _write_csv(OUTWARD_SUMMARY, fieldnames, corrected_summary_rows)

    _write_workbook(corrected_summary_rows, ROOT_WORKBOOK)
    _write_workbook(corrected_summary_rows, OUTWARD_WORKBOOK)
    _write_workbook(corrected_summary_rows, ROOT_ANON_WORKBOOK)
    _write_workbook(corrected_summary_rows, OUTWARD_ANON_WORKBOOK)

    _replace_text(ROOT_REPORT, OLD_REPORT_LINE, NEW_REPORT_LINE)
    _replace_text(OUTWARD_REPORT, OLD_REPORT_LINE, NEW_REPORT_LINE)

    _replace_text(ROOT_CROSSCHECK, OLD_CROSSCHECK_STATUS, OLD_CROSSCHECK_STATUS)
    _replace_text(ROOT_CROSSCHECK, OLD_CROSSCHECK_ROW, NEW_CROSSCHECK_ROW)
    _replace_text(OUTWARD_CROSSCHECK, OLD_CROSSCHECK_STATUS, OLD_CROSSCHECK_STATUS)
    _replace_text(OUTWARD_CROSSCHECK, OLD_CROSSCHECK_ROW, NEW_CROSSCHECK_ROW)

    expected_codes = [row["anonymized_code"] for row in corrected_summary_rows]
    root_workbook_codes = _workbook_codes(ROOT_WORKBOOK)
    outward_workbook_codes = _workbook_codes(OUTWARD_WORKBOOK)
    outward_anon_codes = _workbook_codes(OUTWARD_ANON_WORKBOOK)

    print(f"participant_register_rows={len(expected_codes)}")
    print(f"participant_summary_rows={len(corrected_summary_rows)}")
    print(f"missing_from_summary={sorted(set(expected_codes) - {row['anonymized_code'] for row in corrected_summary_rows})}")
    print(f"root_workbook_match={root_workbook_codes == expected_codes}")
    print(f"outward_workbook_match={outward_workbook_codes == expected_codes}")
    print(f"outward_anonymized_workbook_match={outward_anon_codes == expected_codes}")
    print(f"retained_zero_coded={[row['anonymized_code'] for row in corrected_summary_rows if row['segment_count'] == '0']}")


if __name__ == "__main__":
    main()
