from __future__ import annotations

import csv
import shutil
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(r"c:\Users\baram\OneDrive\Desktop\themnatic")
CASE_DIR = ROOT / "analysis" / "CASE_D3"
WORKING_DIR = CASE_DIR / "_working"
CANDIDATE_SEGMENTS = WORKING_DIR / "CASE_D3_segment_candidates.csv"
CODED_SEGMENTS = CASE_DIR / "CASE_D3_coded_segments.csv"
PARTICIPANT_SUMMARY = CASE_DIR / "CASE_D3_participant_summary.csv"
PARTICIPANT_WORKBOOK = CASE_DIR / "CASE_D3_participant_workbook.xlsx"
QUESTION_EVIDENCE = CASE_DIR / "CASE_D3_question_evidence_table.csv"
QUESTION_ORDER = ["Q1", "Q2", "Q3", "Q4", "Q5", "Q6", "Q7"]
TOP_CODE_LIMIT = 5
TOP_CODE_EXCLUSIONS = {
    "general_response",
    "moderator_context",
    "professional_identity",
    "adult_service_user_voice",
    "provider_patient_dual_position",
    "student_voice",
    "pillar_spiritual",
    "pillar_emotional",
    "pillar_intellectual",
    "pillar_physical",
    "pillar_social",
}
CODE_ORDER = [
    "balance_multidimensional",
    "contentment_acceptance",
    "inner_peace_stability",
    "functionality_coping_capacity",
    "spiritual_moral_anchor",
    "values_action_alignment",
    "relational_connectedness",
    "adult_role_responsibility",
    "family_caregiving_ecology",
    "work_life_strain",
    "self_responsibility_orientation",
    "system_responsibility_orientation",
    "hidden_distress_masking",
    "provider_self_as_adult",
    "trust_questioning_barrier",
    "health_literacy_navigation",
    "patient_provider_relationship",
    "empathy_reassurance",
    "consultation_time_pressure",
    "authority_defensiveness",
    "stigma_help_seeking_barrier",
    "language_cultural_mismatch",
    "fragmented_service_pathway",
    "interdisciplinary_integration",
    "awareness_outreach",
    "differentiated_service_design",
    "digital_admin_relief",
    "top_down_system_change",
    "healthcare_worker_wellbeing",
    "practical_recommendation",
    "pillar_spiritual",
    "pillar_emotional",
    "pillar_intellectual",
    "pillar_physical",
    "pillar_social",
    "professional_identity",
    "adult_service_user_voice",
    "provider_patient_dual_position",
    "student_voice",
    "moderator_context",
    "general_response",
]
ROLE_CODE_PATTERNS = {
    "professional_identity": [
        "consultant",
        "manager",
        "admin",
        "psychologist",
        "social work",
        "occupational therapist",
        "gp",
        "midwife",
        "nurse",
        "academic",
        "research",
        "laboratories",
        "leader",
        "health sciences",
        "طبيب",
        "دكتور",
        "استشاري",
        "أخصائي",
        "ممرضة",
        "باحث",
        "أستاذ",
        "مختبر",
    ],
    "adult_service_user_voice": ["service user", "beneficiary", "patient", "user of care", "service recipient", "مريض", "متلقي خدمة", "مراجع"],
    "student_voice": ["student", "medical student", "طالبة", "طالب"],
}
TEXT_CODE_PATTERNS = {
    "balance_multidimensional": ["توازن", "balance", "holistic", "integrated", "wholeness", "ركائز", "pillars", "all aspects"],
    "contentment_acceptance": ["رضا", "قناعة", "acceptance", "contentment", "satisfied", "at peace with"],
    "inner_peace_stability": ["طمأنينة", "راحة نفسية", "peace of mind", "calm", "stability", "stable", "settled"],
    "functionality_coping_capacity": ["cope", "coping", "function", "functioning", "manage", "carry on", "daily life", "continue", "يتحمل", "قادر"],
    "spiritual_moral_anchor": ["الله", "faith", "prayer", "الدين", "الروحية", "عبادة", "moral", "spiritual"],
    "values_action_alignment": ["values", "actions", "behavior", "behaviour", "aligned", "congruen", "principle", "اخلاق"],
    "relational_connectedness": ["support", "relationships", "belong", "connected", "people around", "family around", "علاقات", "حولك"],
    "adult_role_responsibility": ["responsibility", "responsibilities", "obligation", "burden", "duties", "مسؤولية", "واجب", "ألتفت لأولادي"],
    "family_caregiving_ecology": ["family", "children", "child", "parents", "marriage", "home", "house", "أهلي", "أسرتي", "أولادي", "زوج"],
    "work_life_strain": ["workload", "work", "job", "shift", "schedule", "pressure", "دوام", "شغال", "hours", "students and work"],
    "self_responsibility_orientation": ["start with myself", "ابتديت بنفسي", "مسؤوليتي تجاه نفسي", "my responsibility", "self-awareness", "start from the individual"],
    "system_responsibility_orientation": ["system", "systems", "institution", "institutional", "policy", "society", "community", "pathways", "school", "societal", "المجتمع", "مؤسسية"],
    "hidden_distress_masking": ["hidden", "mask", "surface", "inside", "looks fine", "not obvious", "داخلي", "من داخل", "ظاهر"],
    "provider_self_as_adult": ["for us", "our wellbeing", "كموظفين", "for providers", "as providers", "our own", "we as staff"],
    "trust_questioning_barrier": ["ask questions", "asking questions", "question my authority", "can we question", "offended", "trust", "ثق", "نسأل"],
    "health_literacy_navigation": ["health literate", "literacy", "explain", "understand", "clarify", "navigation", "rights", "map", "يفهم", "شرح"],
    "patient_provider_relationship": ["rapport", "relationship", "trust with doctor", "communication with patient", "doctor-patient", "doctor and patient", "تواصل مع المريض", "ثقة المريض"],
    "empathy_reassurance": ["reassure", "empathy", "human interaction", "compassionate", "acknowledge", "معوضة", "طمأنة", "احتواء", "calm your soul"],
    "consultation_time_pressure": ["no time", "limited time", "30 minute", "half an hour", "waiting list", "eight persons", "ضغط", "وقت", "نص ساعة", "30 دقيقة"],
    "authority_defensiveness": ["offended", "defensive", "authority", "authoritarian", "they don't like", "they feel offended", "لا يحب", "authority or"],
    "stigma_help_seeking_barrier": ["stigma", "taboo", "shame", "mental health", "عيب", "وصمة", "رفض العلاج النفسي"],
    "language_cultural_mismatch": ["language", "arabic", "english", "multilingual", "culture", "cultural", "translator", "culture diversity", "عربي", "اللغة"],
    "fragmented_service_pathway": ["pathway", "clear pathway", "referral", "disconnected", "fragmented", "silo", "coordination", "ما أعرف", "لا يوجد روابط"],
    "interdisciplinary_integration": ["team", "collaboration", "work together", "interdisciplinary", "between specialties", "weekly update", "different specialties", "فريق", "تعاون", "بين التخصصات"],
    "awareness_outreach": ["awareness", "campaign", "education", "educate", "outreach", "workshops", "teaching", "توعية", "حملات"],
    "differentiated_service_design": ["separate clinic", "expanded hours", "different times", "complex cases", "integrated clinic", "extended hours", "عيادة", "ساعات", "الحالات المعقدة"],
    "digital_admin_relief": ["ai", "technology", "digital", "documentation", "admin", "health map", "tool", "ذكاء اصطناعي"],
    "top_down_system_change": ["policy", "policies", "leadership", "ministry", "mandatory", "requirement", "strategy", "institutional", "regulation", "وزارة", "قرار", "سياسات"],
    "healthcare_worker_wellbeing": ["staff wellbeing", "provider wellbeing", "burnout", "workers", "employees", "cry corner", "for staff", "رفاه العاملين", "احتراق"],
    "practical_recommendation": ["suggest", "recommend", "proposal", "should", "need to", "لازم", "يجب", "ينبغي", "مقترح", "اقتراح"],
    "pillar_spiritual": ["ركيزة روحية", "روحية", "spiritual pillar", "الروحية"],
    "pillar_emotional": ["ركيزة عاطفية", "عاطفية", "emotional pillar", "العاطفية"],
    "pillar_intellectual": ["ركيزة فكرية", "فكرية", "intellectual pillar", "الفكرية"],
    "pillar_physical": ["ركيزة بدنية", "بدنية", "physical pillar", "البدنية", "جسدية"],
    "pillar_social": ["ركيزة اجتماعية", "اجتماعية", "social pillar", "الاجتماعية"],
}
SEMANTIC_CODES = {
    code
    for code in CODE_ORDER
    if code
    not in {
        "pillar_spiritual",
        "pillar_emotional",
        "pillar_intellectual",
        "pillar_physical",
        "pillar_social",
        "professional_identity",
        "adult_service_user_voice",
        "provider_patient_dual_position",
        "student_voice",
        "moderator_context",
        "general_response",
    }
}
PROPOSAL_MARKERS = ["suggest", "recommend", "proposal", "should", "need to", "لازم", "يجب", "ينبغي", "مقترح", "اقتراح"]
TIME_MARKERS = ["30 minute", "30 minutes", "half an hour", "waiting list", "queue", "time", "pressure", "وقت", "ضغط", "نص ساعة", "30 دقيقة"]
TEAM_MARKERS = ["team", "collaboration", "interdisciplinary", "work together", "weekly update", "between specialties", "فريق", "تعاون"]


def _load_rows(csv_path: Path) -> list[dict[str, str]]:
    with csv_path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))



def _question_sort_key(question_id: str) -> tuple[int, str]:
    try:
        return (QUESTION_ORDER.index(question_id), question_id)
    except ValueError:
        return (len(QUESTION_ORDER), question_id)



def _normalise_codes(raw_codes: str) -> list[str]:
    return [code.strip() for code in raw_codes.split(";") if code.strip()]



def _format_top_codes(counter: Counter[str]) -> str:
    filtered_counter = Counter({code: count for code, count in counter.items() if code not in TOP_CODE_EXCLUSIONS})
    ranked_source = filtered_counter if filtered_counter else counter
    ranked = sorted(ranked_source.items(), key=lambda item: (-item[1], item[0]))
    return ";".join(code for code, _ in ranked[:TOP_CODE_LIMIT])



def _contains_any(text: str, markers: list[str]) -> bool:
    return any(marker in text for marker in markers)



def _ordered_codes(codes: set[str]) -> list[str]:
    order_map = {code: index for index, code in enumerate(CODE_ORDER)}
    return sorted(codes, key=lambda code: (order_map.get(code, len(order_map)), code))



def _derive_role_codes(role_label: str) -> set[str]:
    lowered_role = role_label.lower()
    derived_codes: set[str] = set()
    for code, markers in ROLE_CODE_PATTERNS.items():
        if _contains_any(lowered_role, markers):
            derived_codes.add(code)
    return derived_codes



def _derive_text_codes(row: dict[str, str]) -> set[str]:
    lowered_text = row["segment_text"].lower()
    lowered_role = row["role_label"].lower()
    derived_codes: set[str] = set()

    for code, markers in TEXT_CODE_PATTERNS.items():
        if _contains_any(lowered_text, markers):
            derived_codes.add(code)

    if _contains_any(lowered_text, ["provider and patient", "as a provider and as a patient", "كمزود خدمة أو كمتلقي خدمة", "كمقدم خدمة أو متلقي خدمة"]):
        derived_codes.add("provider_patient_dual_position")
        derived_codes.add("adult_service_user_voice")

    if _contains_any(lowered_text, ["as a patient", "كمتلقي خدمة", "راجعت", "when i went", "user perspective"]):
        derived_codes.add("adult_service_user_voice")

    if _contains_any(lowered_text, ["medical student", "student", "طالبة", "طالب"]) and "student_voice" not in derived_codes:
        derived_codes.add("student_voice")

    if row["question_id"] == "Q7" and _contains_any(lowered_text, PROPOSAL_MARKERS):
        derived_codes.add("practical_recommendation")

    if row["question_id"] in {"Q6", "Q7"} and _contains_any(lowered_text, ["ai", "technology", "digital", "documentation", "ذكاء اصطناعي"]):
        derived_codes.add("digital_admin_relief")
        derived_codes.add("practical_recommendation")

    if row["question_id"] in {"Q4", "Q5", "Q6", "Q7"} and _contains_any(lowered_text, TIME_MARKERS):
        derived_codes.add("consultation_time_pressure")

    if row["question_id"] in {"Q5", "Q6", "Q7"} and _contains_any(lowered_text, TEAM_MARKERS):
        derived_codes.add("interdisciplinary_integration")

    if row["question_id"] in {"Q6", "Q7"} and _contains_any(lowered_text, ["campaign", "awareness", "education", "workshop", "توعية", "حملات"]):
        derived_codes.add("awareness_outreach")

    if row["question_id"] in {"Q6", "Q7"} and _contains_any(lowered_text, ["longer appointments", "complex cases", "expanded hours", "integrated clinic", "عيادة متكاملة", "الحالات المعقدة", "زيادة أوقات"]):
        derived_codes.add("differentiated_service_design")
        derived_codes.add("practical_recommendation")

    if row["question_id"] in {"Q4", "Q5", "Q6", "Q7"} and _contains_any(lowered_text, ["burnout", "providers", "workers", "employees", "cry corner", "wellbeing for the providers", "لنا كموظفين"]):
        derived_codes.add("healthcare_worker_wellbeing")

    if row["question_id"] in {"Q6", "Q7"} and _contains_any(lowered_text, ["policy", "ministry", "mandatory", "license", "cpd", "strategy", "institutional", "وزارة", "سياسات", "إلزام"]):
        derived_codes.add("top_down_system_change")
        derived_codes.add("practical_recommendation")

    if row["question_id"] in {"Q4", "Q5", "Q6"} and _contains_any(lowered_text, ["pathway", "referral", "clear pathway", "coordination", "لا يوجد روابط", "تحويل"]):
        derived_codes.add("fragmented_service_pathway")

    if row["question_id"] in {"Q2", "Q4", "Q5"} and _contains_any(lowered_text, ["hidden", "looks fine", "inside", "surface", "داخلي", "ظاهر"]):
        derived_codes.add("hidden_distress_masking")

    if _contains_any(lowered_text, ["question", "ask", "asking", "offended", "authority", "نسأل", "offended"]) and _contains_any(lowered_text, ["patient", "doctor", "provider", "gynecologist", "practitioner", "طبيب"]):
        derived_codes.add("trust_questioning_barrier")

    if _contains_any(lowered_text, ["offended", "authority", "challenge my authority", "challenging my authority"]):
        derived_codes.add("authority_defensiveness")

    if _contains_any(lowered_text, ["explain", "clarify", "health literate", "literacy", "rights", "map", "يفهم", "شرح"]):
        derived_codes.add("health_literacy_navigation")

    if _contains_any(lowered_text, ["reassure", "compassionate", "human interaction", "طمأنة", "معوضة", "acknowledgement"]):
        derived_codes.add("empathy_reassurance")

    if _contains_any(lowered_text, ["relationship", "rapport", "trust with doctor", "patient-doctor", "communication with patient", "ثقة المريض"]):
        derived_codes.add("patient_provider_relationship")

    if _contains_any(lowered_text, ["mental health", "عيب", "وصمة", "stigma", "taboo", "رفض العلاج النفسي"]):
        derived_codes.add("stigma_help_seeking_barrier")

    if _contains_any(lowered_text, ["language", "multilingual", "translator", "culture diversity", "arabic", "english", "اللغة", "عربي"]):
        derived_codes.add("language_cultural_mismatch")

    if _contains_any(lowered_role, ["student"]) and _contains_any(lowered_text, ["study", "student", "medical student", "طالبة"]):
        derived_codes.add("student_voice")

    if _contains_any(lowered_role, ["manager", "admin", "psychologist", "nurse", "midwife", "academic", "gp"]) and _contains_any(lowered_text, ["for us", "our wellbeing", "we as providers", "كموظفين", "our own"]):
        derived_codes.add("provider_self_as_adult")

    return derived_codes



def _recode_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    recoded_rows: list[dict[str, str]] = []
    for row in rows:
        updated_row = dict(row)
        if row["speaker_type"] == "moderator":
            updated_row["codes"] = "moderator_context"
            recoded_rows.append(updated_row)
            continue

        derived_codes = _derive_role_codes(row["role_label"])
        derived_codes.update(_derive_text_codes(row))
        if not derived_codes.intersection(SEMANTIC_CODES):
            derived_codes.add("general_response")
        updated_row["codes"] = ";".join(_ordered_codes(derived_codes))
        recoded_rows.append(updated_row)
    return recoded_rows



def _build_participant_summary(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    grouped: dict[tuple[str, str, str, str], list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        if row["speaker_type"] != "participant":
            continue
        key = (row["speaker_code"], row["source_file"], row["table_id"], row["speaker_type"])
        grouped[key].append(row)

    summary_rows: list[dict[str, str]] = []
    for key in sorted(grouped):
        speaker_code, source_file, table_id, speaker_type = key
        group_rows = grouped[key]
        question_ids = sorted({row["question_id"] for row in group_rows if row["question_id"]}, key=_question_sort_key)
        code_counter: Counter[str] = Counter()
        for row in group_rows:
            code_counter.update(_normalise_codes(row["codes"]))
        total_chars = sum(len(row["segment_text"]) for row in group_rows)
        summary_rows.append(
            {
                "anonymized_code": speaker_code,
                "source_file": source_file,
                "table_id": table_id,
                "speaker_type": speaker_type,
                "segment_count": str(len(group_rows)),
                "total_chars": str(total_chars),
                "questions_covered": ";".join(question_ids),
                "top_codes": _format_top_codes(code_counter),
            }
        )
    return summary_rows



def _build_question_evidence(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        grouped[row["question_id"]].append(row)

    evidence_rows: list[dict[str, str]] = []
    for question_id in sorted(grouped, key=_question_sort_key):
        group_rows = grouped[question_id]
        code_counter: Counter[str] = Counter()
        participant_speakers: set[str] = set()
        source_files = sorted({row["source_file"] for row in group_rows if row["source_file"]})
        participant_segments = 0
        moderator_segments = 0
        unclear_segments = 0

        for row in group_rows:
            code_counter.update(_normalise_codes(row["codes"]))
            speaker_type = row["speaker_type"]
            if speaker_type == "participant":
                participant_segments += 1
                if row["speaker_code"]:
                    participant_speakers.add(row["speaker_code"])
            elif speaker_type == "moderator":
                moderator_segments += 1
            else:
                unclear_segments += 1

        evidence_rows.append(
            {
                "question_id": question_id,
                "participant_segments": str(participant_segments),
                "moderator_segments": str(moderator_segments),
                "unclear_segments": str(unclear_segments),
                "unique_participant_speakers": str(len(participant_speakers)),
                "source_files": ";".join(source_files),
                "top_codes": _format_top_codes(code_counter),
            }
        )
    return evidence_rows



def _make_workbook_styles() -> tuple[Font, Alignment, Border, PatternFill]:
    header_font = Font(bold=True, size=11, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    blue_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
    return header_font, wrap, border, blue_fill



def _build_participant_workbook(summary_rows: list[dict[str, str]], target: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Participant_Summary"
    header_font, wrap, border, blue_fill = _make_workbook_styles()
    headers = [
        "Anonymized Code",
        "Source",
        "Table",
        "Speaker Type",
        "Segments",
        "Chars",
        "Questions",
        "Top Codes",
    ]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = wrap
        cell.border = border
    for row_index, row in enumerate(summary_rows, start=2):
        values = [
            row["anonymized_code"],
            row["source_file"],
            row["table_id"],
            row["speaker_type"],
            int(row["segment_count"]),
            int(row["total_chars"]),
            row["questions_covered"],
            row["top_codes"],
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.alignment = wrap
            cell.border = border
    workbook.save(target)



def _write_csv(csv_path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)



def main() -> None:
    """Promote the cleaned Day 3 segment base and build linked summary tables."""
    shutil.copyfile(CANDIDATE_SEGMENTS, CODED_SEGMENTS)
    coded_rows = _recode_rows(_load_rows(CODED_SEGMENTS))
    _write_csv(
        CODED_SEGMENTS,
        [
            "segment_id",
            "source_file",
            "table_id",
            "speaker_code",
            "speaker_type",
            "role_label",
            "attribution_status",
            "question_id",
            "segment_text",
            "codes",
            "language",
        ],
        coded_rows,
    )

    participant_summary_rows = _build_participant_summary(coded_rows)
    question_evidence_rows = _build_question_evidence(coded_rows)

    _write_csv(
        PARTICIPANT_SUMMARY,
        [
            "anonymized_code",
            "source_file",
            "table_id",
            "speaker_type",
            "segment_count",
            "total_chars",
            "questions_covered",
            "top_codes",
        ],
        participant_summary_rows,
    )
    _write_csv(
        QUESTION_EVIDENCE,
        [
            "question_id",
            "participant_segments",
            "moderator_segments",
            "unclear_segments",
            "unique_participant_speakers",
            "source_files",
            "top_codes",
        ],
        question_evidence_rows,
    )
    _build_participant_workbook(participant_summary_rows, PARTICIPANT_WORKBOOK)

    print(f"Promoted {CANDIDATE_SEGMENTS.name} -> {CODED_SEGMENTS.name}")
    print(f"Participant summary rows: {len(participant_summary_rows)}")
    print(f"Question evidence rows: {len(question_evidence_rows)}")
    print(f"Participant workbook: {PARTICIPANT_WORKBOOK.name}")


if __name__ == "__main__":
    main()
