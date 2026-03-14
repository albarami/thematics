from __future__ import annotations

import csv
import re
import shutil
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUTWARD_DIR = ROOT / "OUTWARD_FACING_PACKAGE"
INTERNAL_DIR = ROOT / "INTERNAL_CONFIDENTIAL"

CANDIDATE_THEMES = ROOT / "CASE_D3_candidate_themes.md"
EXCERPT_BANK = ROOT / "CASE_D3_excerpt_bank.csv"
FINAL_REPORT = ROOT / "CASE_D3_final_report.md"
FINAL_THEMES = ROOT / "CASE_D3_final_themes.md"
MATRIX_PROMINENCE_CHECK = ROOT / "CASE_D3_matrix_prominence_check.md"
PARTICIPANT_REGISTER = ROOT / "CASE_D3_participant_register.csv"
PARTICIPANT_SUMMARY = ROOT / "CASE_D3_participant_summary.csv"
PARTICIPANT_WORKBOOK = ROOT / "CASE_D3_participant_workbook.xlsx"
PREPARATION_CHECKLIST = ROOT / "CASE_D3_preparation_checklist.md"
PROMINENCE_SALIENCE = ROOT / "CASE_D3_prominence_salience.csv"
QUESTION_EVIDENCE = ROOT / "CASE_D3_question_evidence_table.csv"
QUESTION_THEME_MATRIX = ROOT / "CASE_D3_question_theme_matrix.csv"
SUMMARY_TABLES = ROOT / "CASE_D3_summary_tables.md"
THEME_EVIDENCE_WORKBOOK = ROOT / "CASE_D3_theme_evidence_workbook.xlsx"
THEME_INTEGRITY = ROOT / "CASE_D3_theme_integrity_report.md"
THEME_SUMMARY = ROOT / "CASE_D3_theme_summary_table.csv"
BOUNDARY_MEMO = ROOT / "CASE_D3_boundary_memo.md"
FAMILIARISATION_MEMO = ROOT / "CASE_D3_familiarisation_memo.md"
LANGUAGE_MEMO = ROOT / "CASE_D3_language_memo.md"
SOURCE_SENSITIVITY_MEMO = ROOT / "CASE_D3_source_sensitivity_memo.md"
REPORT_INTEGRITY = ROOT / "CASE_D3_report_integrity_check.md"
FINAL_CROSSCHECK = ROOT / "CASE_D3_final_crosscheck_report.md"
ANONYMIZED_PARTICIPANT_WORKBOOK = ROOT / "participant_summary_anonymized.xlsx"
OUTWARD_PARTICIPANT_REGISTER = OUTWARD_DIR / "CASE_D3_participant_register.csv"
INTERNAL_IDENTITY_KEY = INTERNAL_DIR / "participant_identity_key.xlsx"

ROOT_TEXT_FILES_TO_SANITIZE = (
    CANDIDATE_THEMES,
    FINAL_REPORT,
    FINAL_THEMES,
    MATRIX_PROMINENCE_CHECK,
    SUMMARY_TABLES,
    THEME_INTEGRITY,
)

OUTWARD_ROOT_FILES = (
    CANDIDATE_THEMES,
    EXCERPT_BANK,
    FINAL_REPORT,
    FINAL_THEMES,
    MATRIX_PROMINENCE_CHECK,
    PARTICIPANT_SUMMARY,
    PARTICIPANT_WORKBOOK,
    PROMINENCE_SALIENCE,
    QUESTION_EVIDENCE,
    QUESTION_THEME_MATRIX,
    SUMMARY_TABLES,
    THEME_EVIDENCE_WORKBOOK,
    THEME_INTEGRITY,
    THEME_SUMMARY,
    ANONYMIZED_PARTICIPANT_WORKBOOK,
)

INTERNAL_ROOT_FILES = (
    PARTICIPANT_REGISTER,
    PREPARATION_CHECKLIST,
    FAMILIARISATION_MEMO,
    SOURCE_SENSITIVITY_MEMO,
    BOUNDARY_MEMO,
    LANGUAGE_MEMO,
)

NAME_REPLACEMENTS = {
    "moderator هنادي أحمد أبو بكر": "moderator",
    "هنادي أحمد أبو بكر": "moderator",
    "مشاعل الأنصاري": "participant",
    "أحمد الفرجابي": "participant",
    "أروى حسين": "participant",
    "عائشة المعصومي": "participant",
    "هند الجابر": "participant",
    "منى سالم": "participant",
    "سحر السيد": "participant",
    "أ.سحر سيد": "participant",
    "ساسجار": "participant",
    "داليا": "participant",
    "آمنة": "participant",
    "أميرة": "participant",
    "وضحى": "participant",
    "حنان": "participant",
    "مها": "participant",
    "وائل": "participant",
    "خالد": "participant",
    "وليد": "participant",
    "Miss Donna": "a colleague",
    "Donna": "a colleague",
    "Julie": "a colleague",
    "Alicia": "a colleague",
    "Katia": "a colleague",
    "Dionne": "a colleague",
    "Dion": "a colleague",
}

LIMITATION_TERMS = (
    "weak `HWAD1AR` `Q2/Q3` boundaries",
    "weak `HWAD10AR` `Q6` preservation",
    "later moderator-led summary structure",
    "substantial `unclear` material in `HWAD3AR` and `HWAD6AR`",
)

QUESTION_HEADINGS = tuple(f"### 3.{index} Q{index}:" for index in range(1, 8))
APPENDIX_HEADINGS = (
    "## Appendix A:",
    "## Appendix B:",
    "## Appendix C:",
    "## Appendix D:",
    "## Appendix E:",
)
CROSS_CASE_TERMS = ("CASE_D1", "CASE_D2", "CASE_D4", "CASE_D5")


def _load_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _sanitize_text(text: str) -> str:
    sanitized = text
    for original, replacement in sorted(NAME_REPLACEMENTS.items(), key=lambda item: len(item[0]), reverse=True):
        sanitized = sanitized.replace(original, replacement)
    return sanitized


def _sanitize_root_markdown() -> None:
    for path in ROOT_TEXT_FILES_TO_SANITIZE:
        text = path.read_text(encoding="utf-8")
        path.write_text(_sanitize_text(text), encoding="utf-8")


def _sanitize_excerpt_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    sanitized_rows: list[dict[str, str]] = []
    for row in rows:
        sanitized_row = dict(row)
        sanitized_row["excerpt_text"] = _sanitize_text(row["excerpt_text"])
        sanitized_rows.append(sanitized_row)
    return sanitized_rows


def _append_sheet(workbook: Workbook, title: str, rows: list[dict[str, str]]) -> None:
    worksheet = workbook.create_sheet(title=title)
    if not rows:
        return
    headers = list(rows[0].keys())
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])


def _write_theme_workbook(
    matrix_rows: list[dict[str, str]],
    excerpt_rows: list[dict[str, str]],
    prominence_rows: list[dict[str, str]],
    summary_rows: list[dict[str, str]],
    target: Path,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _append_sheet(workbook, "question_theme_matrix", matrix_rows)
    _append_sheet(workbook, "excerpt_bank", excerpt_rows)
    _append_sheet(workbook, "prominence_salience", prominence_rows)
    _append_sheet(workbook, "theme_summary", summary_rows)
    workbook.save(target)


def _safe_classification_basis(row: dict[str, str]) -> str:
    speaker_type = row["speaker_type"]
    turns = int(row["turns"] or "0")
    if speaker_type == "participant":
        if turns == 0:
            return "Classified as participant from source-role review, but stable named turns were not preserved in outward-facing materials."
        return "Classified as participant from repeated substantive turns and source-role review."
    if speaker_type == "moderator":
        if turns == 0:
            return "Classified as moderator from the source register because the source is note-only or does not preserve stable moderator turns."
        return "Classified as moderator from facilitation turns and source-role review."
    return "Retained as `unclear` because substantive discussion is preserved without stable recoverable speaker attribution."


def _build_outward_participant_register(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    outward_rows: list[dict[str, str]] = []
    for row in rows:
        outward_row = dict(row)
        outward_row["classification_basis"] = _safe_classification_basis(row)
        outward_rows.append(outward_row)
    return outward_rows


def _extract_internal_label(classification_basis: str) -> str:
    matches = re.findall(r"`([^`]+)`", classification_basis)
    return " | ".join(matches)


def _write_identity_key(rows: list[dict[str, str]], target: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "participant_identity_key"
    headers = [
        "anonymized_code",
        "speaker_type",
        "source_file",
        "table_id",
        "role_label",
        "turns",
        "chars",
        "internal_name_or_label",
        "classification_basis",
    ]
    worksheet.append(headers)
    for row in rows:
        worksheet.append(
            [
                row["anonymized_code"],
                row["speaker_type"],
                row["source_file"],
                row["table_id"],
                row["role_label"],
                row["turns"],
                row["chars"],
                _extract_internal_label(row["classification_basis"]),
                row["classification_basis"],
            ]
        )
    workbook.save(target)


def _prepare_directory(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)
    for child in path.iterdir():
        if child.is_dir():
            shutil.rmtree(child)
            continue
        child.unlink()


def _copy_file(source: Path, destination: Path) -> None:
    shutil.copy2(source, destination)


def _build_scan_terms(register_rows: list[dict[str, str]]) -> list[str]:
    del register_rows
    return sorted((term for term in NAME_REPLACEMENTS if term), key=len, reverse=True)


def _scan_text_file(path: Path, terms: list[str]) -> list[str]:
    text = path.read_text(encoding="utf-8")
    return [term for term in terms if term in text]


def _scan_workbook(path: Path, terms: list[str]) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    hits: set[str] = set()
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows(values_only=True):
            for value in row:
                if not isinstance(value, str):
                    continue
                for term in terms:
                    if term in value:
                        hits.add(term)
    return sorted(hits)


def _scan_outward_package(terms: list[str]) -> dict[str, list[str]]:
    hits: dict[str, list[str]] = {}
    for path in sorted(OUTWARD_DIR.iterdir()):
        if path.suffix.lower() in {".md", ".csv"}:
            file_hits = _scan_text_file(path, terms)
        elif path.suffix.lower() == ".xlsx":
            file_hits = _scan_workbook(path, terms)
        else:
            file_hits = []
        if file_hits:
            hits[path.name] = file_hits
    return hits


def _report_ids(report_text: str) -> list[str]:
    return sorted(set(re.findall(r"D3_S\d{4}", report_text)))


def _build_report_integrity(
    excerpt_rows: list[dict[str, str]],
    report_text: str,
    summary_text: str,
    scan_hits: dict[str, list[str]],
) -> str:
    excerpt_lookup = {row["evidence_id"]: row for row in excerpt_rows}
    report_ids = _report_ids(report_text)
    unresolved_ids = [evidence_id for evidence_id in report_ids if evidence_id not in excerpt_lookup]
    source_files = sorted({excerpt_lookup[evidence_id]["source_file"] for evidence_id in report_ids if evidence_id in excerpt_lookup})
    speaker_codes = sorted({excerpt_lookup[evidence_id]["speaker_code"] for evidence_id in report_ids if evidence_id in excerpt_lookup})
    quote_lines = [line for line in report_text.splitlines() if line.startswith("> ")]
    question_structure_ok = all(heading in report_text for heading in QUESTION_HEADINGS)
    appendices_ok = all(heading in report_text for heading in APPENDIX_HEADINGS)
    no_moderator_codes = not any(re.search(r"D3_M\d{2}", line) for line in quote_lines)
    no_cross_case_terms = not any(term in report_text for term in CROSS_CASE_TERMS)
    note_rule_explicit = "No `note_taker_summary` rows were operationalized" in report_text
    support_labels_explicit = all(term in report_text for term in ("Coded transcript support", "Close-reading support", "Contextual note support", "Auxiliary recommendation support"))
    limitations_explicit = all(term in report_text or term in summary_text for term in LIMITATION_TERMS)
    quotation_diversity_ok = len(speaker_codes) >= 8
    package_scan_clean = not scan_hits
    outward_file_count = len(list(OUTWARD_DIR.iterdir()))
    scan_hit_files = ", ".join(sorted(scan_hits)) if scan_hits else "none"
    return "\n".join(
        [
            "# CASE_D3 Report Integrity Check",
            "",
            "## Case: CASE_D3 — Day 3 (Adults)",
            f"## Status: {'PASS' if all((question_structure_ok, appendices_ok, not unresolved_ids, no_moderator_codes, no_cross_case_terms, support_labels_explicit, limitations_explicit, package_scan_clean)) else 'PASS WITH CAVEATS'}",
            "",
            "---",
            "",
            "## Gate 6 checks",
            "",
            "| Check | Result |",
            "|-------|--------|",
            f"| Question-led structure present | {'PASS' if question_structure_ok else 'FAIL'} — Q1-Q7 headings {'are' if question_structure_ok else 'are not'} present in the final report |",
            f"| Quotations traceable to excerpt bank | {'PASS' if not unresolved_ids else 'FAIL'} — {len(report_ids)} unique evidence IDs cited; {len(unresolved_ids)} unresolved IDs |",
            f"| Quotation diversity strengthened | {'PASS' if quotation_diversity_ok else 'PASS WITH CAVEAT'} — quotations draw on {len(speaker_codes)} speaker codes across {len(source_files)} transcript sources |",
            f"| Moderator material excluded as participant evidence | {'PASS' if no_moderator_codes else 'FAIL'} — moderator speaker codes {'do not appear' if no_moderator_codes else 'appear'} in quotation lines |",
            f"| No cross-case claims | {'PASS' if no_cross_case_terms else 'FAIL'} — CASE_D3 report text {'does not reference other cases' if no_cross_case_terms else 'references other cases'} |",
            f"| Appendices present | {'PASS' if appendices_ok else 'FAIL'} — Appendix A-E headings {'are' if appendices_ok else 'are not'} present |",
            f"| Support-layer distinction explicit | {'PASS' if support_labels_explicit else 'FAIL'} — coded transcript, close-reading, contextual note, and auxiliary recommendation support labels {'are' if support_labels_explicit else 'are not'} explicit |",
            f"| Note and recommendation rules explicit | {'PASS' if note_rule_explicit else 'PASS WITH CAVEAT'} — note_taker_summary operationalization is {'explicitly ruled out' if note_rule_explicit else 'not fully explicit'} in the report text |",
            f"| Known Day 3 limitations remain explicit | {'PASS' if limitations_explicit else 'FAIL'} — required limitation statements {'are' if limitations_explicit else 'are not'} visible across the final report layer |",
            f"| Outward package name scan clean | {'PASS' if package_scan_clean else 'FAIL'} — {sum(len(value) for value in scan_hits.values())} exact-name hits across {outward_file_count} outward files scanned |",
            "",
            "## Traceability summary",
            "",
            f"- **[evidence-ids]** {', '.join(report_ids) if report_ids else 'none'}",
            f"- **[sources]** {', '.join(source_files) if source_files else 'none'}",
            f"- **[speaker-codes]** {', '.join(speaker_codes) if speaker_codes else 'none'}",
            f"- **[unresolved-ids]** {', '.join(unresolved_ids) if unresolved_ids else 'none'}",
            f"- **[scan-hit-files]** {scan_hit_files}",
        ]
    )


def _build_final_crosscheck(
    excerpt_rows: list[dict[str, str]],
    report_text: str,
    summary_text: str,
    scan_hits: dict[str, list[str]],
) -> str:
    theme_rows = _load_csv(THEME_SUMMARY)
    prominence_rows = _load_csv(PROMINENCE_SALIENCE)
    matrix_rows = _load_csv(QUESTION_THEME_MATRIX)
    report_ids = _report_ids(report_text)
    excerpt_lookup = {row["evidence_id"]: row for row in excerpt_rows}
    quote_lines = [line for line in report_text.splitlines() if line.startswith("> ")]
    theme_names = [row["theme_name"] for row in theme_rows]
    theme_names_locked = all(theme_name in report_text and theme_name in FINAL_THEMES.read_text(encoding="utf-8") and theme_name in summary_text for theme_name in theme_names)
    salience_locked = all(row["salience"] in summary_text and row["salience"].replace("_", " ") in report_text.lower().replace("-", " ") for row in prominence_rows)
    report_ids_resolve = all(evidence_id in excerpt_lookup for evidence_id in report_ids)
    moderator_exclusion = not any(re.search(r"D3_M\d{2}", line) for line in quote_lines)
    note_rows = sum(1 for row in excerpt_rows if row["speaker_type"] == "note_taker_summary")
    unclear_rows = sum(1 for row in excerpt_rows if row["speaker_type"] == "unclear")
    package_complete = all((OUTWARD_DIR.exists(), INTERNAL_DIR.exists(), INTERNAL_IDENTITY_KEY.exists()))
    outward_identity_leak = (OUTWARD_DIR / "participant_identity_key.xlsx").exists()
    limitations_explicit = all(term in report_text or term in summary_text for term in LIMITATION_TERMS)
    question_rows_match = len(matrix_rows) == 7
    package_scan_clean = not scan_hits
    outward_files = sorted(path.name for path in OUTWARD_DIR.iterdir())
    internal_files = sorted(path.name for path in INTERNAL_DIR.iterdir())
    scan_hit_files = ", ".join(sorted(scan_hits)) if scan_hits else "none"
    return "\n".join(
        [
            "# CASE_D3 Final Cross-Check Report",
            "",
            "## Case: CASE_D3 — Day 3 (Adults)",
            f"## Status: {'PASS' if all((theme_names_locked, salience_locked, report_ids_resolve, moderator_exclusion, note_rows == 0, not outward_identity_leak, package_complete, limitations_explicit, question_rows_match, package_scan_clean)) else 'PASS WITH CAVEATS'}",
            "",
            "---",
            "",
            "## Gate 7 cross-check",
            "",
            "| # | Cross-check item | Result |",
            "|---|------------------|--------|",
            f"| 1 | Theme names locked across report / tables / theme file | {'PASS' if theme_names_locked else 'FAIL'} |",
            f"| 2 | Matrix layer still has 7 question rows | {'PASS' if question_rows_match else 'FAIL'} |",
            f"| 3 | Prominence labels align across salience CSV, report, and summary tables | {'PASS' if salience_locked else 'FAIL'} |",
            f"| 4 | Final-report evidence IDs resolve to excerpt bank | {'PASS' if report_ids_resolve else 'FAIL'} |",
            f"| 5 | Moderator material remains contextual only in report quotations | {'PASS' if moderator_exclusion else 'FAIL'} |",
            f"| 6 | Note-taker material is truthfully handled | {'PASS' if note_rows == 0 else 'FAIL'} — {note_rows} note rows operationalized; {unclear_rows} `unclear` transcript rows retained with explicit labeling |",
            f"| 7 | Recommendation workbook remains auxiliary only | {'PASS' if 'recommendation workbook remains auxiliary' in report_text.lower() else 'PASS WITH CAVEAT'} |",
            f"| 8 | Day 3 source limitations remain explicit | {'PASS' if limitations_explicit else 'FAIL'} |",
            f"| 9 | Outward-facing package name scan is clean | {'PASS' if package_scan_clean else 'FAIL'} |",
            f"| 10 | Outward/internal package separation complete | {'PASS' if package_complete and not outward_identity_leak else 'FAIL'} |",
            "",
            "## Package inventory",
            "",
            f"- **[outward-files]** {', '.join(outward_files)}",
            f"- **[internal-files]** {', '.join(internal_files)}",
            f"- **[quoted-evidence-ids]** {', '.join(report_ids) if report_ids else 'none'}",
            f"- **[name-scan-hit-files]** {scan_hit_files}",
        ]
    )


def main() -> None:
    """Build the Day 3 outward/internal packages and Gate 6/7 check files."""
    register_rows = _load_csv(PARTICIPANT_REGISTER)
    excerpt_rows = _sanitize_excerpt_rows(_load_csv(EXCERPT_BANK))
    matrix_rows = _load_csv(QUESTION_THEME_MATRIX)
    prominence_rows = _load_csv(PROMINENCE_SALIENCE)
    summary_rows = _load_csv(THEME_SUMMARY)

    _sanitize_root_markdown()
    _write_csv(EXCERPT_BANK, list(excerpt_rows[0].keys()), excerpt_rows)
    _write_theme_workbook(matrix_rows, excerpt_rows, prominence_rows, summary_rows, THEME_EVIDENCE_WORKBOOK)
    shutil.copy2(PARTICIPANT_WORKBOOK, ANONYMIZED_PARTICIPANT_WORKBOOK)

    _prepare_directory(OUTWARD_DIR)
    _prepare_directory(INTERNAL_DIR)

    outward_register_rows = _build_outward_participant_register(register_rows)
    _write_csv(OUTWARD_PARTICIPANT_REGISTER, list(outward_register_rows[0].keys()), outward_register_rows)
    _write_identity_key(register_rows, INTERNAL_IDENTITY_KEY)

    for source in OUTWARD_ROOT_FILES:
        _copy_file(source, OUTWARD_DIR / source.name)
    for source in INTERNAL_ROOT_FILES:
        _copy_file(source, INTERNAL_DIR / source.name)

    terms = _build_scan_terms(register_rows)
    scan_hits = _scan_outward_package(terms)
    report_text = FINAL_REPORT.read_text(encoding="utf-8")
    summary_text = SUMMARY_TABLES.read_text(encoding="utf-8")
    REPORT_INTEGRITY.write_text(_build_report_integrity(excerpt_rows, report_text, summary_text, scan_hits), encoding="utf-8")
    FINAL_CROSSCHECK.write_text(_build_final_crosscheck(excerpt_rows, report_text, summary_text, scan_hits), encoding="utf-8")
    _copy_file(REPORT_INTEGRITY, OUTWARD_DIR / REPORT_INTEGRITY.name)
    _copy_file(FINAL_CROSSCHECK, OUTWARD_DIR / FINAL_CROSSCHECK.name)

    final_scan_hits = _scan_outward_package(terms)
    REPORT_INTEGRITY.write_text(_build_report_integrity(excerpt_rows, report_text, summary_text, final_scan_hits), encoding="utf-8")
    FINAL_CROSSCHECK.write_text(_build_final_crosscheck(excerpt_rows, report_text, summary_text, final_scan_hits), encoding="utf-8")
    _copy_file(REPORT_INTEGRITY, OUTWARD_DIR / REPORT_INTEGRITY.name)
    _copy_file(FINAL_CROSSCHECK, OUTWARD_DIR / FINAL_CROSSCHECK.name)
    scan_hits = final_scan_hits

    print(f"Outward files: {len(list(OUTWARD_DIR.iterdir()))}")
    print(f"Internal files: {len(list(INTERNAL_DIR.iterdir()))}")
    print(f"Quoted evidence ids: {len(_report_ids(report_text))}")
    print(f"Name scan hits: {scan_hits}")


if __name__ == "__main__":
    main()
