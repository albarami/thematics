"""Build Excel workbooks for CASE_D1: participant summary and theme evidence."""
import csv
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("openpyxl not available")
    exit(1)

root = Path(r'c:\Users\baram\OneDrive\Desktop\themnatic')
d1_out = root / 'analysis' / 'CASE_D1'

# Read data
with open(d1_out / 'CASE_D1_coded_segments.csv', 'r', encoding='utf-8') as f:
    segments = list(csv.DictReader(f))

with open(d1_out / 'CASE_D1_participant_summary.csv', 'r', encoding='utf-8') as f:
    participants = list(csv.DictReader(f))

with open(d1_out / 'CASE_D1_excerpt_bank.csv', 'r', encoding='utf-8') as f:
    excerpts = list(csv.DictReader(f))

with open(d1_out / 'CASE_D1_prominence_salience.csv', 'r', encoding='utf-8') as f:
    prominence = list(csv.DictReader(f))

with open(d1_out / 'CASE_D1_question_theme_matrix.csv', 'r', encoding='utf-8') as f:
    matrix = list(csv.DictReader(f))

# Styles
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font_white = Font(bold=True, size=11, color='FFFFFF')
wrap = Alignment(wrap_text=True, vertical='top')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, row_num, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = wrap
        cell.border = thin_border

def style_data(ws, start_row, end_row, num_cols):
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = wrap
            cell.border = thin_border

# ═══════════════════════════════════════════════════════
# WORKBOOK 1: Participant Summary
# ═══════════════════════════════════════════════════════
wb1 = Workbook()

# Sheet 1: All participants
ws1 = wb1.active
ws1.title = 'Participant_Summary'
headers = ['Speaker Label', 'Source File', 'Table', 'Speaker Type', 'Segments', 'Total Chars', 'Questions Covered', 'Top Codes']
for col, h in enumerate(headers, 1):
    ws1.cell(row=1, column=col, value=h)
style_header(ws1, 1, len(headers))

# Sort: participants first, then by total_chars descending
p_sorted = sorted(participants, key=lambda x: (0 if x['speaker_type'] == 'participant' else 1 if x['speaker_type'] == 'unclear' else 2, -int(x['total_chars'])))
for i, p in enumerate(p_sorted, 2):
    ws1.cell(row=i, column=1, value=p['speaker_label'])
    ws1.cell(row=i, column=2, value=p['source_file'])
    ws1.cell(row=i, column=3, value=p['table_id'])
    ws1.cell(row=i, column=4, value=p['speaker_type'])
    ws1.cell(row=i, column=5, value=int(p['segment_count']))
    ws1.cell(row=i, column=6, value=int(p['total_chars']))
    ws1.cell(row=i, column=7, value=p['questions_covered'])
    ws1.cell(row=i, column=8, value=p['top_codes'])
style_data(ws1, 2, len(p_sorted) + 1, len(headers))

# Column widths
ws1.column_dimensions['A'].width = 40
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 8
ws1.column_dimensions['D'].width = 15
ws1.column_dimensions['E'].width = 10
ws1.column_dimensions['F'].width = 12
ws1.column_dimensions['G'].width = 25
ws1.column_dimensions['H'].width = 45

# Sheet 2: Source contribution
ws2 = wb1.create_sheet('Source_Contribution')
src_headers = ['Source File', 'Total Segments', 'Participant Segments', 'Unique Speakers', 'Questions', 'Themes Present']
with open(d1_out / 'CASE_D1_source_contribution_table.csv', 'r', encoding='utf-8') as f:
    src_data = list(csv.DictReader(f))
for col, h in enumerate(src_headers, 1):
    ws2.cell(row=1, column=col, value=h)
style_header(ws2, 1, len(src_headers))
for i, s in enumerate(src_data, 2):
    ws2.cell(row=i, column=1, value=s['source_file'])
    ws2.cell(row=i, column=2, value=int(s['total_segments']))
    ws2.cell(row=i, column=3, value=int(s['participant_segments']))
    ws2.cell(row=i, column=4, value=int(s['unique_speakers']))
    ws2.cell(row=i, column=5, value=s['questions_covered'])
    ws2.cell(row=i, column=6, value=s['themes_present'])
style_data(ws2, 2, len(src_data) + 1, len(src_headers))
ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 20
ws2.column_dimensions['D'].width = 15
ws2.column_dimensions['E'].width = 30
ws2.column_dimensions['F'].width = 60

wb1.save(d1_out / 'CASE_D1_participant_workbook.xlsx')
print('Participant workbook saved.')

# ═══════════════════════════════════════════════════════
# WORKBOOK 2: Theme Evidence
# ═══════════════════════════════════════════════════════
wb2 = Workbook()

# Sheet 1: Theme summary
ws_ts = wb2.active
ws_ts.title = 'Theme_Summary'
ts_headers = ['Theme', 'Segments', 'Speakers', 'Tables', 'Total Chars', 'Questions', 'Salience', 'Explanation']
for col, h in enumerate(ts_headers, 1):
    ws_ts.cell(row=1, column=col, value=h)
style_header(ws_ts, 1, len(ts_headers))
for i, p in enumerate(prominence, 2):
    ws_ts.cell(row=i, column=1, value=p['theme'])
    ws_ts.cell(row=i, column=2, value=int(p['participant_segments']))
    ws_ts.cell(row=i, column=3, value=int(p['unique_speakers']))
    ws_ts.cell(row=i, column=4, value=int(p['unique_tables']))
    ws_ts.cell(row=i, column=5, value=int(p['total_chars']))
    ws_ts.cell(row=i, column=6, value=p['questions_present'])
    ws_ts.cell(row=i, column=7, value=p['salience'])
    ws_ts.cell(row=i, column=8, value=p['salience_explanation'])
style_data(ws_ts, 2, len(prominence) + 1, len(ts_headers))
ws_ts.column_dimensions['A'].width = 45
ws_ts.column_dimensions['B'].width = 12
ws_ts.column_dimensions['C'].width = 12
ws_ts.column_dimensions['D'].width = 10
ws_ts.column_dimensions['E'].width = 12
ws_ts.column_dimensions['F'].width = 25
ws_ts.column_dimensions['G'].width = 28
ws_ts.column_dimensions['H'].width = 80

# Sheet 2: Q x Theme Matrix
ws_mx = wb2.create_sheet('QxTheme_Matrix')
theme_names = ['Theme_1_Balanced_Contentment', 'Theme_2_Care_Ecology', 'Theme_3_Service_Fragmentation', 'Theme_4_Culturally_Grounded_Solutions']
mx_headers = ['Question']
for t in theme_names:
    short = t.split('_', 2)[-1]
    mx_headers.extend([f'{short}_segs', f'{short}_spk', f'{short}_tbl'])
for col, h in enumerate(mx_headers, 1):
    ws_mx.cell(row=1, column=col, value=h)
style_header(ws_mx, 1, len(mx_headers))
for i, m in enumerate(matrix, 2):
    ws_mx.cell(row=i, column=1, value=m['question_id'])
    col = 2
    for t in theme_names:
        ws_mx.cell(row=i, column=col, value=int(m.get(f'{t}_segments', 0)))
        ws_mx.cell(row=i, column=col+1, value=int(m.get(f'{t}_speakers', 0)))
        ws_mx.cell(row=i, column=col+2, value=int(m.get(f'{t}_tables', 0)))
        col += 3
style_data(ws_mx, 2, len(matrix) + 1, len(mx_headers))
for c in range(1, len(mx_headers) + 1):
    ws_mx.column_dimensions[chr(64 + c) if c < 27 else 'A'].width = 14

# Sheet 3: Excerpt Bank
ws_eb = wb2.create_sheet('Excerpt_Bank')
eb_headers = ['Theme', 'Segment ID', 'Source', 'Table', 'Speaker', 'Speaker Type', 'Question', 'Excerpt Text', 'Codes', 'Language']
for col, h in enumerate(eb_headers, 1):
    ws_eb.cell(row=1, column=col, value=h)
style_header(ws_eb, 1, len(eb_headers))
for i, e in enumerate(excerpts, 2):
    ws_eb.cell(row=i, column=1, value=e['theme'])
    ws_eb.cell(row=i, column=2, value=e['segment_id'])
    ws_eb.cell(row=i, column=3, value=e['source_file'])
    ws_eb.cell(row=i, column=4, value=e['table_id'])
    ws_eb.cell(row=i, column=5, value=e['speaker_label'])
    ws_eb.cell(row=i, column=6, value=e['speaker_type'])
    ws_eb.cell(row=i, column=7, value=e['question_id'])
    ws_eb.cell(row=i, column=8, value=e['excerpt_text'])
    ws_eb.cell(row=i, column=9, value=e['codes'])
    ws_eb.cell(row=i, column=10, value=e['language'])
style_data(ws_eb, 2, len(excerpts) + 1, len(eb_headers))
ws_eb.column_dimensions['A'].width = 40
ws_eb.column_dimensions['B'].width = 12
ws_eb.column_dimensions['C'].width = 16
ws_eb.column_dimensions['D'].width = 8
ws_eb.column_dimensions['E'].width = 35
ws_eb.column_dimensions['F'].width = 14
ws_eb.column_dimensions['G'].width = 10
ws_eb.column_dimensions['H'].width = 80
ws_eb.column_dimensions['I'].width = 40
ws_eb.column_dimensions['J'].width = 10

# Sheet 4: Prominence/Salience
ws_pr = wb2.create_sheet('Prominence_Salience')
pr_headers = ['Theme', 'Segments', 'Speakers', 'Tables', 'Chars', 'Questions', 'Composite', 'Salience', 'Explanation']
for col, h in enumerate(pr_headers, 1):
    ws_pr.cell(row=1, column=col, value=h)
style_header(ws_pr, 1, len(pr_headers))
for i, p in enumerate(prominence, 2):
    ws_pr.cell(row=i, column=1, value=p['theme'])
    ws_pr.cell(row=i, column=2, value=int(p['participant_segments']))
    ws_pr.cell(row=i, column=3, value=int(p['unique_speakers']))
    ws_pr.cell(row=i, column=4, value=int(p['unique_tables']))
    ws_pr.cell(row=i, column=5, value=int(p['total_chars']))
    ws_pr.cell(row=i, column=6, value=p['questions_present'])
    ws_pr.cell(row=i, column=7, value=int(p['composite_score']))
    ws_pr.cell(row=i, column=8, value=p['salience'])
    ws_pr.cell(row=i, column=9, value=p['salience_explanation'])
style_data(ws_pr, 2, len(prominence) + 1, len(pr_headers))
ws_pr.column_dimensions['A'].width = 45
ws_pr.column_dimensions['H'].width = 28
ws_pr.column_dimensions['I'].width = 80

wb2.save(d1_out / 'CASE_D1_theme_evidence_workbook.xlsx')
print('Theme evidence workbook saved.')
print('Both Excel workbooks created.')
