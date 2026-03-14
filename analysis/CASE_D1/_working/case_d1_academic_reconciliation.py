from __future__ import annotations

import ast
import csv
import shutil
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(r"c:\Users\baram\OneDrive\Desktop\themnatic")
CASE_DIR = ROOT / "analysis" / "CASE_D1"
WORKING_DIR = CASE_DIR / "_working"
OUTWARD_DIR = CASE_DIR / "OUTWARD_FACING_PACKAGE"
INTERNAL_DIR = CASE_DIR / "INTERNAL_CONFIDENTIAL"
LABEL_MAP_SOURCE = WORKING_DIR / "reconciliation_fix_v2.py"

THEME_CODES: dict[str, list[str]] = {
    "Theme_1_Balanced_Contentment": [
        "contentment_as_core",
        "balance_multidimensional",
        "spiritual_moral_anchor",
        "safety_security",
        "inner_peace",
        "pillar_physical",
        "pillar_emotional",
        "pillar_social",
        "pillar_intellectual",
        "pillar_spiritual",
    ],
    "Theme_2_Care_Ecology": [
        "interdependent_care",
        "school_clinical_link",
        "family_support_ecology",
        "community_volunteer_bridge",
        "professional_identity",
        "service_recipient_voice",
    ],
    "Theme_3_Service_Fragmentation": [
        "system_disconnection",
        "training_deficit",
        "resource_strain",
        "healthcare_system_critique",
        "consultation_time_pressure",
        "staffing_shortage_isolation",
        "medical_model_dominance",
        "missing_stepdown_services",
        "bureaucratic_barrier",
        "child_vulnerability",
        "child_mental_health",
        "bullying_violence",
        "child_rights_inclusion",
    ],
    "Theme_4_Culturally_Grounded_Solutions": [
        "cultural_local_adaptation",
        "transcultural_workforce",
        "awareness_education",
        "early_detection_prevention",
        "practical_recommendation",
        "top_down_systemic_change",
        "healthcare_worker_wellbeing",
        "gradual_implementation",
        "digital_platform_proposal",
        "respite_inclusive_resources",
    ],
}

THEME_LABELS = {
    "Theme_1_Balanced_Contentment": "Balanced contentment, safety, and moral steadiness",
    "Theme_2_Care_Ecology": "Interdependent professional support ecology",
    "Theme_3_Service_Fragmentation": "Fragmented, strained childhood service conditions",
    "Theme_4_Culturally_Grounded_Solutions": "Culturally grounded coordination and implementation",
}

REGISTER_ID_MAP: dict[str, tuple[str, str, str]] = {
    "D1_T0_MOD_Kulood": ("D1_M01", "moderator", "Moderator, Table 0"),
    "D1_T0_P1_DrAhmed": ("D1_P01", "participant", "Psychiatrist, Table 0"),
    "D1_T0_P2_Sp4": ("D1_P02", "participant", "Healthcare professional, Table 0"),
    "D1_T0_P3_Sp5": ("D1_P03", "participant", "Mental health professional, Table 0"),
    "D1_T0_P4_Sp7": ("D1_P04", "participant", "Wellbeing researcher, Table 0"),
    "D1_T0_P5_Sp6": ("D1_P05", "participant", "Healthcare professional, Table 0"),
    "D1_T0_P6_Sp2": ("D1_U01", "unclear", "Unclear speaker identity, Table 0"),
    "D1_T2_P1_DrHessa": ("D1_P06", "participant", "Healthcare professional, Table 2"),
    "D1_T2_P2_Mudakhala": ("D1_P07", "participant", "Healthcare professional, Table 2"),
    "D1_T2_MOD_Hanadi": ("D1_M02", "moderator", "Moderator, Table 2"),
    "D1_T3_MOD_Hala": ("D1_M03", "moderator", "Moderator, Table 3"),
    "D1_T4_MOD_SessionMgr": ("D1_M04", "moderator", "Moderator, Table 4"),
    "D1_T4_P1_DrYaseen": ("D1_P08", "participant", "Healthcare professional, Table 4"),
    "D1_T4_P2_Ali": ("D1_P09", "participant", "Service recipient, Table 4"),
    "D1_T6_MOD_Nour": ("D1_M05", "moderator", "Moderator, Table 6"),
    "D1_T6_P1_Manal": ("D1_P10", "participant", "Psychologist, Table 6"),
    "D1_T6_P2_Hafsa": ("D1_P11", "participant", "Community volunteer, Table 6"),
    "D1_T6_P3_HamadPR": ("D1_P12", "participant", "Hospital public-relations manager, Table 6"),
    "D1_T6_P4_DrAmna": ("D1_P13", "participant", "Paediatrician, Table 6"),
    "D1_T6_P5_DrAmal": ("D1_P14", "participant", "Paediatrician, Table 6"),
    "D1_T6_P6_Jumana": ("D1_U02", "unclear", "Unclear speaker identity, Table 6"),
    "D1_T7_MOD_SessionMgr": ("D1_M06", "moderator", "Moderator, Table 7"),
    "D1_T7_P1_Alaa": ("D1_P15", "participant", "Healthcare professional, Table 7"),
    "D1_T7_P2_DrWael": ("D1_P16", "participant", "Healthcare professional, Table 7"),
    "D1_T7_P3_Fatma": ("D1_P17", "participant", "Healthcare professional, Table 7"),
    "D1_T7_P4_Shafaq": ("D1_P18", "participant", "Healthcare professional, Table 7"),
    "D1_T7_P5_Amani": ("D1_P19", "participant", "Healthcare professional, Table 7"),
    "D1_T7_P6_Noura": ("D1_P20", "participant", "Healthcare professional, Table 7"),
    "D1_T7_P7_Latifa": ("D1_P21", "participant", "Healthcare professional, Table 7"),
    "D1_T10_MOD1": ("D1_M07", "moderator", "Moderator, Table 10"),
    "D1_T10_MOD2": ("D1_M08", "moderator", "Moderator, Table 10"),
    "D1_T10_P1_Muhammad": ("D1_P22", "participant", "Family consultant, Table 10"),
    "D1_T10_P2_Melissa": ("D1_P23", "participant", "Occupational therapist, Table 10"),
    "D1_T10_P3_Nawal": ("D1_P24", "participant", "Resident doctor, Table 10"),
    "D1_T10_P4_DrAfaf": ("D1_P25", "participant", "Psychologist, Table 10"),
    "D1_T10_P5_Angela": ("D1_P26", "participant", "Social worker, Table 10"),
    "D1_T10_P6_DrKhalid": ("D1_P27", "participant", "Child psychiatrist, Table 10"),
}

REAL_NAME_BY_CODE: dict[str, str] = {
    "D1_M01": "Kulood / Dr. Kholoud",
    "D1_M02": "Hanadi Ahmad Abu Bakr",
    "D1_M03": "Hala Fayed Hassan Fathi",
    "D1_M04": "Abdulkarim Shaaban",
    "D1_M05": "Nour Ahmad Al-Watadi",
    "D1_M06": "Abla Ahmad Khalil",
    "D1_M07": "Prof. Arokiasamy Penanayagam",
    "D1_M08": "Moderator 2 (unnamed)",
    "D1_P01": "Dr Ahmed Al Emadi",
    "D1_P02": "Speaker 4",
    "D1_P03": "Speaker 5",
    "D1_P04": "Speaker 7",
    "D1_P05": "Speaker 6",
    "D1_P06": "Dr. Hessa / د. حصة",
    "D1_P07": "مداخلة / intervention label",
    "D1_P08": "Dr. Yaseen",
    "D1_P09": "Ali (service recipient)",
    "D1_P10": "Manal",
    "D1_P11": "Hafsa Abdullah",
    "D1_P12": "PR Manager (Hamad Hospital)",
    "D1_P13": "Dr. Amna",
    "D1_P14": "Dr. Amal",
    "D1_P15": "Alaa Karmala",
    "D1_P16": "Dr. Wael Mahmoud",
    "D1_P17": "Fatma Al-Obaidan",
    "D1_P18": "Shafaq Al-Khalidi",
    "D1_P19": "Amani Al-Yafei",
    "D1_P20": "Noura",
    "D1_P21": "Latifa Al-Sulaiti",
    "D1_P22": "Muhammad Ali",
    "D1_P23": "Melissa Toon",
    "D1_P24": "Nawal Yosul",
    "D1_P25": "Dr. Afaf Asuhou",
    "D1_P26": "Angela Lwage",
    "D1_P27": "Dr. Khalid",
    "D1_U01": "Speaker 2",
    "D1_U02": "Jumana",
}

TEXT_REPLACEMENTS = {
    "Sarah Aicha Belouidiane- Clinical Psychologist- PHCC- HC Al Thumama": "clinical psychologist participant",
    "Sarah Aicha Belouidiane- Clinic": "clinical psychologist participant",
    "Ms. Maleeha Khamis M A Alshamali- Admin- Mental Health Service -HMC": "mental health service administrator participant",
    "Maleeha Khamis M A Alshamali- Admin- Mental Health Service -HMC": "mental health service administrator participant",
    "Emad Kora- Dietitian- KHR HC": "dietitian participant",
    "Amale Antoine Issa- Specialist Family Medicine- GHW HC": "family medicine specialist participant",
    "Amale Ant": "family medicine specialist participant",
    "Professor Abdullah ALkhater": "a professor",
    "professor Abdullah ALkhater": "a professor",
    "Professor Abdullah Akhater": "a professor",
    "professor Abdullah Akhater": "a professor",
    "Abdullah ALkhater": "a professor",
    "Abdullah Akhater": "a professor",
    "k Abdullah": "a",
    "Dr. A Lee": "a colleague",
    "Dr. Marge": "a colleague",
    "Dr. Manger": "a colleague",
    "Shahrazad": "a colleague",
    "الدكتورة أمال": "participant",
    "الدكتورة امال": "participant",
    "أمال": "participant",
    "د. أمال": "participant",
    "دكتورة امل": "participant",
    "د. امال": "participant",
    "أختي نورة": "participant",
    "نورة": "participant",
    "فاطمة": "participant",
    "دكتور وائل": "participant",
    "شكرا صافية": "شكرا participant",
    "صافية participant": "participant",
    "أ. صافية": "participant",
    "أ/ صافية": "participant",
    "أ/ حسين": "participant",
    "أ. حسين": "participant",
    "د/ دانية": "participant",
    "دكتورة دانية": "participant",
    "د. دانية": "participant",
    "أستاذة نبراس": "participant",
    "أستاذة/ نبراس": "participant",
    "أ/ نبراس": "participant",
    "الأستاذة لولوا": "participant",
    "الأستاذة ايناس": "participant",
    "الأستاذة إيناس": "participant",
    "أستاذة صافية": "participant",
    "مثلما ذكرت سارة": "مثلما ذكرت participant",
    "مثلما ذكرت مليحة": "مثلما ذكرت participant",
}

NOTE_TAKER_ROWS: list[dict[str, str]] = [
    {
        "evidence_id": "D1_E001",
        "theme": "Theme_1_Balanced_Contentment",
        "segment_id": "",
        "source_file": "HWCH0NT1.docx",
        "table_id": "0",
        "speaker_code": "",
        "speaker_type": "note_taker_summary",
        "role_label": "Note-taker summary",
        "attribution_status": "note_taker_summary",
        "evidence_type": "note_taker_summary",
        "question_id": "Q1",
        "excerpt_text": "My wellbeing is a sense of being content, not happy or sad, just content.",
        "codes": "contentment_as_core",
        "language": "en",
        "report_use": "Q1 definition of wellbeing",
    },
    {
        "evidence_id": "D1_E002",
        "theme": "Theme_1_Balanced_Contentment",
        "segment_id": "",
        "source_file": "HWCH6NT3.docx",
        "table_id": "6",
        "speaker_code": "",
        "speaker_type": "note_taker_summary",
        "role_label": "Note-taker summary",
        "attribution_status": "note_taker_summary",
        "evidence_type": "note_taker_summary",
        "question_id": "Q2",
        "excerpt_text": "حالة تكاملية من التوازن بين جميع الركائز (الروحية / الجسدية / الاجتماعية / العاطفية / الفكرية)",
        "codes": "balance_multidimensional",
        "language": "ar",
        "report_use": "Q2 definition of wellness",
    },
    {
        "evidence_id": "D1_E003",
        "theme": "Theme_1_Balanced_Contentment",
        "segment_id": "",
        "source_file": "HWCH0NT1.docx",
        "table_id": "0",
        "speaker_code": "",
        "speaker_type": "note_taker_summary",
        "role_label": "Note-taker summary",
        "attribution_status": "note_taker_summary",
        "evidence_type": "note_taker_summary",
        "question_id": "Q2",
        "excerpt_text": "I haven't thought about the pillars until we sat today.",
        "codes": "pillar_spiritual",
        "language": "en",
        "report_use": "Q2 pillar-familiarity limitation",
    },
]


class ReconciliationError(RuntimeError):
    """Raised when the CASE_D1 academic reconciliation cannot proceed safely."""


def load_csv(path: Path) -> list[dict[str, str]]:
    """Load a UTF-8 CSV file into memory."""
    with open(path, "r", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    """Write rows to a CSV file using a fixed schema."""
    with open(path, "w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def load_label_map(path: Path) -> dict[str, tuple[str, str, str]]:
    """Parse the LABEL_MAP literal from the prior reconciliation script."""
    tree = ast.parse(path.read_text(encoding="utf-8"))
    for node in tree.body:
        if isinstance(node, ast.Assign):
            for target in node.targets:
                if isinstance(target, ast.Name) and target.id == "LABEL_MAP":
                    value = ast.literal_eval(node.value)
                    if not isinstance(value, dict):
                        raise ReconciliationError("LABEL_MAP did not parse as a dictionary.")
                    return value
    raise ReconciliationError("Could not locate LABEL_MAP in reconciliation_fix_v2.py.")


def normalize_ws_text(text: str) -> str:
    """Normalize whitespace in narrative text exports."""
    return " ".join(text.split())


def build_dynamic_replacements(label_map: dict[str, tuple[str, str, str]]) -> list[tuple[str, str]]:
    """Build deterministic text-replacement pairs for anonymization inside exported text fields."""
    replacements: list[tuple[str, str]] = []
    for label, (code, speaker_type, role_label) in sorted(label_map.items(), key=lambda item: len(item[0]), reverse=True):
        if label == "_UNLABELED_":
            continue
        if code in {"D1_U03", "D1_U04", "D1_UNK"}:
            replacement = "participant"
        elif speaker_type == "moderator":
            replacement = code
        else:
            replacement = code
        replacements.append((label, replacement))
    for key, value in sorted(TEXT_REPLACEMENTS.items(), key=lambda item: len(item[0]), reverse=True):
        replacements.append((key, value))
    return replacements


def sanitize_text(text: str, replacements: list[tuple[str, str]]) -> str:
    """Remove real-name leakage from exported evidence text."""
    cleaned = text
    if " NOTE " in cleaned:
        cleaned = cleaned.split(" NOTE ", 1)[0].strip()
    for old, new in replacements:
        cleaned = cleaned.replace(old, new)
    cleaned = cleaned.replace("Dr. Alaf AsuhouPsychologist Sidra", "participant")
    cleaned = cleaned.replace("Dr. Khaled", "D1_P27")
    return normalize_ws_text(cleaned)


def speaker_metadata(
    speaker_label: str,
    label_map: dict[str, tuple[str, str, str]],
) -> tuple[str, str, str, str]:
    """Return outward-facing speaker metadata for a coded segment row."""
    if speaker_label not in label_map:
        raise ReconciliationError(f"Unmapped speaker label in coded segments: {speaker_label!r}")
    code, mapped_type, role_label = label_map[speaker_label]
    attribution_note = "identified"
    if code in {"D1_UNK", "D1_U03", "D1_U04"}:
        code = ""
        mapped_type = "unclear"
        role_label = "Attribution indeterminate"
        attribution_note = "indeterminate"
    return code, mapped_type, role_label, attribution_note


def build_classification_basis(speaker_type: str, role_label: str) -> str:
    """Create a safe outward-facing classification note without identity leakage."""
    if speaker_type == "moderator":
        return "Classified as moderator from facilitation pattern and moderator-assignment review."
    if speaker_type == "unclear":
        return "Attribution remained indeterminate after review; retained as unclear because role could not be confirmed from available labels."
    if role_label == "Service recipient, Table 4":
        return "Classified as participant because the source explicitly indicates a service-recipient role and the contribution is substantive."
    return "Classified as participant after review of substantive contribution pattern and source-role context."


def build_outward_participant_register(
    original_register: list[dict[str, str]],
    replacements: list[tuple[str, str]],
) -> list[dict[str, str]]:
    """Create the outward-facing anonymized participant register."""
    rows: list[dict[str, str]] = []
    for row in original_register:
        if "participant_id" in row:
            participant_id = row["participant_id"]
            if participant_id not in REGISTER_ID_MAP:
                raise ReconciliationError(f"No register mapping for {participant_id}")
            code, speaker_type, role_label = REGISTER_ID_MAP[participant_id]
            classification_basis = build_classification_basis(speaker_type, role_label)
            turns = row["turns"]
            chars = row["chars"]
            source_file = row["source_file"]
            table_id = row["table_id"]
        else:
            code = row["anonymized_code"]
            speaker_type = row["speaker_type"]
            role_label = row["role_label"]
            classification_basis = build_classification_basis(speaker_type, role_label)
            turns = row["turns"]
            chars = row["chars"]
            source_file = row["source_file"]
            table_id = row["table_id"]
        rows.append(
            {
                "anonymized_code": code,
                "source_file": source_file,
                "table_id": table_id,
                "speaker_type": speaker_type,
                "role_label": role_label,
                "turns": turns,
                "chars": chars,
                "classification_basis": classification_basis,
            }
        )
    return rows


def build_clean_segments(
    raw_segments: list[dict[str, str]],
    label_map: dict[str, tuple[str, str, str]],
    replacements: list[tuple[str, str]],
) -> list[dict[str, str]]:
    """Create an outward-facing coded-segment export with stable anonymization only."""
    clean_rows: list[dict[str, str]] = []
    for row in raw_segments:
        if "speaker_label" in row:
            code, speaker_type, role_label, attribution_note = speaker_metadata(row["speaker_label"], label_map)
        else:
            code = row.get("speaker_code", "")
            speaker_type = row["speaker_type"]
            role_label = row["role_label"]
            attribution_note = row.get("attribution_status", "identified")
        clean_rows.append(
            {
                "segment_id": row["segment_id"],
                "source_file": row["source_file"],
                "table_id": row["table_id"],
                "speaker_code": code,
                "speaker_type": speaker_type,
                "role_label": role_label,
                "attribution_status": attribution_note,
                "question_id": row["question_id"],
                "segment_text": sanitize_text(row["segment_text"], replacements),
                "codes": row["codes"],
                "language": row["language"],
            }
        )
    return clean_rows


def classify_evidence_type(source_file: str, text: str) -> str:
    """Assign one truthful evidence type to each quotation row."""
    note_style_markers = (
        "Described ",
        "Noted ",
        "Called for ",
        "Argued ",
        "Highlighted ",
        "Identified ",
        "Introduced ",
        "Proposed ",
        "Recommended ",
        "Provided ",
        "Recounted ",
        "Suggested ",
        "Emphasized ",
        "Critiqued ",
        "Shared ",
        "Speaking as ",
    )
    if "NT" in source_file:
        return "note_taker_summary"
    if "Health_Workshop_Suggestions Day 1.xlsx" in source_file:
        return "auxiliary_recommendation_reference"
    if source_file == "HWCH10AR.docx" and text.startswith(note_style_markers):
        return "note_style_transcript_summary"
    return "verbatim_transcript"


def build_excerpt_bank(
    raw_excerpt_rows: list[dict[str, str]],
    segment_lookup: dict[str, dict[str, str]],
    label_map: dict[str, tuple[str, str, str]],
    replacements: list[tuple[str, str]],
) -> list[dict[str, str]]:
    """Create the outward-facing excerpt bank with truthful evidence-type coverage."""
    rows: list[dict[str, str]] = []
    for row in raw_excerpt_rows:
        segment_id = row["segment_id"]
        if not segment_id:
            clean_text = sanitize_text(row["excerpt_text"], replacements)
            rows.append(
                {
                    "evidence_id": row["evidence_id"],
                    "theme": row["theme"],
                    "segment_id": "",
                    "source_file": row["source_file"],
                    "table_id": row["table_id"],
                    "speaker_code": row.get("speaker_code", ""),
                    "speaker_type": row.get("speaker_type", "note_taker_summary"),
                    "role_label": row.get("role_label", "Note-taker summary"),
                    "attribution_status": row.get("attribution_status") or "note_taker_summary",
                    "evidence_type": "note_taker_summary",
                    "question_id": row["question_id"],
                    "excerpt_text": clean_text,
                    "codes": row["codes"],
                    "language": row["language"],
                    "report_use": row.get("report_use", "theme_evidence"),
                }
            )
            continue
        if segment_id not in segment_lookup:
            raise ReconciliationError(f"Excerpt bank segment missing from coded segments: {segment_id}")
        original_segment = segment_lookup[segment_id]
        if "speaker_label" in original_segment:
            code, speaker_type, role_label, attribution_note = speaker_metadata(original_segment["speaker_label"], label_map)
        else:
            code = original_segment.get("speaker_code", "")
            speaker_type = original_segment["speaker_type"]
            role_label = original_segment["role_label"]
            attribution_note = original_segment.get("attribution_status", "identified")
        clean_text = sanitize_text(row["excerpt_text"], replacements)
        rows.append(
            {
                "evidence_id": segment_id,
                "theme": row["theme"],
                "segment_id": segment_id,
                "source_file": row["source_file"],
                "table_id": row["table_id"],
                "speaker_code": code,
                "speaker_type": speaker_type,
                "role_label": role_label,
                "attribution_status": attribution_note,
                "evidence_type": classify_evidence_type(row["source_file"], clean_text),
                "question_id": row["question_id"],
                "excerpt_text": clean_text,
                "codes": row["codes"],
                "language": row["language"],
                "report_use": "theme_evidence",
            }
        )
    existing_evidence_ids = {row["evidence_id"] for row in rows}
    for note_row in NOTE_TAKER_ROWS:
        if note_row["evidence_id"] not in existing_evidence_ids:
            rows.append(note_row)
    return rows


def build_participant_summary(
    clean_segments: list[dict[str, str]],
    outward_register: list[dict[str, str]],
) -> list[dict[str, str]]:
    """Aggregate per-speaker summaries from the cleaned coded segments and register."""
    grouped: dict[tuple[str, str, str, str], dict[str, Any]] = defaultdict(
        lambda: {"segment_count": 0, "total_chars": 0, "questions": set(), "codes": Counter()}
    )
    for row in outward_register:
        grouped[(row["anonymized_code"], row["source_file"], row["table_id"], row["speaker_type"])]
    for row in clean_segments:
        speaker_code = row["speaker_code"]
        if not speaker_code:
            continue
        key = (speaker_code, row["source_file"], row["table_id"], row["speaker_type"])
        grouped[key]["segment_count"] += 1
        grouped[key]["total_chars"] += len(row["segment_text"])
        grouped[key]["questions"].add(row["question_id"])
        for code in row["codes"].split(";"):
            if code:
                grouped[key]["codes"][code] += 1
    summary_rows: list[dict[str, str]] = []
    for (speaker_code, source_file, table_id, speaker_type), payload in sorted(grouped.items()):
        summary_rows.append(
            {
                "anonymized_code": speaker_code,
                "source_file": source_file,
                "table_id": table_id,
                "speaker_type": speaker_type,
                "segment_count": str(payload["segment_count"]),
                "total_chars": str(payload["total_chars"]),
                "questions_covered": ";".join(sorted(payload["questions"])),
                "top_codes": ";".join(code for code, _ in payload["codes"].most_common(5)),
            }
        )
    return summary_rows


def build_source_contribution(clean_segments: list[dict[str, str]]) -> list[dict[str, str]]:
    """Recompute source contribution from the corrected coded segments."""
    buckets: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"total": 0, "participant": 0, "speakers": set(), "questions": set(), "themes": set()}
    )
    for row in clean_segments:
        bucket = buckets[row["source_file"]]
        bucket["total"] += 1
        bucket["questions"].add(row["question_id"])
        if row["speaker_type"] == "participant":
            bucket["participant"] += 1
            if row["speaker_code"]:
                bucket["speakers"].add(row["speaker_code"])
            code_set = set(row["codes"].split(";"))
            for theme, theme_codes in THEME_CODES.items():
                if code_set & set(theme_codes):
                    bucket["themes"].add(theme)
    return [
        {
            "source_file": source,
            "total_segments": str(payload["total"]),
            "participant_segments": str(payload["participant"]),
            "unique_speakers": str(len(payload["speakers"])),
            "questions_covered": ";".join(sorted(payload["questions"])),
            "themes_present": ";".join(sorted(payload["themes"])),
        }
        for source, payload in sorted(buckets.items())
    ]


def build_question_theme_matrix(clean_segments: list[dict[str, str]]) -> list[dict[str, str]]:
    """Recompute the question-by-theme matrix using corrected participant typing."""
    participant_rows = [row for row in clean_segments if row["speaker_type"] == "participant"]
    questions = sorted({row["question_id"] for row in clean_segments}, key=lambda value: int(value[1:]))
    matrix_rows: list[dict[str, str]] = []
    for question_id in questions:
        question_rows = [row for row in participant_rows if row["question_id"] == question_id]
        out: dict[str, str] = {"question_id": question_id}
        for theme, theme_codes in THEME_CODES.items():
            matches = [row for row in question_rows if set(row["codes"].split(";")) & set(theme_codes)]
            out[f"{theme}_segments"] = str(len(matches))
            out[f"{theme}_speakers"] = str(len({row['speaker_code'] for row in matches if row['speaker_code']}))
            out[f"{theme}_tables"] = str(len({row['table_id'] for row in matches}))
        matrix_rows.append(out)
    return matrix_rows


def build_question_evidence(clean_segments: list[dict[str, str]]) -> list[dict[str, str]]:
    """Recompute question-level evidence counts using corrected speaker typing."""
    questions = sorted({row["question_id"] for row in clean_segments}, key=lambda value: int(value[1:]))
    rows: list[dict[str, str]] = []
    for question_id in questions:
        qrows = [row for row in clean_segments if row["question_id"] == question_id]
        participant_rows = [row for row in qrows if row["speaker_type"] == "participant"]
        code_counter: Counter[str] = Counter()
        for row in participant_rows:
            for code in row["codes"].split(";"):
                if code:
                    code_counter[code] += 1
        rows.append(
            {
                "question_id": question_id,
                "participant_segments": str(sum(1 for row in qrows if row["speaker_type"] == "participant")),
                "moderator_segments": str(sum(1 for row in qrows if row["speaker_type"] == "moderator")),
                "unclear_segments": str(sum(1 for row in qrows if row["speaker_type"] == "unclear")),
                "unique_participant_speakers": str(len({row['speaker_code'] for row in participant_rows if row['speaker_code']})),
                "source_files": str(len({row['source_file'] for row in participant_rows})),
                "top_codes": ";".join(code for code, _ in code_counter.most_common(5)),
            }
        )
    return rows


def build_prominence(clean_segments: list[dict[str, str]]) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    """Recompute theme prominence and the derived summary table."""
    participant_rows = [row for row in clean_segments if row["speaker_type"] == "participant"]
    theme_data: dict[str, dict[str, Any]] = {}
    for theme, theme_codes in THEME_CODES.items():
        rows = [row for row in participant_rows if set(row["codes"].split(";")) & set(theme_codes)]
        theme_data[theme] = {
            "segments": len(rows),
            "speakers": len({row["speaker_code"] for row in rows if row["speaker_code"]}),
            "tables": len({row["table_id"] for row in rows}),
            "chars": sum(len(row["segment_text"]) for row in rows),
            "questions": sorted({row["question_id"] for row in rows}),
        }
        theme_data[theme]["composite"] = (
            theme_data[theme]["segments"]
            + (theme_data[theme]["speakers"] * 3)
            + (theme_data[theme]["tables"] * 5)
            + (len(theme_data[theme]["questions"]) * 2)
        )
    ranked = sorted(theme_data.items(), key=lambda item: -item[1]["composite"])
    salience_labels = [
        "most_prominent",
        "highly_prominent",
        "moderately_prominent",
        "present_but_less_prominent",
    ]
    prominence_rows: list[dict[str, str]] = []
    for index, (theme, payload) in enumerate(ranked):
        salience = salience_labels[min(index, len(salience_labels) - 1)]
        prominence_rows.append(
            {
                "theme": theme,
                "participant_segments": str(payload["segments"]),
                "unique_speakers": str(payload["speakers"]),
                "unique_tables": str(payload["tables"]),
                "total_chars": str(payload["chars"]),
                "questions_present": ";".join(payload["questions"]),
                "composite_score": str(payload["composite"]),
                "salience": salience,
                "salience_explanation": (
                    f"{payload['segments']} participant-coded segments from {payload['speakers']} speakers "
                    f"across {payload['tables']} tables covering {len(payload['questions'])} questions."
                ),
            }
        )
    theme_summary_rows = [
        {
            "theme_number": theme.split("_")[1],
            "theme_name": THEME_LABELS[theme],
            "participant_segments": row["participant_segments"],
            "unique_speakers": row["unique_speakers"],
            "unique_tables": row["unique_tables"],
            "questions_present": row["questions_present"],
            "salience": row["salience"],
        }
        for row in prominence_rows
        for theme in [row["theme"]]
    ]
    return prominence_rows, theme_summary_rows


def make_workbook_styles() -> tuple[Font, Alignment, Border, PatternFill, PatternFill]:
    """Return shared workbook styles."""
    header_font = Font(bold=True, size=11, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    blue_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
    green_fill = PatternFill(start_color="1E8449", end_color="1E8449", fill_type="solid")
    return header_font, wrap, border, blue_fill, green_fill


def build_participant_workbook(summary_rows: list[dict[str, str]], target: Path) -> None:
    """Write the outward-facing participant workbook."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Participant_Summary"
    header_font, wrap, border, blue_fill, _ = make_workbook_styles()
    headers = [
        "Anonymized Code",
        "Source",
        "Table",
        "Speaker Type",
        "Segments",
        "Chars",
        "Questions",
        "Top Codes",
    ]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = wrap
        cell.border = border
    for row_index, row in enumerate(summary_rows, start=2):
        values = [
            row["anonymized_code"],
            row["source_file"],
            row["table_id"],
            row["speaker_type"],
            int(row["segment_count"]),
            int(row["total_chars"]),
            row["questions_covered"],
            row["top_codes"],
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.alignment = wrap
            cell.border = border
    workbook.save(target)


def build_theme_evidence_workbook(
    excerpt_rows: list[dict[str, str]],
    matrix_rows: list[dict[str, str]],
    source_rows: list[dict[str, str]],
    prominence_rows: list[dict[str, str]],
    target: Path,
) -> None:
    """Write the outward-facing theme evidence workbook."""
    workbook = Workbook()
    header_font, wrap, border, blue_fill, green_fill = make_workbook_styles()
    matrix_sheet = workbook.active
    matrix_sheet.title = "Question_Theme_Matrix"
    matrix_headers = list(matrix_rows[0].keys())
    for column, header in enumerate(matrix_headers, start=1):
        cell = matrix_sheet.cell(row=1, column=column, value=header)
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = wrap
        cell.border = border
    for row_index, row in enumerate(matrix_rows, start=2):
        for column, header in enumerate(matrix_headers, start=1):
            cell = matrix_sheet.cell(row=row_index, column=column, value=row[header])
            cell.alignment = wrap
            cell.border = border

    excerpt_sheet = workbook.create_sheet("Excerpt_Bank")
    excerpt_headers = list(excerpt_rows[0].keys())
    for column, header in enumerate(excerpt_headers, start=1):
        cell = excerpt_sheet.cell(row=1, column=column, value=header)
        cell.font = header_font
        cell.fill = green_fill
        cell.alignment = wrap
        cell.border = border
    for row_index, row in enumerate(excerpt_rows, start=2):
        for column, header in enumerate(excerpt_headers, start=1):
            cell = excerpt_sheet.cell(row=row_index, column=column, value=row[header])
            cell.alignment = wrap
            cell.border = border

    source_sheet = workbook.create_sheet("Source_Contribution")
    source_headers = list(source_rows[0].keys())
    for column, header in enumerate(source_headers, start=1):
        cell = source_sheet.cell(row=1, column=column, value=header)
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = wrap
        cell.border = border
    for row_index, row in enumerate(source_rows, start=2):
        for column, header in enumerate(source_headers, start=1):
            cell = source_sheet.cell(row=row_index, column=column, value=row[header])
            cell.alignment = wrap
            cell.border = border

    prominence_sheet = workbook.create_sheet("Prominence")
    prominence_headers = list(prominence_rows[0].keys())
    for column, header in enumerate(prominence_headers, start=1):
        cell = prominence_sheet.cell(row=1, column=column, value=header)
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = wrap
        cell.border = border
    for row_index, row in enumerate(prominence_rows, start=2):
        for column, header in enumerate(prominence_headers, start=1):
            cell = prominence_sheet.cell(row=row_index, column=column, value=row[header])
            cell.alignment = wrap
            cell.border = border
    workbook.save(target)


def build_identity_key(
    outward_register: list[dict[str, str]],
    target: Path,
) -> None:
    """Write the internal/confidential identity key workbook."""
    workbook = Workbook()
    warning_sheet = workbook.active
    warning_sheet.title = "CONFIDENTIAL"
    warning_sheet["A1"] = "CONFIDENTIAL — INTERNAL USE ONLY"
    warning_sheet["A1"].font = Font(bold=True, size=16, color="FF0000")
    warning_sheet["A3"] = "This workbook contains identity traceability and must not be included in the outward-facing academic package."
    key_sheet = workbook.create_sheet("Identity_Key")
    header_font, wrap, border, blue_fill, _ = make_workbook_styles()
    headers = [
        "original_participant_id",
        "real_name_or_label",
        "anonymized_code",
        "speaker_type",
        "source_file",
        "table_id",
        "role_label",
        "classification_basis",
    ]
    for column, header in enumerate(headers, start=1):
        cell = key_sheet.cell(row=1, column=column, value=header)
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = wrap
        cell.border = border
    for row_index, row in enumerate(outward_register, start=2):
        code = row["anonymized_code"]
        values = [
            "",
            REAL_NAME_BY_CODE.get(code, ""),
            code,
            row["speaker_type"],
            row["source_file"],
            row["table_id"],
            row["role_label"],
            row["classification_basis"],
        ]
        for column, value in enumerate(values, start=1):
            cell = key_sheet.cell(row=row_index, column=column, value=value)
            cell.alignment = wrap
            cell.border = border
    workbook.save(target)


def copy_to_package(files: list[Path], package_dir: Path) -> None:
    """Copy revised files into the target package folder."""
    package_dir.mkdir(parents=True, exist_ok=True)
    for file_path in files:
        shutil.copy2(file_path, package_dir / file_path.name)


def main() -> None:
    """Run the strict CASE_D1 academic reconciliation pass."""
    OUTWARD_DIR.mkdir(parents=True, exist_ok=True)
    INTERNAL_DIR.mkdir(parents=True, exist_ok=True)

    label_map = load_label_map(LABEL_MAP_SOURCE)
    replacements = build_dynamic_replacements(label_map)

    original_register = load_csv(CASE_DIR / "CASE_D1_participant_register.csv")
    raw_segments = load_csv(CASE_DIR / "CASE_D1_coded_segments.csv")
    raw_excerpt_rows = load_csv(CASE_DIR / "CASE_D1_excerpt_bank.csv")

    clean_segments = build_clean_segments(raw_segments, label_map, replacements)
    segment_lookup = {row["segment_id"]: row for row in raw_segments}
    excerpt_rows = build_excerpt_bank(raw_excerpt_rows, segment_lookup, label_map, replacements)
    outward_register = build_outward_participant_register(original_register, replacements)
    summary_rows = build_participant_summary(clean_segments, outward_register)
    source_rows = build_source_contribution(clean_segments)
    matrix_rows = build_question_theme_matrix(clean_segments)
    evidence_rows = build_question_evidence(clean_segments)
    prominence_rows, theme_summary_rows = build_prominence(clean_segments)

    clean_segment_fields = [
        "segment_id",
        "source_file",
        "table_id",
        "speaker_code",
        "speaker_type",
        "role_label",
        "attribution_status",
        "question_id",
        "segment_text",
        "codes",
        "language",
    ]
    excerpt_fields = [
        "evidence_id",
        "theme",
        "segment_id",
        "source_file",
        "table_id",
        "speaker_code",
        "speaker_type",
        "role_label",
        "attribution_status",
        "evidence_type",
        "question_id",
        "excerpt_text",
        "codes",
        "language",
        "report_use",
    ]
    register_fields = [
        "anonymized_code",
        "source_file",
        "table_id",
        "speaker_type",
        "role_label",
        "turns",
        "chars",
        "classification_basis",
    ]
    summary_fields = [
        "anonymized_code",
        "source_file",
        "table_id",
        "speaker_type",
        "segment_count",
        "total_chars",
        "questions_covered",
        "top_codes",
    ]

    write_csv(CASE_DIR / "CASE_D1_coded_segments.csv", clean_segment_fields, clean_segments)
    write_csv(CASE_DIR / "CASE_D1_excerpt_bank.csv", excerpt_fields, excerpt_rows)
    write_csv(CASE_DIR / "CASE_D1_participant_register.csv", register_fields, outward_register)
    write_csv(CASE_DIR / "CASE_D1_participant_summary.csv", summary_fields, summary_rows)
    write_csv(CASE_DIR / "CASE_D1_source_contribution_table.csv", list(source_rows[0].keys()), source_rows)
    write_csv(CASE_DIR / "CASE_D1_question_theme_matrix.csv", list(matrix_rows[0].keys()), matrix_rows)
    write_csv(CASE_DIR / "CASE_D1_question_evidence_table.csv", list(evidence_rows[0].keys()), evidence_rows)
    write_csv(CASE_DIR / "CASE_D1_prominence_salience.csv", list(prominence_rows[0].keys()), prominence_rows)
    write_csv(CASE_DIR / "CASE_D1_theme_summary_table.csv", list(theme_summary_rows[0].keys()), theme_summary_rows)

    build_participant_workbook(summary_rows, CASE_DIR / "CASE_D1_participant_workbook.xlsx")
    build_theme_evidence_workbook(
        excerpt_rows,
        matrix_rows,
        source_rows,
        prominence_rows,
        CASE_DIR / "CASE_D1_theme_evidence_workbook.xlsx",
    )
    build_participant_workbook(summary_rows, CASE_DIR / "participant_summary_anonymized.xlsx")
    build_identity_key(outward_register, INTERNAL_DIR / "participant_identity_key.xlsx")

    shutil.copy2(CASE_DIR / "CASE_D1_participant_register.csv", INTERNAL_DIR / "CASE_D1_participant_register_anonymized.csv")
    shutil.copy2(CASE_DIR / "CASE_D1_participant_workbook.xlsx", OUTWARD_DIR / "CASE_D1_participant_workbook.xlsx")
    shutil.copy2(CASE_DIR / "CASE_D1_theme_evidence_workbook.xlsx", OUTWARD_DIR / "CASE_D1_theme_evidence_workbook.xlsx")
    shutil.copy2(CASE_DIR / "participant_summary_anonymized.xlsx", OUTWARD_DIR / "participant_summary_anonymized.xlsx")

    internal_files = [
        CASE_DIR / "CASE_D1_familiarisation_memo.md",
        CASE_DIR / "CASE_D1_preparation_checklist.md",
        CASE_DIR / "CASE_D1_source_sensitivity_memo.md",
    ]
    for file_path in internal_files:
        if file_path.exists():
            shutil.copy2(file_path, INTERNAL_DIR / file_path.name)
    shutil.copy2(CASE_DIR / "CASE_D1_participant_register.csv", OUTWARD_DIR / "CASE_D1_participant_register.csv")
    outward_files = [
        CASE_DIR / "CASE_D1_coded_segments.csv",
        CASE_DIR / "CASE_D1_excerpt_bank.csv",
        CASE_DIR / "CASE_D1_participant_register.csv",
        CASE_DIR / "CASE_D1_participant_summary.csv",
        CASE_DIR / "CASE_D1_candidate_themes.md",
        CASE_DIR / "CASE_D1_final_themes.md",
        CASE_DIR / "CASE_D1_final_report.md",
        CASE_DIR / "CASE_D1_crosscheck_report.md",
        CASE_DIR / "CASE_D1_source_contribution_table.csv",
        CASE_DIR / "CASE_D1_question_theme_matrix.csv",
        CASE_DIR / "CASE_D1_question_evidence_table.csv",
        CASE_DIR / "CASE_D1_prominence_salience.csv",
        CASE_DIR / "CASE_D1_theme_summary_table.csv",
    ]
    copy_to_package(outward_files, OUTWARD_DIR)

    counts = Counter(row["speaker_type"] for row in clean_segments)
    evidence_counts = Counter(row["evidence_type"] for row in excerpt_rows)
    print("CASE_D1 academic reconciliation data rebuild complete")
    print(f"speaker_type counts: {dict(counts)}")
    print(f"evidence_type counts: {dict(evidence_counts)}")
    print(f"outward package dir: {OUTWARD_DIR}")
    print(f"internal package dir: {INTERNAL_DIR}")


if __name__ == "__main__":
    main()
