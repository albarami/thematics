"""CASE_D1 Reconciliation Fix v2 — complete anonymization with proper Arabic label mapping."""
import csv
from pathlib import Path
from collections import Counter, defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("ERROR: openpyxl required"); exit(1)

root = Path(r'c:\Users\baram\OneDrive\Desktop\themnatic')
d1_out = root / 'analysis' / 'CASE_D1'

# ═══════════════════════════════════════════════════════
# COMPLETE LABEL → ANON MAPPING (all actual CSV labels)
# ═══════════════════════════════════════════════════════
LABEL_MAP = {
    # Table 0
    'Kulood': ('D1_M01', 'moderator', 'Moderator, Table 0'),
    'Dr Ahmed Al Emadi': ('D1_P01', 'participant', 'Psychiatrist, Table 0'),
    'Speaker 4': ('D1_P02', 'participant', 'Healthcare professional, Table 0'),
    'Speaker 5': ('D1_P03', 'participant', 'Mental health professional, Table 0'),
    'Speaker 7': ('D1_P04', 'participant', 'Wellbeing researcher, Table 0'),
    'Speaker 6': ('D1_P05', 'participant', 'Healthcare professional, Table 0'),
    'Speaker 2': ('D1_U01', 'unclear', 'Unknown role, Table 0'),
    # Table 2
    '\u0645\u062f\u0627\u062e\u0644\u0629': ('D1_P07', 'participant', 'Healthcare professional, Table 2'),  # مداخلة
    '\u0623\u0645\u062f\u0627\u062e\u0644\u0629': ('D1_P07', 'participant', 'Healthcare professional, Table 2'),  # أمداخلة (variant)
    '\u062f. \u062d\u0635\u0629': ('D1_P06', 'participant', 'Healthcare professional, Table 2'),  # د. حصة
    # Table 3
    '\u0637\u064a\u0628\u060c \u062f\u0639\u0646\u0627 \u0646\u0646\u062a\u0642\u0644 \u0644\u0644\u0633\u0624\u0627\u0644 \u0627\u0644\u0644\u064a \u0628\u0639\u062f\u0647\u060c \u0627\u0644\u0633\u0624\u0627\u0644 \u0627\u0644\u0631\u0627\u0628\u0639': ('D1_M03', 'moderator', 'Moderator, Table 3'),  # طيب، دعنا ننتقل...
    # Table 4
    '\u0645\u062f\u064a\u0631 \u0627\u0644\u062c\u0644\u0633\u0629': ('D1_M04', 'moderator', 'Moderator, Table 4'),  # مدير الجلسة
    '\u062f.\u064a\u0627\u0633\u064a\u0646': ('D1_P08', 'participant', 'Healthcare professional, Table 4'),  # د.ياسين
    '\u0623.\u0639\u0644\u064a (\u0645\u062a\u0644\u0642\u064a \u062e\u062f\u0645\u0629)': ('D1_P09', 'participant', 'Service recipient, Table 4'),  # أ.علي (متلقي خدمة)
    '\u0639\u0644\u064a (\u0645\u062a\u0644\u0642\u064a \u062e\u062f\u0645\u0629)': ('D1_P09', 'participant', 'Service recipient, Table 4'),  # علي (متلقي خدمة)
    '\u0639\u0644\u064a(\u0645\u062a\u0644\u0642\u064a \u062e\u062f\u0645\u0629)': ('D1_P09', 'participant', 'Service recipient, Table 4'),  # علي(متلقي خدمة)
    # Table 6
    '\u0646\u0648\u0631 \u0627\u0644\u0648\u062a\u0627\u0631\u064a': ('D1_M05', 'moderator', 'Moderator, Table 6'),  # نور الوتاري
    '( \u0645\u0646\u0627\u0644 \u0627\u062e\u0635\u0627\u0626\u064a\u0629 \u0646\u0641\u0633\u064a\u0629 \u0645\u0633\u062a\u0634\u0641\u0649 \u0627\u0644\u0633\u062f\u0631\u0629)': ('D1_P10', 'participant', 'Psychologist, Table 6'),
    '(\u0627\u062e\u0635\u0627\u0626\u064a\u0629 \u0646\u0641\u0633\u064a\u0629 \u0645\u0633\u062a\u0634\u0641\u0649 \u0627\u0644\u0633\u062f\u0631\u0629)': ('D1_P10', 'participant', 'Psychologist, Table 6'),
    '.( \u0645\u0646\u0627\u0644 \u0627\u062e\u0635\u0627\u0626\u064a\u0629 \u0646\u0641\u0633\u064a\u0629 \u0645\u0633\u062a\u0634\u0641\u0649 \u0627\u0644\u0633\u062f\u0631\u0629)': ('D1_P10', 'participant', 'Psychologist, Table 6'),
    '. ( \u0645\u0646\u0627\u0644 \u0627\u062e\u0635\u0627\u0626\u064a\u0629 \u0646\u0641\u0633\u064a\u0629 \u0645\u0633\u062a\u0634\u0641\u0649 \u0627\u0644\u0633\u062f\u0631\u0629)': ('D1_P10', 'participant', 'Psychologist, Table 6'),
    '(\u0645\u062a\u0637\u0648\u0639\u0629 \u062d\u0641\u0635\u0629 \u0639\u0628\u062f\u0627\u0644\u0644\u0647)': ('D1_P11', 'participant', 'Community volunteer, Table 6'),
    '\u0645\u062a\u0637\u0648\u0639\u0629 \u062d\u0641\u0635\u0629 \u0639\u0628\u062f\u0627\u0644\u0644\u0647': ('D1_P11', 'participant', 'Community volunteer, Table 6'),
    '(\u0645\u062f\u064a\u0631\u0629 \u0639\u0644\u0627\u0642\u0627\u062a \u0645\u0633\u062a\u0634\u0641\u0649 \u062d\u0645\u062f)': ('D1_P12', 'participant', 'Hospital PR manager, Table 6'),
    '(\u062f \u0627\u0645\u0646\u0629 \u0637\u0628\u064a\u0628\u0629 \u0623\u0637\u0641\u0627\u0644)': ('D1_P13', 'participant', 'Paediatrician, Table 6'),
    '. (\u062f\u0643\u062a\u0648\u0631\u0629 \u0627\u0645\u0644 \u0637\u0628\u064a\u0628\u0629 \u0623\u0637\u0641\u0627\u0644)': ('D1_P14', 'participant', 'Paediatrician, Table 6'),
    '\u062c\u0645\u0627\u0646\u0629': ('D1_U02', 'unclear', 'Unknown role, Table 6'),  # جمانة
    # Table 7
    '\u0645\u062f\u064a\u0631 \u0627\u0644\u062c\u0644\u0633\u0647': ('D1_M06', 'moderator', 'Moderator, Table 7'),  # مدير الجلسه
    '\u0627\u0644\u0627\u0621 \u0643\u0631\u0645\u0627\u0644\u0647': ('D1_P15', 'participant', 'Healthcare professional, Table 7'),  # الاء كرماله
    '\u062f/\u0648\u0627\u0626\u0644 \u0645\u062d\u0645\u0648\u062f': ('D1_P16', 'participant', 'Healthcare professional, Table 7'),  # د/وائل محمود
    '\u0641\u0627\u0637\u0645\u0647 \u0627\u0644\u0639\u0628\u064a\u062f\u0627\u0646': ('D1_P17', 'participant', 'Healthcare professional, Table 7'),  # فاطمه العبيدان
    '\u0634\u0641\u0642 \u0627\u0644\u062e\u0627\u0644\u062f\u0649': ('D1_P18', 'participant', 'Healthcare professional, Table 7'),  # شفق الخالدى
    '\u0627\u0645\u0627\u0646\u064a \u0627\u0644\u064a\u0627\u0641\u0639\u0649': ('D1_P19', 'participant', 'Healthcare professional, Table 7'),  # اماني اليافعى
    '\u0646\u0648\u0631\u0627': ('D1_P20', 'participant', 'Healthcare professional, Table 7'),  # نورا
    '\u0644\u0637\u064a\u0641\u0647 \u0627\u0644\u0633\u0644\u064a\u0637\u0649': ('D1_P21', 'participant', 'Healthcare professional, Table 7'),  # لطيفه السليطى
    # Table 10
    'Moderator 1': ('D1_M07', 'moderator', 'Moderator, Table 10'),
    'Moderator 1 (Mental/Spiritual)': ('D1_M07', 'moderator', 'Moderator, Table 10'),
    'Moderator 2': ('D1_M08', 'moderator', 'Moderator, Table 10'),
    'Dr. Afaf Asuhou': ('D1_P25', 'participant', 'Psychologist, Table 10'),
    'Dr. Afaf Asuhou (Physical)': ('D1_P25', 'participant', 'Psychologist, Table 10'),
    'Melissa Toon': ('D1_P23', 'participant', 'Occupational therapist, Table 10'),
    'Melissa Toon (Integrated)': ('D1_P23', 'participant', 'Occupational therapist, Table 10'),
    'Melissa Toon (Challenge)': ('D1_P23', 'participant', 'Occupational therapist, Table 10'),
    'Melissa Toon (Emotional)': ('D1_P23', 'participant', 'Occupational therapist, Table 10'),
    'Melissa Toon (Opportunity)': ('D1_P23', 'participant', 'Occupational therapist, Table 10'),
    'Angela Lwage': ('D1_P26', 'participant', 'Social worker, Table 10'),
    'Angela Lwage (Suggestion)': ('D1_P26', 'participant', 'Social worker, Table 10'),
    'Angela Lwage (Challenge)': ('D1_P26', 'participant', 'Social worker, Table 10'),
    'Angela Lwage (Social)': ('D1_P26', 'participant', 'Social worker, Table 10'),
    'Muhammad Ali': ('D1_P22', 'participant', 'Family consultant, Table 10'),
    'Muhammad Ali (Physical)': ('D1_P22', 'participant', 'Family consultant, Table 10'),
    'Muhammad Ali (Challenge)': ('D1_P22', 'participant', 'Family consultant, Table 10'),
    'Muhammad Ali (Suggestion)': ('D1_P22', 'participant', 'Family consultant, Table 10'),
    'Dr. Khalid': ('D1_P27', 'participant', 'Child psychiatrist, Table 10'),
    'Dr. Khalid (Suggestion)': ('D1_P27', 'participant', 'Child psychiatrist, Table 10'),
    'Dr. Khalid (Social/Emotional)': ('D1_P27', 'participant', 'Child psychiatrist, Table 10'),
    'Dr. Khalid (Challenge)': ('D1_P27', 'participant', 'Child psychiatrist, Table 10'),
    'Dr. Khalid (Rebuttal)': ('D1_P27', 'participant', 'Child psychiatrist, Table 10'),
    'Nawal Yosul': ('D1_P24', 'participant', 'Resident doctor, Table 10'),
    'Nawal Yosul (Intellectual)': ('D1_P24', 'participant', 'Resident doctor, Table 10'),
    'Nawal Yosul (Challenge)': ('D1_P24', 'participant', 'Resident doctor, Table 10'),
    'Nawal Yosul (Spiritual)': ('D1_P24', 'participant', 'Resident doctor, Table 10'),
    'Nawal Yosul (Physical)': ('D1_P24', 'participant', 'Resident doctor, Table 10'),
    'Nawal Yosul (Suggestion)': ('D1_P24', 'participant', 'Resident doctor, Table 10'),
    # Generic / unclear
    '_UNLABELED_': ('D1_UNK', 'unclear', 'Unlabeled speaker'),
    # Table 4 unclear
    '\u0623\u0646\u0627  \u0645\u0631\u0651 \u0639\u0644\u064a \u0645\u0648\u0636\u0648\u0639\u064a\u0646 \u0635\u0631\u0627\u062d\u0629. \u0627\u0644\u0648\u0639\u064a': ('D1_U03', 'unclear', 'Unknown, Table 4'),
    '\u0627\u0644\u0634\u062e\u0635 \u064a\u0643\u0648\u0646 \u0639\u0627\u0631\u0641 \u0625\u0630\u0627 \u0645\u062b\u0644\u0627 \u0627\u062c\u062a\u0647\u062f \u0648 \u0642\u0627\u0645 \u0628\u0634\u064a\u0621 \u0645\u0639\u064a\u0646': ('D1_U04', 'unclear', 'Unknown, Table 4'),
}

def get_anon(label):
    if label in LABEL_MAP:
        return LABEL_MAP[label][0]
    return f'D1_UNKNOWN'

def get_role(label):
    if label in LABEL_MAP:
        return LABEL_MAP[label][2]
    return 'Unknown'

# ═══════════════════════════════════════════════════════
# LOAD DATA
# ═══════════════════════════════════════════════════════
with open(d1_out / 'CASE_D1_coded_segments.csv', 'r', encoding='utf-8') as f:
    segments = list(csv.DictReader(f))

with open(d1_out / 'CASE_D1_excerpt_bank.csv', 'r', encoding='utf-8') as f:
    excerpts = list(csv.DictReader(f))

with open(d1_out / 'CASE_D1_participant_summary.csv', 'r', encoding='utf-8') as f:
    part_summary = list(csv.DictReader(f))

theme_codes = {
    'Theme_1_Balanced_Contentment': [
        'contentment_as_core', 'balance_multidimensional', 'spiritual_moral_anchor',
        'safety_security', 'inner_peace', 'pillar_physical', 'pillar_emotional',
        'pillar_social', 'pillar_intellectual', 'pillar_spiritual'],
    'Theme_2_Care_Ecology': [
        'interdependent_care', 'school_clinical_link', 'family_support_ecology',
        'community_volunteer_bridge', 'professional_identity', 'service_recipient_voice'],
    'Theme_3_Service_Fragmentation': [
        'system_disconnection', 'training_deficit', 'resource_strain',
        'healthcare_system_critique', 'consultation_time_pressure',
        'staffing_shortage_isolation', 'medical_model_dominance',
        'missing_stepdown_services', 'bureaucratic_barrier',
        'child_vulnerability', 'child_mental_health', 'bullying_violence',
        'child_rights_inclusion'],
    'Theme_4_Culturally_Grounded_Solutions': [
        'cultural_local_adaptation', 'transcultural_workforce',
        'awareness_education', 'early_detection_prevention',
        'practical_recommendation', 'top_down_systemic_change',
        'healthcare_worker_wellbeing', 'gradual_implementation',
        'digital_platform_proposal', 'respite_inclusive_resources'],
}

# Verify mapping completeness
all_labels = set(s['speaker_label'] for s in segments)
unmapped = all_labels - set(LABEL_MAP.keys())
if unmapped:
    print(f"WARNING: {len(unmapped)} unmapped labels:")
    for u in sorted(unmapped):
        print(f"  '{u[:60]}'")
else:
    print("All speaker labels mapped successfully.")

# ═══════════════════════════════════════════════════════
# EVIDENCE TYPE CLASSIFIER
# ═══════════════════════════════════════════════════════
TP_MARKERS = ['Described ', 'Noted ', 'Called for', 'Argued ', 'Highlighted ',
              'Identified ', 'Introduced ', 'Proposed ', 'Recommended ', 'Provided ',
              'Recounted ', 'Suggested ', 'Emphasized ', 'Critiqued ', 'Shared ',
              'Speaking as']

def classify_evidence(source_file, text):
    if 'NT' in source_file and 'AR' not in source_file:
        return 'note_taker_summary'
    if 'HWCH10AR' in source_file and any(text[:50].startswith(m) for m in TP_MARKERS):
        return 'note_style_transcript_summary'
    return 'verbatim_transcript'

# ═══════════════════════════════════════════════════════
# APPLY ANONYMIZATION TO CODED SEGMENTS
# ═══════════════════════════════════════════════════════
for seg in segments:
    seg['anonymized_speaker'] = get_anon(seg['speaker_label'])

# Write back
seg_fields = list(segments[0].keys())
if 'anonymized_speaker' not in seg_fields:
    seg_fields.append('anonymized_speaker')
# Remove old 'assigned_themes' if present from prior run
seg_fields = [f for f in seg_fields if f != 'assigned_themes']

with open(d1_out / 'CASE_D1_coded_segments.csv', 'w', encoding='utf-8', newline='') as f:
    w = csv.DictWriter(f, fieldnames=seg_fields, extrasaction='ignore')
    w.writeheader()
    w.writerows(segments)

print(f"Coded segments: anonymized_speaker applied to {len(segments)} rows")

# ═══════════════════════════════════════════════════════
# APPLY ANONYMIZATION TO EXCERPT BANK + ADD evidence_type
# ═══════════════════════════════════════════════════════
# Re-read the original excerpt bank (before prior run modified it)
# The current excerpt bank may have anonymized_speaker from prior run
# We need speaker_label — check if it's still there
if 'speaker_label' in excerpts[0]:
    label_field = 'speaker_label'
elif 'anonymized_speaker' in excerpts[0]:
    # Prior run already replaced — we need the original labels
    # They should still be in the segment data
    label_field = None
else:
    label_field = None

for e in excerpts:
    if label_field:
        orig_label = e[label_field]
    else:
        # Look up from segments by segment_id
        matching = [s for s in segments if s['segment_id'] == e['segment_id']]
        orig_label = matching[0]['speaker_label'] if matching else e.get('anonymized_speaker', '')
    e['anonymized_speaker'] = get_anon(orig_label)
    e['evidence_type'] = classify_evidence(e['source_file'], e.get('excerpt_text', ''))

exc_fields = ['theme', 'segment_id', 'source_file', 'table_id', 'anonymized_speaker',
              'speaker_type', 'evidence_type', 'question_id', 'excerpt_text', 'codes', 'language']
with open(d1_out / 'CASE_D1_excerpt_bank.csv', 'w', encoding='utf-8', newline='') as f:
    w = csv.DictWriter(f, fieldnames=exc_fields, extrasaction='ignore')
    w.writeheader()
    w.writerows(excerpts)

ev_counts = Counter(e['evidence_type'] for e in excerpts)
print(f"Excerpt bank: evidence_type added — {dict(ev_counts)}")

# ═══════════════════════════════════════════════════════
# APPLY ANONYMIZATION TO PARTICIPANT SUMMARY
# ═══════════════════════════════════════════════════════
# Re-read original if needed
with open(d1_out / 'CASE_D1_participant_summary.csv', 'r', encoding='utf-8') as f:
    ps_raw = list(csv.DictReader(f))

# Check which field has the label
if 'speaker_label' in ps_raw[0]:
    for ps in ps_raw:
        ps['anonymized_code'] = get_anon(ps['speaker_label'])
    ps_fields = ['anonymized_code', 'source_file', 'table_id', 'speaker_type',
                 'segment_count', 'total_chars', 'questions_covered', 'top_codes']
elif 'anonymized_code' in ps_raw[0]:
    # Already has it from prior run — but codes may be wrong, recompute from speaker_label if available
    ps_fields = list(ps_raw[0].keys())
else:
    ps_fields = list(ps_raw[0].keys())

with open(d1_out / 'CASE_D1_participant_summary.csv', 'w', encoding='utf-8', newline='') as f:
    w = csv.DictWriter(f, fieldnames=ps_fields, extrasaction='ignore')
    w.writeheader()
    w.writerows(ps_raw)

print(f"Participant summary: anonymized ({len(ps_raw)} rows)")

# ═══════════════════════════════════════════════════════
# RECOMPUTE SOURCE CONTRIBUTION TABLE (Issue 5)
# ═══════════════════════════════════════════════════════
p_segs = [s for s in segments if s['speaker_type'] == 'participant']

source_data = defaultdict(lambda: {
    'total': 0, 'p_segs': 0, 'speakers': set(), 'questions': set(), 'themes': set()
})
for seg in segments:
    src = seg['source_file']
    source_data[src]['total'] += 1
    if seg['speaker_type'] == 'participant':
        source_data[src]['p_segs'] += 1
        source_data[src]['speakers'].add(seg['speaker_label'])
    source_data[src]['questions'].add(seg['question_id'])
    seg_codes = set(seg['codes'].split(';'))
    for theme, codes in theme_codes.items():
        if seg_codes & set(codes) and seg['speaker_type'] == 'participant':
            source_data[src]['themes'].add(theme)

src_rows = []
for src in sorted(source_data):
    d = source_data[src]
    src_rows.append({
        'source_file': src,
        'total_segments': d['total'],
        'participant_segments': d['p_segs'],
        'unique_speakers': len(d['speakers']),
        'questions_covered': ';'.join(sorted(d['questions'])),
        'themes_present': ';'.join(sorted(d['themes'])) if d['themes'] else '',
    })

with open(d1_out / 'CASE_D1_source_contribution_table.csv', 'w', encoding='utf-8', newline='') as f:
    w = csv.DictWriter(f, fieldnames=list(src_rows[0].keys()))
    w.writeheader()
    w.writerows(src_rows)

print("\nSource contribution (corrected):")
for r in src_rows:
    t = r['themes_present'] if r['themes_present'] else 'NONE (general_response only)'
    print(f"  {r['source_file']:20s} p={r['participant_segments']:4d} themes={t[:60]}")

# ═══════════════════════════════════════════════════════
# RECOMPUTE QUESTION-THEME MATRIX
# ═══════════════════════════════════════════════════════
questions = sorted(set(s['question_id'] for s in segments),
                   key=lambda x: int(x[1:]) if x[1:].isdigit() else 99)
matrix_rows = []
for qid in questions:
    row = {'question_id': qid}
    q_p_segs = [s for s in p_segs if s['question_id'] == qid]
    for theme in theme_codes:
        t_codes = set(theme_codes[theme])
        matching = [s for s in q_p_segs if set(s['codes'].split(';')) & t_codes]
        row[f'{theme}_segments'] = len(matching)
        row[f'{theme}_speakers'] = len(set(s['speaker_label'] for s in matching))
        row[f'{theme}_tables'] = len(set(s['table_id'] for s in matching))
    matrix_rows.append(row)

with open(d1_out / 'CASE_D1_question_theme_matrix.csv', 'w', encoding='utf-8', newline='') as f:
    w = csv.DictWriter(f, fieldnames=list(matrix_rows[0].keys()))
    w.writeheader()
    w.writerows(matrix_rows)

# ═══════════════════════════════════════════════════════
# RECOMPUTE PROMINENCE/SALIENCE
# ═══════════════════════════════════════════════════════
prominence_data = {}
for theme in theme_codes:
    t_codes = set(theme_codes[theme])
    theme_p_segs = [s for s in p_segs if set(s['codes'].split(';')) & t_codes]
    d = {
        'segments': len(theme_p_segs),
        'speakers': len(set(s['speaker_label'] for s in theme_p_segs)),
        'tables': len(set(s['table_id'] for s in theme_p_segs)),
        'chars': sum(len(s['segment_text']) for s in theme_p_segs),
        'questions': sorted(set(s['question_id'] for s in theme_p_segs)),
    }
    d['composite'] = d['segments'] + (d['speakers'] * 3) + (d['tables'] * 5) + (len(d['questions']) * 2)
    prominence_data[theme] = d

ranked = sorted(prominence_data.items(), key=lambda x: -x[1]['composite'])
sal_labels = ['most_prominent', 'highly_prominent', 'moderately_prominent', 'present_but_less_prominent']

prominence_rows = []
for i, (theme, d) in enumerate(ranked):
    sal = sal_labels[min(i, len(sal_labels)-1)]
    explanation = (f'{d["segments"]} participant segments from {d["speakers"]} speakers '
                   f'across {d["tables"]} tables covering {len(d["questions"])} questions '
                   f'({";".join(d["questions"])}). '
                   f'Composite score: {d["composite"]}.')
    prominence_rows.append({
        'theme': theme, 'participant_segments': d['segments'],
        'unique_speakers': d['speakers'], 'unique_tables': d['tables'],
        'total_chars': d['chars'], 'questions_present': ';'.join(d['questions']),
        'composite_score': d['composite'], 'salience': sal,
        'salience_explanation': explanation,
    })

with open(d1_out / 'CASE_D1_prominence_salience.csv', 'w', encoding='utf-8', newline='') as f:
    w = csv.DictWriter(f, fieldnames=list(prominence_rows[0].keys()))
    w.writeheader()
    w.writerows(prominence_rows)

# Theme summary table
theme_short = {
    'Theme_1_Balanced_Contentment': 'Balanced contentment, safety, and moral steadiness',
    'Theme_2_Care_Ecology': 'Interdependent professional support ecology',
    'Theme_3_Service_Fragmentation': 'Fragmented, strained childhood service conditions',
    'Theme_4_Culturally_Grounded_Solutions': 'Culturally grounded coordination and implementation',
}
sal_map = {r['theme']: r['salience'] for r in prominence_rows}
ts_rows = []
for theme in theme_codes:
    d = prominence_data[theme]
    ts_rows.append({
        'theme_number': theme.split('_')[1], 'theme_name': theme_short[theme],
        'participant_segments': d['segments'], 'unique_speakers': d['speakers'],
        'unique_tables': d['tables'], 'questions_present': ';'.join(d['questions']),
        'salience': sal_map[theme],
    })
with open(d1_out / 'CASE_D1_theme_summary_table.csv', 'w', encoding='utf-8', newline='') as f:
    w = csv.DictWriter(f, fieldnames=list(ts_rows[0].keys()))
    w.writeheader()
    w.writerows(ts_rows)

# ═══════════════════════════════════════════════════════
# QUESTION EVIDENCE TABLE (recompute)
# ═══════════════════════════════════════════════════════
qe_rows = []
for qid in questions:
    q_segs = [s for s in segments if s['question_id'] == qid]
    qe_rows.append({
        'question_id': qid,
        'participant_segments': sum(1 for s in q_segs if s['speaker_type'] == 'participant'),
        'moderator_segments': sum(1 for s in q_segs if s['speaker_type'] == 'moderator'),
        'unclear_segments': sum(1 for s in q_segs if s['speaker_type'] == 'unclear'),
        'unique_participant_speakers': len(set(s['speaker_label'] for s in q_segs if s['speaker_type'] == 'participant')),
        'source_files': len(set(s['source_file'] for s in q_segs)),
        'top_codes': ';'.join(c for c, _ in Counter(
            code for s in q_segs if s['speaker_type'] == 'participant'
            for code in s['codes'].split(';')).most_common(5)),
    })
with open(d1_out / 'CASE_D1_question_evidence_table.csv', 'w', encoding='utf-8', newline='') as f:
    w = csv.DictWriter(f, fieldnames=list(qe_rows[0].keys()))
    w.writeheader()
    w.writerows(qe_rows)

# ═══════════════════════════════════════════════════════
# CREATE IDENTITY KEY (INTERNAL ONLY)
# ═══════════════════════════════════════════════════════
# Build identity rows from actual data
seen_codes = set()
identity_rows = []
for label, (anon, stype, role) in sorted(LABEL_MAP.items(), key=lambda x: x[1][0]):
    if anon in seen_codes or anon == 'D1_UNK':
        continue
    seen_codes.add(anon)
    # Find matching segments
    matching = [s for s in segments if s['speaker_label'] == label]
    if not matching:
        continue
    tid = matching[0]['table_id']
    src = matching[0]['source_file']
    identity_rows.append({
        'real_name': label,
        'anonymized_code': anon,
        'role': role,
        'case_id': 'CASE_D1',
        'day_label': 'Day 1 — Childhood',
        'table_id': tid,
        'source_files': src,
        'notes': f'{stype}, {len(matching)} segments',
    })

wb_key = Workbook()
ws_warn = wb_key.active
ws_warn.title = 'CONFIDENTIAL'
ws_warn.cell(row=1, column=1, value='CONFIDENTIAL — INTERNAL USE ONLY').font = Font(bold=True, size=16, color='FF0000')
ws_warn.cell(row=3, column=1, value='This file maps anonymized participant codes to real identities.')
ws_warn.cell(row=4, column=1, value='Must NOT be included in any outward-facing academic submission.')

ws_key = wb_key.create_sheet('Identity_Key')
key_headers = ['real_name', 'anonymized_code', 'role', 'case_id', 'day_label', 'table_id', 'source_files', 'notes']
hfill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
hfont = Font(bold=True, size=11, color='FFFFFF')
wrap = Alignment(wrap_text=True, vertical='top')
thin = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'), bottom=Side(style='thin'))

for col, h in enumerate(key_headers, 1):
    c = ws_key.cell(row=1, column=col, value=h)
    c.font = hfont; c.fill = hfill; c.alignment = wrap; c.border = thin

for i, row in enumerate(identity_rows, 2):
    for col, h in enumerate(key_headers, 1):
        c = ws_key.cell(row=i, column=col, value=row.get(h, ''))
        c.alignment = wrap; c.border = thin

ws_key.column_dimensions['A'].width = 50
ws_key.column_dimensions['B'].width = 14
ws_key.column_dimensions['C'].width = 35
ws_key.column_dimensions['H'].width = 40

wb_key.save(d1_out / 'participant_identity_key.xlsx')
print(f"\nIdentity key: {len(identity_rows)} entries")

# ═══════════════════════════════════════════════════════
# CREATE ANONYMIZED PARTICIPANT SUMMARY (REPORTING)
# ═══════════════════════════════════════════════════════
anon_agg = defaultdict(lambda: {
    'role': '', 'table': '', 'questions': set(), 'segs': 0, 'chars': 0,
    'themes': set(), 'in_bank': False, 'lang': set()
})
for seg in p_segs:
    ac = seg['anonymized_speaker']
    anon_agg[ac]['segs'] += 1
    anon_agg[ac]['chars'] += len(seg['segment_text'])
    anon_agg[ac]['questions'].add(seg['question_id'])
    anon_agg[ac]['table'] = seg['table_id']
    anon_agg[ac]['role'] = get_role(seg['speaker_label'])
    seg_codes = set(seg['codes'].split(';'))
    for theme, codes in theme_codes.items():
        if seg_codes & set(codes):
            anon_agg[ac]['themes'].add(theme)
    if any('\u0600' <= c <= '\u06FF' for c in seg['segment_text'][:100]):
        anon_agg[ac]['lang'].add('ar')
    else:
        anon_agg[ac]['lang'].add('en')

for e in excerpts:
    ac = e.get('anonymized_speaker', '')
    if ac in anon_agg:
        anon_agg[ac]['in_bank'] = True

wb_as = Workbook()
ws_as = wb_as.active
ws_as.title = 'Participant_Summary'
as_headers = ['anonymized_code', 'role_label', 'case_id', 'day_label', 'table_id',
              'question_ids', 'segment_count', 'total_chars', 'linked_themes',
              'quotation_in_excerpt_bank', 'language_note']
bfill = PatternFill(start_color='2980B9', end_color='2980B9', fill_type='solid')
for col, h in enumerate(as_headers, 1):
    c = ws_as.cell(row=1, column=col, value=h)
    c.font = hfont; c.fill = bfill; c.alignment = wrap; c.border = thin

row_num = 2
for ac in sorted(anon_agg.keys()):
    d = anon_agg[ac]
    ws_as.cell(row=row_num, column=1, value=ac)
    ws_as.cell(row=row_num, column=2, value=d['role'])
    ws_as.cell(row=row_num, column=3, value='CASE_D1')
    ws_as.cell(row=row_num, column=4, value='Day 1 — Childhood')
    ws_as.cell(row=row_num, column=5, value=d['table'])
    ws_as.cell(row=row_num, column=6, value=';'.join(sorted(d['questions'])))
    ws_as.cell(row=row_num, column=7, value=d['segs'])
    ws_as.cell(row=row_num, column=8, value=d['chars'])
    ws_as.cell(row=row_num, column=9, value=';'.join(sorted(d['themes'])) if d['themes'] else 'general_response only')
    ws_as.cell(row=row_num, column=10, value='Yes' if d['in_bank'] else 'No')
    ws_as.cell(row=row_num, column=11, value=';'.join(sorted(d['lang'])))
    for col in range(1, len(as_headers) + 1):
        ws_as.cell(row=row_num, column=col).alignment = wrap
        ws_as.cell(row=row_num, column=col).border = thin
    row_num += 1

ws_as.column_dimensions['A'].width = 14
ws_as.column_dimensions['B'].width = 35
ws_as.column_dimensions['I'].width = 50

wb_as.save(d1_out / 'participant_summary_anonymized.xlsx')
print(f"Anonymized summary: {len(anon_agg)} participants")

# ═══════════════════════════════════════════════════════
# REBUILD EXCEL WORKBOOKS (anonymized)
# ═══════════════════════════════════════════════════════
# Participant workbook
wb_pw = Workbook()
ws_pw = wb_pw.active
ws_pw.title = 'Participant_Summary'
pw_h = ['Anonymized Code', 'Source', 'Table', 'Type', 'Segments', 'Chars', 'Questions', 'Top Codes']
for col, h in enumerate(pw_h, 1):
    c = ws_pw.cell(row=1, column=col, value=h)
    c.font = hfont; c.fill = bfill; c.alignment = wrap; c.border = thin

ps_sorted = sorted(ps_raw, key=lambda x: (
    0 if x.get('speaker_type','') == 'participant' else 1,
    -int(x.get('total_chars', 0))
))
for i, p in enumerate(ps_sorted, 2):
    ac = p.get('anonymized_code', get_anon(p.get('speaker_label', '')))
    ws_pw.cell(row=i, column=1, value=ac)
    ws_pw.cell(row=i, column=2, value=p.get('source_file', ''))
    ws_pw.cell(row=i, column=3, value=p.get('table_id', ''))
    ws_pw.cell(row=i, column=4, value=p.get('speaker_type', ''))
    ws_pw.cell(row=i, column=5, value=int(p.get('segment_count', 0)))
    ws_pw.cell(row=i, column=6, value=int(p.get('total_chars', 0)))
    ws_pw.cell(row=i, column=7, value=p.get('questions_covered', ''))
    ws_pw.cell(row=i, column=8, value=p.get('top_codes', ''))

ws_sc = wb_pw.create_sheet('Source_Contribution')
sc_h = ['Source', 'Total', 'Participant', 'Speakers', 'Questions', 'Themes']
for col, h in enumerate(sc_h, 1):
    c = ws_sc.cell(row=1, column=col, value=h)
    c.font = hfont; c.fill = bfill
for i, r in enumerate(src_rows, 2):
    ws_sc.cell(row=i, column=1, value=r['source_file'])
    ws_sc.cell(row=i, column=2, value=r['total_segments'])
    ws_sc.cell(row=i, column=3, value=r['participant_segments'])
    ws_sc.cell(row=i, column=4, value=r['unique_speakers'])
    ws_sc.cell(row=i, column=5, value=r['questions_covered'])
    ws_sc.cell(row=i, column=6, value=r['themes_present'])

wb_pw.save(d1_out / 'CASE_D1_participant_workbook.xlsx')

# Theme evidence workbook
wb_te = Workbook()
ws_ts = wb_te.active
ws_ts.title = 'Theme_Summary'
for col, h in enumerate(['Theme', 'Segments', 'Speakers', 'Tables', 'Chars', 'Questions', 'Salience', 'Explanation'], 1):
    c = ws_ts.cell(row=1, column=col, value=h)
    c.font = hfont; c.fill = bfill
for i, r in enumerate(prominence_rows, 2):
    ws_ts.cell(row=i, column=1, value=r['theme'])
    ws_ts.cell(row=i, column=2, value=r['participant_segments'])
    ws_ts.cell(row=i, column=3, value=r['unique_speakers'])
    ws_ts.cell(row=i, column=4, value=r['unique_tables'])
    ws_ts.cell(row=i, column=5, value=r['total_chars'])
    ws_ts.cell(row=i, column=6, value=r['questions_present'])
    ws_ts.cell(row=i, column=7, value=r['salience'])
    ws_ts.cell(row=i, column=8, value=r['salience_explanation'])
ws_ts.column_dimensions['A'].width = 45
ws_ts.column_dimensions['H'].width = 80

ws_mx = wb_te.create_sheet('QxTheme_Matrix')
mx_h = ['Question']
for t in theme_codes:
    sh = t.split('_', 2)[-1]
    mx_h.extend([f'{sh}_segs', f'{sh}_spk', f'{sh}_tbl'])
for col, h in enumerate(mx_h, 1):
    c = ws_mx.cell(row=1, column=col, value=h)
    c.font = hfont; c.fill = bfill
for i, m in enumerate(matrix_rows, 2):
    ws_mx.cell(row=i, column=1, value=m['question_id'])
    col = 2
    for t in theme_codes:
        ws_mx.cell(row=i, column=col, value=m[f'{t}_segments'])
        ws_mx.cell(row=i, column=col+1, value=m[f'{t}_speakers'])
        ws_mx.cell(row=i, column=col+2, value=m[f'{t}_tables'])
        col += 3

ws_eb = wb_te.create_sheet('Excerpt_Bank')
eb_h = ['Theme', 'ID', 'Source', 'Table', 'Speaker', 'Type', 'Evidence', 'Q', 'Text', 'Codes', 'Lang']
for col, h in enumerate(eb_h, 1):
    c = ws_eb.cell(row=1, column=col, value=h)
    c.font = hfont; c.fill = bfill
for i, e in enumerate(excerpts, 2):
    ws_eb.cell(row=i, column=1, value=e['theme'])
    ws_eb.cell(row=i, column=2, value=e['segment_id'])
    ws_eb.cell(row=i, column=3, value=e['source_file'])
    ws_eb.cell(row=i, column=4, value=e['table_id'])
    ws_eb.cell(row=i, column=5, value=e.get('anonymized_speaker', ''))
    ws_eb.cell(row=i, column=6, value=e['speaker_type'])
    ws_eb.cell(row=i, column=7, value=e['evidence_type'])
    ws_eb.cell(row=i, column=8, value=e['question_id'])
    ws_eb.cell(row=i, column=9, value=e.get('excerpt_text', ''))
    ws_eb.cell(row=i, column=10, value=e['codes'])
    ws_eb.cell(row=i, column=11, value=e['language'])
ws_eb.column_dimensions['I'].width = 80

ws_pr = wb_te.create_sheet('Prominence_Salience')
for col, h in enumerate(['Theme', 'Segs', 'Spk', 'Tbl', 'Chars', 'Qs', 'Score', 'Salience', 'Explanation'], 1):
    c = ws_pr.cell(row=1, column=col, value=h)
    c.font = hfont; c.fill = bfill
for i, r in enumerate(prominence_rows, 2):
    ws_pr.cell(row=i, column=1, value=r['theme'])
    ws_pr.cell(row=i, column=2, value=r['participant_segments'])
    ws_pr.cell(row=i, column=3, value=r['unique_speakers'])
    ws_pr.cell(row=i, column=4, value=r['unique_tables'])
    ws_pr.cell(row=i, column=5, value=r['total_chars'])
    ws_pr.cell(row=i, column=6, value=r['questions_present'])
    ws_pr.cell(row=i, column=7, value=r['composite_score'])
    ws_pr.cell(row=i, column=8, value=r['salience'])
    ws_pr.cell(row=i, column=9, value=r['salience_explanation'])

wb_te.save(d1_out / 'CASE_D1_theme_evidence_workbook.xlsx')
print("Excel workbooks rebuilt (anonymized)")

# ═══════════════════════════════════════════════════════
# FINAL COUNTS FOR CODEBOOK FIX
# ═══════════════════════════════════════════════════════
tc = Counter(s['speaker_type'] for s in segments)
print(f"\n=== COUNTS FOR CODEBOOK FIX ===")
print(f"participant: {tc['participant']}, moderator: {tc['moderator']}, unclear: {tc['unclear']}, total: {len(segments)}")
print(f"unique codes: {len(set(c for s in segments for c in s['codes'].split(';')))}")

print("\n=== ALL STRUCTURED DATA FILES FIXED ===")
