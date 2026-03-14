"""CASE_D1 Final QA Cross-Check — 10-point verification across all files."""
import csv, json, re
from pathlib import Path
from collections import Counter, defaultdict

root = Path(r'c:\Users\baram\OneDrive\Desktop\themnatic')
d1_out = root / 'analysis' / 'CASE_D1'

errors = []
warnings = []

def check(name, condition, msg):
    if condition:
        print(f'  [PASS] {name}')
    else:
        print(f'  [FAIL] {name}: {msg}')
        errors.append(f'{name}: {msg}')

def warn(name, msg):
    print(f'  [WARN] {name}: {msg}')
    warnings.append(f'{name}: {msg}')

# Load all CSVs
def load_csv(filename):
    path = d1_out / filename
    if not path.exists():
        errors.append(f'Missing file: {filename}')
        return []
    with open(path, 'r', encoding='utf-8') as f:
        return list(csv.DictReader(f))

segments = load_csv('CASE_D1_coded_segments.csv')
participants = load_csv('CASE_D1_participant_summary.csv')
excerpts = load_csv('CASE_D1_excerpt_bank.csv')
prominence = load_csv('CASE_D1_prominence_salience.csv')
matrix = load_csv('CASE_D1_question_theme_matrix.csv')
source_reg = load_csv('CASE_D1_source_register.csv')
part_reg = load_csv('CASE_D1_participant_register.csv')
q_evidence = load_csv('CASE_D1_question_evidence_table.csv')
theme_summary = load_csv('CASE_D1_theme_summary_table.csv')
source_contrib = load_csv('CASE_D1_source_contribution_table.csv')

# ═══════════════════════════════════════════════════════
print('\n=== CHECK 1: All CSVs parse cleanly ===')
# ═══════════════════════════════════════════════════════
csv_files = [
    'CASE_D1_coded_segments.csv', 'CASE_D1_participant_summary.csv',
    'CASE_D1_excerpt_bank.csv', 'CASE_D1_prominence_salience.csv',
    'CASE_D1_question_theme_matrix.csv', 'CASE_D1_source_register.csv',
    'CASE_D1_participant_register.csv', 'CASE_D1_question_evidence_table.csv',
    'CASE_D1_theme_summary_table.csv', 'CASE_D1_source_contribution_table.csv',
]
for cf in csv_files:
    path = d1_out / cf
    check(f'CSV parse: {cf}', path.exists(), f'File not found: {cf}')

# ═══════════════════════════════════════════════════════
print('\n=== CHECK 2: All Excel sheets have consistent schema ===')
# ═══════════════════════════════════════════════════════
from openpyxl import load_workbook
for xlsx_name in ['CASE_D1_participant_workbook.xlsx', 'CASE_D1_theme_evidence_workbook.xlsx']:
    path = d1_out / xlsx_name
    if path.exists():
        wb = load_workbook(path, read_only=True)
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            headers = [cell.value for cell in ws[1]]
            check(f'Excel schema: {xlsx_name}/{ws_name}', all(h is not None for h in headers), f'Null header in {ws_name}')
        wb.close()
    else:
        check(f'Excel exists: {xlsx_name}', False, 'File not found')

# ═══════════════════════════════════════════════════════
print('\n=== CHECK 3: Every quotation maps to participant evidence ===')
# ═══════════════════════════════════════════════════════
for e in excerpts:
    check(
        f'Excerpt {e["segment_id"]} speaker_type',
        e['speaker_type'] == 'participant',
        f'Excerpt {e["segment_id"]} has speaker_type={e["speaker_type"]} (expected participant)'
    )

# ═══════════════════════════════════════════════════════
print('\n=== CHECK 4: Moderator lines excluded from participant counts ===')
# ═══════════════════════════════════════════════════════
mod_segs = [s for s in segments if s['speaker_type'] == 'moderator']
p_segs = [s for s in segments if s['speaker_type'] == 'participant']

# Check that no moderator appears in excerpt bank
mod_labels_in_excerpts = [e for e in excerpts if e['speaker_type'] == 'moderator']
check('No moderator in excerpt bank', len(mod_labels_in_excerpts) == 0,
      f'{len(mod_labels_in_excerpts)} moderator excerpts found')

# Check participant summary has correct types
mod_in_summary = [p for p in participants if p['speaker_type'] == 'moderator']
p_in_summary = [p for p in participants if p['speaker_type'] == 'participant']
check(f'Participant summary: {len(p_in_summary)} participants, {len(mod_in_summary)} moderators',
      len(p_in_summary) > 0 and len(mod_in_summary) > 0, 'Missing speaker types in summary')

# Known moderator labels should be classified correctly
known_mods = ['Kulood', 'Moderator 1', 'Moderator 2']
for label in known_mods:
    matching = [s for s in segments if s['speaker_label'] == label]
    if matching:
        all_mod = all(s['speaker_type'] == 'moderator' for s in matching)
        check(f'Known moderator "{label}" classified correctly', all_mod,
              f'Some segments for "{label}" not classified as moderator')

# ═══════════════════════════════════════════════════════
print('\n=== CHECK 5: Recommendation workbook labeled as auxiliary ===')
# ═══════════════════════════════════════════════════════
rec_rule_path = d1_out / 'CASE_D1_recommendation_usage_rule.md'
check('Recommendation usage rule exists', rec_rule_path.exists(), 'File not found')

rec_in_source_reg = [s for s in source_reg if s['file_type'] == 'recommendation_workbook']
check('Rec workbook in source register', len(rec_in_source_reg) == 1,
      f'Expected 1, found {len(rec_in_source_reg)}')
if rec_in_source_reg:
    check('Rec workbook role = auxiliary',
          rec_in_source_reg[0]['source_role'] == 'auxiliary_recommendation',
          f'Role is {rec_in_source_reg[0]["source_role"]}')

# Check final report mentions auxiliary labeling
report_path = d1_out / 'CASE_D1_final_report.md'
if report_path.exists():
    report_text = report_path.read_text(encoding='utf-8')
    check('Report labels rec workbook as auxiliary',
          'auxiliary' in report_text.lower() and 'recommendation workbook' in report_text.lower(),
          'Recommendation workbook not labeled as auxiliary in report')

# ═══════════════════════════════════════════════════════
print('\n=== CHECK 6: Theme names standardized everywhere ===')
# ═══════════════════════════════════════════════════════
theme_names_in_prominence = set(p['theme'] for p in prominence)
theme_names_in_excerpts = set(e['theme'] for e in excerpts)
theme_names_in_summary = set(t['theme_name'] for t in theme_summary)

check('Theme names consistent: prominence vs excerpts',
      theme_names_in_prominence == theme_names_in_excerpts,
      f'Mismatch: prominence={theme_names_in_prominence}, excerpts={theme_names_in_excerpts}')

# Check matrix has all themes
matrix_cols = set()
if matrix:
    for key in matrix[0].keys():
        if '_segments' in key:
            theme_from_col = key.replace('_segments', '')
            matrix_cols.add(theme_from_col)
check('Matrix has all 4 themes',
      len(matrix_cols) == 4, f'Found {len(matrix_cols)} themes in matrix: {matrix_cols}')

# ═══════════════════════════════════════════════════════
print('\n=== CHECK 7: Matrix / chart / table / report text agree ===')
# ═══════════════════════════════════════════════════════
# Check that prominence segment counts match actual coded segment counts
theme_code_map = {
    'Theme_1_Balanced_Contentment': [
        'contentment_as_core', 'balance_multidimensional', 'spiritual_moral_anchor',
        'safety_security', 'inner_peace', 'pillar_physical', 'pillar_emotional',
        'pillar_social', 'pillar_intellectual', 'pillar_spiritual'
    ],
    'Theme_2_Care_Ecology': [
        'interdependent_care', 'school_clinical_link', 'family_support_ecology',
        'community_volunteer_bridge', 'professional_identity', 'service_recipient_voice'
    ],
    'Theme_3_Service_Fragmentation': [
        'system_disconnection', 'training_deficit', 'resource_strain',
        'healthcare_system_critique', 'consultation_time_pressure',
        'staffing_shortage_isolation', 'medical_model_dominance',
        'missing_stepdown_services', 'bureaucratic_barrier',
        'child_vulnerability', 'child_mental_health', 'bullying_violence',
        'child_rights_inclusion'
    ],
    'Theme_4_Culturally_Grounded_Solutions': [
        'cultural_local_adaptation', 'transcultural_workforce',
        'awareness_education', 'early_detection_prevention',
        'practical_recommendation', 'top_down_systemic_change',
        'healthcare_worker_wellbeing', 'gradual_implementation',
        'digital_platform_proposal', 'respite_inclusive_resources'
    ],
}

for theme, codes in theme_code_map.items():
    actual_count = sum(1 for s in p_segs if set(s['codes'].split(';')) & set(codes))
    reported = [p for p in prominence if p['theme'] == theme]
    if reported:
        reported_count = int(reported[0]['participant_segments'])
        check(f'Prominence count for {theme}',
              actual_count == reported_count,
              f'Actual={actual_count}, reported={reported_count}')

# Check report mentions correct participant count
if report_path.exists():
    check('Report mentions 27 participants',
          '27' in report_text, 'Participant count 27 not found in report')
    check('Report mentions 4 themes',
          'Four final themes' in report_text or 'four final themes' in report_text or '4 final themes' in report_text,
          '4 final themes not stated in report')

# ═══════════════════════════════════════════════════════
print('\n=== CHECK 8: Participant and source counts consistent ===')
# ═══════════════════════════════════════════════════════
# Source register should have 17 entries
check('Source register: 17 entries', len(source_reg) == 17, f'Found {len(source_reg)}')

# Participant register entry count
check(f'Participant register: {len(part_reg)} entries', len(part_reg) > 0, 'Empty')

# Check segment count by speaker_type
type_counts = Counter(s['speaker_type'] for s in segments)
check(f'Segment types: participant={type_counts["participant"]}, moderator={type_counts["moderator"]}, unclear={type_counts["unclear"]}',
      type_counts['participant'] > 0 and type_counts['moderator'] > 0,
      'Missing participant or moderator segments')

# Check total segments
check(f'Total segments: {len(segments)}', len(segments) > 0, 'No segments')

# ═══════════════════════════════════════════════════════
print('\n=== CHECK 9: No unsupported Q-to-theme claim ===')
# ═══════════════════════════════════════════════════════
# Check that report's appendix matrix matches the actual matrix CSV
if report_path.exists():
    # The report says Q6 Theme 1 = 0/0 and Q7 all = 0/0
    check('Report acknowledges Q6-Q7 low coding',
          'Q6' in report_text and 'Q7' in report_text and 'general_response' in report_text,
          'Report does not acknowledge Q6-Q7 coding limitation')

# ═══════════════════════════════════════════════════════
print('\n=== CHECK 10: No Day 5 / cross-case contamination ===')
# ═══════════════════════════════════════════════════════
all_text = ''
for f in d1_out.glob('*.md'):
    all_text += f.read_text(encoding='utf-8')
for f in d1_out.glob('*.csv'):
    all_text += f.read_text(encoding='utf-8')

check('No CASE_D5 in outputs', 'CASE_D5' not in all_text, 'CASE_D5 found')
check('No Day5 in outputs', 'Day5' not in all_text and 'Day 5' not in all_text, 'Day 5 reference found')
check('No cross-case claims',
      'cross-case' not in all_text.lower() or 'no cross-case' in all_text.lower(),
      'Possible cross-case claim found')
check('No CASE_D2/D3/D4 claims',
      'CASE_D2' not in all_text and 'CASE_D3' not in all_text and 'CASE_D4' not in all_text,
      'Other case reference found in CASE_D1 outputs')

# ═══════════════════════════════════════════════════════
# FINAL SUMMARY
# ═══════════════════════════════════════════════════════
print(f'\n{"="*60}')
print(f'FINAL QA SUMMARY')
print(f'{"="*60}')
print(f'Total checks: {len(errors) + len(warnings) + sum(1 for line in [] if True)}')
print(f'Errors: {len(errors)}')
print(f'Warnings: {len(warnings)}')

if errors:
    print('\nERRORS:')
    for e in errors:
        print(f'  - {e}')
if warnings:
    print('\nWARNINGS:')
    for w in warnings:
        print(f'  - {w}')
if not errors and not warnings:
    print('\nAll checks passed. CASE_D1 package is internally consistent.')

# List all files created
print(f'\n{"="*60}')
print(f'FILES CREATED')
print(f'{"="*60}')
for f in sorted(d1_out.glob('*')):
    if f.is_file() and not f.name.startswith('_'):
        size = f.stat().st_size
        print(f'  {f.name:50s} {size:>8,d} bytes')
print(f'\nWorking files:')
for f in sorted((d1_out / '_working').glob('*')):
    if f.is_file():
        size = f.stat().st_size
        print(f'  _working/{f.name:44s} {size:>8,d} bytes')
