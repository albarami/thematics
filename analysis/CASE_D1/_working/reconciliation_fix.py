"""CASE_D1 Reconciliation & Anonymization — comprehensive fix for Issues 1-6 + anonymization.

This script:
1. Builds the anonymization mapping from participant register + coded segments
2. Fixes source contribution table (Issue 5) — acknowledges general_response limitation
3. Fixes question-theme matrix with accurate Theme 4 Q6/Q7 = 0 (Issue 2 — no change needed in matrix)
4. Recomputes prominence/salience with corrected data
5. Adds evidence_type column to excerpt bank (Issue 1)
6. Anonymizes all outward-facing CSVs
7. Regenerates Excel workbooks (anonymized)
8. Creates participant_identity_key.xlsx and participant_summary_anonymized.xlsx
"""
import csv
import re
from pathlib import Path
from collections import Counter, defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("ERROR: openpyxl required")
    exit(1)

root = Path(r'c:\Users\baram\OneDrive\Desktop\themnatic')
d1_out = root / 'analysis' / 'CASE_D1'

# ═══════════════════════════════════════════════════════
# LOAD DATA
# ═══════════════════════════════════════════════════════
with open(d1_out / 'CASE_D1_coded_segments.csv', 'r', encoding='utf-8') as f:
    segments = list(csv.DictReader(f))

with open(d1_out / 'CASE_D1_participant_register.csv', 'r', encoding='utf-8') as f:
    register = list(csv.DictReader(f))

with open(d1_out / 'CASE_D1_excerpt_bank.csv', 'r', encoding='utf-8') as f:
    excerpts = list(csv.DictReader(f))

with open(d1_out / 'CASE_D1_participant_summary.csv', 'r', encoding='utf-8') as f:
    part_summary = list(csv.DictReader(f))

theme_codes = {
    'Theme_1_Balanced_Contentment': [
        'contentment_as_core', 'balance_multidimensional', 'spiritual_moral_anchor',
        'safety_security', 'inner_peace', 'pillar_physical', 'pillar_emotional',
        'pillar_social', 'pillar_intellectual', 'pillar_spiritual'
    ],
    'Theme_2_Care_Ecology': [
        'interdependent_care', 'school_clinical_link', 'family_support_ecology',
        'community_volunteer_bridge', 'professional_identity', 'service_recipient_voice'
    ],
    'Theme_3_Service_Fragmentation': [
        'system_disconnection', 'training_deficit', 'resource_strain',
        'healthcare_system_critique', 'consultation_time_pressure',
        'staffing_shortage_isolation', 'medical_model_dominance',
        'missing_stepdown_services', 'bureaucratic_barrier',
        'child_vulnerability', 'child_mental_health', 'bullying_violence',
        'child_rights_inclusion'
    ],
    'Theme_4_Culturally_Grounded_Solutions': [
        'cultural_local_adaptation', 'transcultural_workforce',
        'awareness_education', 'early_detection_prevention',
        'practical_recommendation', 'top_down_systemic_change',
        'healthcare_worker_wellbeing', 'gradual_implementation',
        'digital_platform_proposal', 'respite_inclusive_resources'
    ],
}

# ═══════════════════════════════════════════════════════
# BUILD ANONYMIZATION MAPPING
# ═══════════════════════════════════════════════════════
# Map participant_id from register → anonymized code
# Order: by table, then by contribution size within table

anon_map = {}  # speaker_label → anon_code
identity_key = []  # rows for the identity key Excel

# Role labels for generalized reporting
role_map = {
    'D1_T0_MOD_Kulood': ('D1_M01', 'moderator', 'Moderator, Table 0'),
    'D1_T0_P1_DrAhmed': ('D1_P01', 'participant', 'Psychiatrist, Table 0'),
    'D1_T0_P2_Sp4': ('D1_P02', 'participant', 'Healthcare professional, Table 0'),
    'D1_T0_P3_Sp5': ('D1_P03', 'participant', 'Mental health professional, Table 0'),
    'D1_T0_P4_Sp7': ('D1_P04', 'participant', 'Wellbeing researcher, Table 0'),
    'D1_T0_P5_Sp6': ('D1_P05', 'participant', 'Healthcare professional, Table 0'),
    'D1_T0_P6_Sp2': ('D1_U01', 'unclear', 'Unknown role, Table 0'),
    'D1_T2_P1_DrHessa': ('D1_P06', 'participant', 'Healthcare professional, Table 2'),
    'D1_T2_P2_Mudakhala': ('D1_P07', 'participant', 'Healthcare professional, Table 2'),
    'D1_T2_MOD_Hanadi': ('D1_M02', 'moderator', 'Moderator, Table 2'),
    'D1_T3_MOD_Hala': ('D1_M03', 'moderator', 'Moderator, Table 3'),
    'D1_T4_MOD_SessionMgr': ('D1_M04', 'moderator', 'Moderator, Table 4'),
    'D1_T4_P1_DrYaseen': ('D1_P08', 'participant', 'Healthcare professional, Table 4'),
    'D1_T4_P2_Ali': ('D1_P09', 'participant', 'Service recipient, Table 4'),
    'D1_T6_MOD_Nour': ('D1_M05', 'moderator', 'Moderator, Table 6'),
    'D1_T6_P1_Manal': ('D1_P10', 'participant', 'Psychologist, Table 6'),
    'D1_T6_P2_Hafsa': ('D1_P11', 'participant', 'Community volunteer, Table 6'),
    'D1_T6_P3_HamadPR': ('D1_P12', 'participant', 'Hospital PR manager, Table 6'),
    'D1_T6_P4_DrAmna': ('D1_P13', 'participant', 'Paediatrician, Table 6'),
    'D1_T6_P5_DrAmal': ('D1_P14', 'participant', 'Paediatrician, Table 6'),
    'D1_T6_P6_Jumana': ('D1_U02', 'unclear', 'Unknown role, Table 6'),
    'D1_T7_MOD_SessionMgr': ('D1_M06', 'moderator', 'Moderator, Table 7'),
    'D1_T7_P1_Alaa': ('D1_P15', 'participant', 'Healthcare professional, Table 7'),
    'D1_T7_P2_DrWael': ('D1_P16', 'participant', 'Healthcare professional, Table 7'),
    'D1_T7_P3_Fatma': ('D1_P17', 'participant', 'Healthcare professional, Table 7'),
    'D1_T7_P4_Shafaq': ('D1_P18', 'participant', 'Healthcare professional, Table 7'),
    'D1_T7_P5_Amani': ('D1_P19', 'participant', 'Healthcare professional, Table 7'),
    'D1_T7_P6_Noura': ('D1_P20', 'participant', 'Healthcare professional, Table 7'),
    'D1_T7_P7_Latifa': ('D1_P21', 'participant', 'Healthcare professional, Table 7'),
    'D1_T10_MOD1': ('D1_M07', 'moderator', 'Moderator, Table 10'),
    'D1_T10_MOD2': ('D1_M08', 'moderator', 'Moderator, Table 10'),
    'D1_T10_P1_Muhammad': ('D1_P22', 'participant', 'Family consultant, Table 10'),
    'D1_T10_P2_Melissa': ('D1_P23', 'participant', 'Occupational therapist, Table 10'),
    'D1_T10_P3_Nawal': ('D1_P24', 'participant', 'Resident doctor, Table 10'),
    'D1_T10_P4_DrAfaf': ('D1_P25', 'participant', 'Psychologist, Table 10'),
    'D1_T10_P5_Angela': ('D1_P26', 'participant', 'Social worker, Table 10'),
    'D1_T10_P6_DrKhalid': ('D1_P27', 'participant', 'Child psychiatrist, Table 10'),
}

# Build speaker_label → anon_code map from register
for reg in register:
    pid = reg['participant_id']
    if pid in role_map:
        anon_code, stype, role_label = role_map[pid]
        label = reg['speaker_label']
        anon_map[label] = anon_code
        identity_key.append({
            'real_name': reg['name_normalized'],
            'anonymized_code': anon_code,
            'role': role_label,
            'case_id': 'CASE_D1',
            'day_label': 'Day 1 — Childhood',
            'table_id': reg['table_id'],
            'source_files': reg['source_file'],
            'speaker_label_in_csv': label,
            'notes': reg.get('classification_basis', ''),
        })

# Also map variant labels that appear in coded segments but not register
# Collect all unique speaker labels from segments
seg_labels = set(s['speaker_label'] for s in segments)
reg_labels = set(anon_map.keys())
unmapped = seg_labels - reg_labels

print(f"Mapped labels: {len(reg_labels)}")
print(f"Unmapped segment labels: {len(unmapped)}")
for ul in sorted(unmapped):
    # Try fuzzy matching — find closest register entry by table_id
    matching_segs = [s for s in segments if s['speaker_label'] == ul]
    if matching_segs:
        tid = matching_segs[0]['table_id']
        stype = matching_segs[0]['speaker_type']
        # Find register entries for same table
        table_regs = [r for r in register if r['table_id'] == tid]
        matched = False
        for tr in table_regs:
            if tr['speaker_label'] in anon_map:
                # Check if label is a substring match
                if ul[:10] in tr['speaker_label'] or tr['speaker_label'][:10] in ul:
                    anon_map[ul] = anon_map[tr['speaker_label']]
                    matched = True
                    break
        if not matched:
            # Assign based on speaker_type
            if stype == 'moderator':
                # Find existing moderator anon for this table
                for tr in table_regs:
                    if tr['speaker_type'] == 'moderator' and tr['speaker_label'] in anon_map:
                        anon_map[ul] = anon_map[tr['speaker_label']]
                        matched = True
                        break
            if not matched:
                print(f"  UNMAPPED: '{ul[:50]}' T{tid} ({stype}) — assigning generic")
                # Generate a fallback code
                if stype == 'participant':
                    fallback = f"D1_P_T{tid}"
                elif stype == 'moderator':
                    fallback = f"D1_M_T{tid}"
                else:
                    fallback = f"D1_U_T{tid}"
                anon_map[ul] = fallback

def anonymize_label(label):
    """Return anonymized code for a speaker label."""
    if label in anon_map:
        return anon_map[label]
    # Try partial match
    for key in anon_map:
        if key in label or label in key:
            return anon_map[key]
    return label  # fallback — keep as-is if no match

# ═══════════════════════════════════════════════════════
# DETERMINE EVIDENCE TYPE FOR EXCERPTS (Issue 1)
# ═══════════════════════════════════════════════════════
third_person_markers = ['Described ', 'Noted ', 'Called for', 'Argued ', 'Highlighted ',
                        'Identified ', 'Introduced ', 'Proposed ', 'Recommended ', 'Provided ',
                        'Recounted ', 'Suggested ', 'Emphasized ', 'Critiqued ', 'Shared ',
                        'Speaking as']
note_sources = {'HWCH0NT1', 'HWCH2NT1', 'HWCH3NT2', 'HWCH4NT1', 'HWCH5NT1',
                'HWCH6NT3', 'HWCH7NT1', 'HWCH10NT1'}

def classify_evidence_type(source_file, text):
    """Classify a segment's evidence type."""
    base = source_file.replace('.docx', '')
    if base in note_sources:
        return 'note_taker_summary'
    if 'HWCH10AR' in source_file:
        if any(text[:50].startswith(m) for m in third_person_markers):
            return 'note_style_transcript_summary'
        return 'verbatim_transcript'
    return 'verbatim_transcript'

# ═══════════════════════════════════════════════════════
# ASSIGN THEMES TO SEGMENTS (for recomputation)
# ═══════════════════════════════════════════════════════
for seg in segments:
    seg_codes = set(seg['codes'].split(';'))
    themes = []
    for theme, codes in theme_codes.items():
        if seg_codes & set(codes):
            themes.append(theme)
    seg['assigned_themes'] = ';'.join(themes) if themes else ''

p_segs = [s for s in segments if s['speaker_type'] == 'participant']

# ═══════════════════════════════════════════════════════
# FIX: SOURCE CONTRIBUTION TABLE (Issue 5)
# ═══════════════════════════════════════════════════════
source_data = defaultdict(lambda: {
    'total': 0, 'p_segs': 0, 'speakers': set(),
    'questions': set(), 'themes': set()
})
for seg in segments:
    src = seg['source_file']
    source_data[src]['total'] += 1
    if seg['speaker_type'] == 'participant':
        source_data[src]['p_segs'] += 1
        source_data[src]['speakers'].add(seg['speaker_label'])
    source_data[src]['questions'].add(seg['question_id'])
    for t in seg.get('assigned_themes', '').split(';'):
        if t:
            source_data[src]['themes'].add(t)

src_rows = []
for src in sorted(source_data):
    d = source_data[src]
    src_rows.append({
        'source_file': src,
        'total_segments': d['total'],
        'participant_segments': d['p_segs'],
        'unique_speakers': len(d['speakers']),
        'questions_covered': ';'.join(sorted(d['questions'])),
        'themes_present': ';'.join(sorted(d['themes'])),
    })

with open(d1_out / 'CASE_D1_source_contribution_table.csv', 'w', encoding='utf-8', newline='') as f:
    w = csv.DictWriter(f, fieldnames=list(src_rows[0].keys()))
    w.writeheader()
    w.writerows(src_rows)

print("\nFixed source contribution table:")
for r in src_rows:
    themes = r['themes_present'] if r['themes_present'] else 'NONE (all general_response)'
    print(f"  {r['source_file']:20s} p={r['participant_segments']:4d} themes={themes[:60]}")

# ═══════════════════════════════════════════════════════
# FIX: QUESTION-THEME MATRIX (Issue 2 — already correct, just confirm)
# ═══════════════════════════════════════════════════════
matrix_rows = []
questions = sorted(set(s['question_id'] for s in segments), key=lambda x: int(x[1:]) if x[1:].isdigit() else 99)
for qid in questions:
    row = {'question_id': qid}
    q_p_segs = [s for s in p_segs if s['question_id'] == qid]
    for theme in theme_codes:
        t_codes = set(theme_codes[theme])
        matching = [s for s in q_p_segs if set(s['codes'].split(';')) & t_codes]
        speakers = set(s['speaker_label'] for s in matching)
        tables = set(s['table_id'] for s in matching)
        row[f'{theme}_segments'] = len(matching)
        row[f'{theme}_speakers'] = len(speakers)
        row[f'{theme}_tables'] = len(tables)
    matrix_rows.append(row)

with open(d1_out / 'CASE_D1_question_theme_matrix.csv', 'w', encoding='utf-8', newline='') as f:
    w = csv.DictWriter(f, fieldnames=list(matrix_rows[0].keys()))
    w.writeheader()
    w.writerows(matrix_rows)

# ═══════════════════════════════════════════════════════
# FIX: PROMINENCE/SALIENCE (recompute)
# ═══════════════════════════════════════════════════════
prominence_data = {}
for theme in theme_codes:
    t_codes = set(theme_codes[theme])
    theme_p_segs = [s for s in p_segs if set(s['codes'].split(';')) & t_codes]
    total_segs = len(theme_p_segs)
    unique_speakers = len(set(s['speaker_label'] for s in theme_p_segs))
    unique_tables = len(set(s['table_id'] for s in theme_p_segs))
    total_chars = sum(len(s['segment_text']) for s in theme_p_segs)
    questions_present = sorted(set(s['question_id'] for s in theme_p_segs))
    composite = total_segs + (unique_speakers * 3) + (unique_tables * 5) + (len(questions_present) * 2)
    prominence_data[theme] = {
        'segments': total_segs, 'speakers': unique_speakers,
        'tables': unique_tables, 'chars': total_chars,
        'questions': questions_present, 'composite': composite
    }

ranked = sorted(prominence_data.items(), key=lambda x: -x[1]['composite'])
salience_labels = ['most_prominent', 'highly_prominent', 'moderately_prominent', 'present_but_less_prominent']

prominence_rows = []
for i, (theme, d) in enumerate(ranked):
    sal = salience_labels[min(i, len(salience_labels)-1)]
    if sal == 'most_prominent':
        explanation = (f'Highest composite evidence: {d["segments"]} segments from {d["speakers"]} '
                      f'speakers across {d["tables"]} tables covering {len(d["questions"])} questions.')
    elif sal == 'highly_prominent':
        explanation = (f'Strong evidence: {d["segments"]} segments from {d["speakers"]} speakers '
                      f'across {d["tables"]} tables. Broad scope but slightly narrower than most prominent.')
    elif sal == 'moderately_prominent':
        explanation = (f'Solid evidence: {d["segments"]} segments from {d["speakers"]} speakers '
                      f'across {d["tables"]} tables. Well-supported with broad speaker base.')
    else:
        explanation = (f'Present with meaningful evidence: {d["segments"]} segments from {d["speakers"]} '
                      f'speakers across {d["tables"]} tables. Concentrated, analytically deep content.')

    prominence_rows.append({
        'theme': theme, 'participant_segments': d['segments'],
        'unique_speakers': d['speakers'], 'unique_tables': d['tables'],
        'total_chars': d['chars'], 'questions_present': ';'.join(d['questions']),
        'composite_score': d['composite'], 'salience': sal,
        'salience_explanation': explanation,
    })

with open(d1_out / 'CASE_D1_prominence_salience.csv', 'w', encoding='utf-8', newline='') as f:
    w = csv.DictWriter(f, fieldnames=list(prominence_rows[0].keys()))
    w.writeheader()
    w.writerows(prominence_rows)

print("\nProminence/salience (recomputed):")
for r in prominence_rows:
    print(f"  {r['theme']:45s} segs={r['participant_segments']:4d} → {r['salience']}")

# ═══════════════════════════════════════════════════════
# FIX: THEME SUMMARY TABLE (recompute)
# ═══════════════════════════════════════════════════════
theme_short_names = {
    'Theme_1_Balanced_Contentment': 'Balanced contentment, safety, and moral steadiness',
    'Theme_2_Care_Ecology': 'Interdependent professional support ecology',
    'Theme_3_Service_Fragmentation': 'Fragmented, strained childhood service conditions',
    'Theme_4_Culturally_Grounded_Solutions': 'Culturally grounded coordination and implementation',
}
salience_map = {r['theme']: r['salience'] for r in prominence_rows}

theme_summary_rows = []
for theme in theme_codes:
    d = prominence_data[theme]
    theme_summary_rows.append({
        'theme_number': theme.split('_')[1],
        'theme_name': theme_short_names[theme],
        'participant_segments': d['segments'],
        'unique_speakers': d['speakers'],
        'unique_tables': d['tables'],
        'questions_present': ';'.join(d['questions']),
        'salience': salience_map[theme],
    })

with open(d1_out / 'CASE_D1_theme_summary_table.csv', 'w', encoding='utf-8', newline='') as f:
    w = csv.DictWriter(f, fieldnames=list(theme_summary_rows[0].keys()))
    w.writeheader()
    w.writerows(theme_summary_rows)

# ═══════════════════════════════════════════════════════
# FIX: EXCERPT BANK — add evidence_type + anonymize (Issue 1)
# ═══════════════════════════════════════════════════════
for e in excerpts:
    e['evidence_type'] = classify_evidence_type(e['source_file'], e['excerpt_text'])
    e['anonymized_speaker'] = anonymize_label(e['speaker_label'])

exc_fields = ['theme', 'segment_id', 'source_file', 'table_id', 'anonymized_speaker',
              'speaker_type', 'evidence_type', 'question_id', 'excerpt_text', 'codes', 'language']
with open(d1_out / 'CASE_D1_excerpt_bank.csv', 'w', encoding='utf-8', newline='') as f:
    w = csv.DictWriter(f, fieldnames=exc_fields)
    w.writeheader()
    for e in excerpts:
        row = {k: e.get(k, '') for k in exc_fields}
        w.writerow(row)

print(f"\nExcerpt bank: added evidence_type + anonymized_speaker")
ev_types = Counter(e['evidence_type'] for e in excerpts)
for t, c in ev_types.most_common():
    print(f"  {t}: {c}")

# ═══════════════════════════════════════════════════════
# FIX: PARTICIPANT SUMMARY — anonymize
# ═══════════════════════════════════════════════════════
for ps in part_summary:
    ps['anonymized_code'] = anonymize_label(ps['speaker_label'])

ps_fields = ['anonymized_code', 'source_file', 'table_id', 'speaker_type',
             'segment_count', 'total_chars', 'questions_covered', 'top_codes']
with open(d1_out / 'CASE_D1_participant_summary.csv', 'w', encoding='utf-8', newline='') as f:
    w = csv.DictWriter(f, fieldnames=ps_fields)
    w.writeheader()
    for ps in part_summary:
        row = {k: ps.get(k, '') for k in ps_fields}
        w.writerow(row)

print(f"\nParticipant summary: anonymized ({len(part_summary)} rows)")

# ═══════════════════════════════════════════════════════
# FIX: CODED SEGMENTS — anonymize speaker_label
# ═══════════════════════════════════════════════════════
seg_fields = list(segments[0].keys())
# Add anonymized_speaker column, keep original speaker_label for internal traceability
if 'anonymized_speaker' not in seg_fields:
    seg_fields.append('anonymized_speaker')

for seg in segments:
    seg['anonymized_speaker'] = anonymize_label(seg['speaker_label'])

with open(d1_out / 'CASE_D1_coded_segments.csv', 'w', encoding='utf-8', newline='') as f:
    w = csv.DictWriter(f, fieldnames=seg_fields)
    w.writeheader()
    w.writerows(segments)

print(f"\nCoded segments: added anonymized_speaker column")

# ═══════════════════════════════════════════════════════
# FIX: QUESTION EVIDENCE TABLE — add note about note_taker_summary (Issue 3)
# ═══════════════════════════════════════════════════════
with open(d1_out / 'CASE_D1_question_evidence_table.csv', 'r', encoding='utf-8') as f:
    q_evidence = list(csv.DictReader(f))

# Recompute with explicit note about speaker_type coverage
qe_rows = []
for qid in questions:
    q_segs = [s for s in segments if s['question_id'] == qid]
    row = {
        'question_id': qid,
        'participant_segments': sum(1 for s in q_segs if s['speaker_type'] == 'participant'),
        'moderator_segments': sum(1 for s in q_segs if s['speaker_type'] == 'moderator'),
        'unclear_segments': sum(1 for s in q_segs if s['speaker_type'] == 'unclear'),
        'unique_participant_speakers': len(set(s['speaker_label'] for s in q_segs if s['speaker_type'] == 'participant')),
        'source_files': len(set(s['source_file'] for s in q_segs)),
        'top_codes': ';'.join(c for c, _ in Counter(
            code for s in q_segs if s['speaker_type'] == 'participant'
            for code in s['codes'].split(';')
        ).most_common(5)),
    }
    qe_rows.append(row)

with open(d1_out / 'CASE_D1_question_evidence_table.csv', 'w', encoding='utf-8', newline='') as f:
    w = csv.DictWriter(f, fieldnames=list(qe_rows[0].keys()))
    w.writeheader()
    w.writerows(qe_rows)

# ═══════════════════════════════════════════════════════
# CREATE: participant_identity_key.xlsx (INTERNAL ONLY)
# ═══════════════════════════════════════════════════════
wb_key = Workbook()
ws_key = wb_key.active
ws_key.title = 'Identity_Key'
key_headers = ['real_name', 'anonymized_code', 'role', 'case_id', 'day_label',
               'table_id', 'source_files', 'speaker_label_in_csv', 'notes']
header_fill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
header_font = Font(bold=True, size=11, color='FFFFFF')
wrap = Alignment(wrap_text=True, vertical='top')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for col, h in enumerate(key_headers, 1):
    cell = ws_key.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap
    cell.border = thin_border

for i, row in enumerate(identity_key, 2):
    for col, h in enumerate(key_headers, 1):
        cell = ws_key.cell(row=i, column=col, value=row.get(h, ''))
        cell.alignment = wrap
        cell.border = thin_border

ws_key.column_dimensions['A'].width = 35
ws_key.column_dimensions['B'].width = 14
ws_key.column_dimensions['C'].width = 30
ws_key.column_dimensions['D'].width = 12
ws_key.column_dimensions['E'].width = 22
ws_key.column_dimensions['F'].width = 10
ws_key.column_dimensions['G'].width = 18
ws_key.column_dimensions['H'].width = 40
ws_key.column_dimensions['I'].width = 60

# Add confidentiality warning
ws_warn = wb_key.create_sheet('CONFIDENTIAL', 0)
ws_warn.cell(row=1, column=1, value='CONFIDENTIAL — INTERNAL USE ONLY').font = Font(bold=True, size=16, color='FF0000')
ws_warn.cell(row=3, column=1, value='This file contains the mapping between anonymized participant codes and real names.')
ws_warn.cell(row=4, column=1, value='It must NOT be included in any outward-facing academic submission package.')
ws_warn.cell(row=5, column=1, value='Store securely and share only with authorized members of the research team.')

wb_key.save(d1_out / 'participant_identity_key.xlsx')
print(f"\nCreated participant_identity_key.xlsx ({len(identity_key)} entries)")

# ═══════════════════════════════════════════════════════
# CREATE: participant_summary_anonymized.xlsx (REPORTING)
# ═══════════════════════════════════════════════════════
wb_anon = Workbook()
ws_anon = wb_anon.active
ws_anon.title = 'Participant_Summary'
anon_headers = ['anonymized_code', 'role_label', 'case_id', 'day_label', 'table_id',
                'question_ids', 'segment_count', 'total_chars', 'linked_themes',
                'quotation_in_excerpt_bank', 'language_note']

blue_fill = PatternFill(start_color='2980B9', end_color='2980B9', fill_type='solid')
for col, h in enumerate(anon_headers, 1):
    cell = ws_anon.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, size=11, color='FFFFFF')
    cell.fill = blue_fill
    cell.alignment = wrap
    cell.border = thin_border

# Build per-anonymized-code summary
anon_summary = defaultdict(lambda: {
    'role': '', 'table': '', 'questions': set(), 'segs': 0, 'chars': 0,
    'themes': set(), 'in_excerpt_bank': False, 'lang': set()
})

for seg in segments:
    if seg['speaker_type'] != 'participant':
        continue
    acode = seg['anonymized_speaker']
    anon_summary[acode]['segs'] += 1
    anon_summary[acode]['chars'] += len(seg['segment_text'])
    anon_summary[acode]['questions'].add(seg['question_id'])
    anon_summary[acode]['table'] = seg['table_id']
    for t in seg.get('assigned_themes', '').split(';'):
        if t:
            anon_summary[acode]['themes'].add(t)
    # Detect language
    if any('\u0600' <= c <= '\u06FF' for c in seg['segment_text'][:100]):
        anon_summary[acode]['lang'].add('ar')
    else:
        anon_summary[acode]['lang'].add('en')

# Mark excerpt bank presence
for e in excerpts:
    acode = e['anonymized_speaker']
    if acode in anon_summary:
        anon_summary[acode]['in_excerpt_bank'] = True

# Get role labels
for ik in identity_key:
    acode = ik['anonymized_code']
    if acode in anon_summary:
        anon_summary[acode]['role'] = ik['role']

row_num = 2
for acode in sorted(anon_summary.keys()):
    d = anon_summary[acode]
    ws_anon.cell(row=row_num, column=1, value=acode)
    ws_anon.cell(row=row_num, column=2, value=d['role'])
    ws_anon.cell(row=row_num, column=3, value='CASE_D1')
    ws_anon.cell(row=row_num, column=4, value='Day 1 — Childhood')
    ws_anon.cell(row=row_num, column=5, value=d['table'])
    ws_anon.cell(row=row_num, column=6, value=';'.join(sorted(d['questions'])))
    ws_anon.cell(row=row_num, column=7, value=d['segs'])
    ws_anon.cell(row=row_num, column=8, value=d['chars'])
    ws_anon.cell(row=row_num, column=9, value=';'.join(sorted(d['themes'])) if d['themes'] else 'general_response only')
    ws_anon.cell(row=row_num, column=10, value='Yes' if d['in_excerpt_bank'] else 'No')
    ws_anon.cell(row=row_num, column=11, value=';'.join(sorted(d['lang'])))
    for col in range(1, len(anon_headers) + 1):
        ws_anon.cell(row=row_num, column=col).alignment = wrap
        ws_anon.cell(row=row_num, column=col).border = thin_border
    row_num += 1

ws_anon.column_dimensions['A'].width = 14
ws_anon.column_dimensions['B'].width = 35
ws_anon.column_dimensions['C'].width = 12
ws_anon.column_dimensions['D'].width = 22
ws_anon.column_dimensions['E'].width = 10
ws_anon.column_dimensions['F'].width = 25
ws_anon.column_dimensions['G'].width = 14
ws_anon.column_dimensions['H'].width = 12
ws_anon.column_dimensions['I'].width = 50
ws_anon.column_dimensions['J'].width = 20
ws_anon.column_dimensions['K'].width = 12

wb_anon.save(d1_out / 'participant_summary_anonymized.xlsx')
print(f"Created participant_summary_anonymized.xlsx ({len(anon_summary)} participants)")

# ═══════════════════════════════════════════════════════
# REBUILD: participant_workbook.xlsx (anonymized)
# ═══════════════════════════════════════════════════════
wb_pw = Workbook()
ws_pw = wb_pw.active
ws_pw.title = 'Participant_Summary'
pw_headers = ['Anonymized Code', 'Source File', 'Table', 'Speaker Type',
              'Segments', 'Total Chars', 'Questions Covered', 'Top Codes']
for col, h in enumerate(pw_headers, 1):
    cell = ws_pw.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, size=11, color='FFFFFF')
    cell.fill = blue_fill
    cell.alignment = wrap
    cell.border = thin_border

ps_sorted = sorted(part_summary, key=lambda x: (
    0 if x['speaker_type'] == 'participant' else 1 if x['speaker_type'] == 'unclear' else 2,
    -int(x['total_chars'])
))
for i, p in enumerate(ps_sorted, 2):
    ws_pw.cell(row=i, column=1, value=p['anonymized_code'])
    ws_pw.cell(row=i, column=2, value=p['source_file'])
    ws_pw.cell(row=i, column=3, value=p['table_id'])
    ws_pw.cell(row=i, column=4, value=p['speaker_type'])
    ws_pw.cell(row=i, column=5, value=int(p['segment_count']))
    ws_pw.cell(row=i, column=6, value=int(p['total_chars']))
    ws_pw.cell(row=i, column=7, value=p['questions_covered'])
    ws_pw.cell(row=i, column=8, value=p['top_codes'])
    for col in range(1, len(pw_headers) + 1):
        ws_pw.cell(row=i, column=col).alignment = wrap
        ws_pw.cell(row=i, column=col).border = thin_border

ws_pw.column_dimensions['A'].width = 16
ws_pw.column_dimensions['B'].width = 18
ws_pw.column_dimensions['H'].width = 45

# Source contribution sheet
ws_sc = wb_pw.create_sheet('Source_Contribution')
sc_headers = ['Source File', 'Total Segments', 'Participant Segments',
              'Unique Speakers', 'Questions', 'Themes Present']
for col, h in enumerate(sc_headers, 1):
    cell = ws_sc.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, size=11, color='FFFFFF')
    cell.fill = blue_fill
    cell.alignment = wrap
    cell.border = thin_border
for i, r in enumerate(src_rows, 2):
    ws_sc.cell(row=i, column=1, value=r['source_file'])
    ws_sc.cell(row=i, column=2, value=r['total_segments'])
    ws_sc.cell(row=i, column=3, value=r['participant_segments'])
    ws_sc.cell(row=i, column=4, value=r['unique_speakers'])
    ws_sc.cell(row=i, column=5, value=r['questions_covered'])
    ws_sc.cell(row=i, column=6, value=r['themes_present'])

wb_pw.save(d1_out / 'CASE_D1_participant_workbook.xlsx')
print("Rebuilt CASE_D1_participant_workbook.xlsx (anonymized)")

# ═══════════════════════════════════════════════════════
# REBUILD: theme_evidence_workbook.xlsx (anonymized)
# ═══════════════════════════════════════════════════════
wb_te = Workbook()

# Sheet 1: Theme Summary
ws_ts = wb_te.active
ws_ts.title = 'Theme_Summary'
ts_headers = ['Theme', 'Segments', 'Speakers', 'Tables', 'Total Chars',
              'Questions', 'Salience', 'Explanation']
for col, h in enumerate(ts_headers, 1):
    cell = ws_ts.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, size=11, color='FFFFFF')
    cell.fill = blue_fill
    cell.alignment = wrap
    cell.border = thin_border
for i, r in enumerate(prominence_rows, 2):
    ws_ts.cell(row=i, column=1, value=r['theme'])
    ws_ts.cell(row=i, column=2, value=r['participant_segments'])
    ws_ts.cell(row=i, column=3, value=r['unique_speakers'])
    ws_ts.cell(row=i, column=4, value=r['unique_tables'])
    ws_ts.cell(row=i, column=5, value=r['total_chars'])
    ws_ts.cell(row=i, column=6, value=r['questions_present'])
    ws_ts.cell(row=i, column=7, value=r['salience'])
    ws_ts.cell(row=i, column=8, value=r['salience_explanation'])

ws_ts.column_dimensions['A'].width = 45
ws_ts.column_dimensions['H'].width = 80

# Sheet 2: QxTheme Matrix
ws_mx = wb_te.create_sheet('QxTheme_Matrix')
theme_names_list = list(theme_codes.keys())
mx_headers = ['Question']
for t in theme_names_list:
    short = t.split('_', 2)[-1]
    mx_headers.extend([f'{short}_segs', f'{short}_spk', f'{short}_tbl'])
for col, h in enumerate(mx_headers, 1):
    cell = ws_mx.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, size=11, color='FFFFFF')
    cell.fill = blue_fill
for i, m in enumerate(matrix_rows, 2):
    ws_mx.cell(row=i, column=1, value=m['question_id'])
    col = 2
    for t in theme_names_list:
        ws_mx.cell(row=i, column=col, value=m[f'{t}_segments'])
        ws_mx.cell(row=i, column=col+1, value=m[f'{t}_speakers'])
        ws_mx.cell(row=i, column=col+2, value=m[f'{t}_tables'])
        col += 3

# Sheet 3: Excerpt Bank (anonymized)
ws_eb = wb_te.create_sheet('Excerpt_Bank')
eb_headers = ['Theme', 'Segment ID', 'Source', 'Table', 'Anonymized Speaker',
              'Speaker Type', 'Evidence Type', 'Question', 'Excerpt Text', 'Codes', 'Language']
for col, h in enumerate(eb_headers, 1):
    cell = ws_eb.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, size=11, color='FFFFFF')
    cell.fill = blue_fill
for i, e in enumerate(excerpts, 2):
    ws_eb.cell(row=i, column=1, value=e['theme'])
    ws_eb.cell(row=i, column=2, value=e['segment_id'])
    ws_eb.cell(row=i, column=3, value=e['source_file'])
    ws_eb.cell(row=i, column=4, value=e['table_id'])
    ws_eb.cell(row=i, column=5, value=e['anonymized_speaker'])
    ws_eb.cell(row=i, column=6, value=e['speaker_type'])
    ws_eb.cell(row=i, column=7, value=e['evidence_type'])
    ws_eb.cell(row=i, column=8, value=e['question_id'])
    ws_eb.cell(row=i, column=9, value=e['excerpt_text'])
    ws_eb.cell(row=i, column=10, value=e['codes'])
    ws_eb.cell(row=i, column=11, value=e['language'])

ws_eb.column_dimensions['E'].width = 16
ws_eb.column_dimensions['G'].width = 28
ws_eb.column_dimensions['I'].width = 80

# Sheet 4: Prominence
ws_pr = wb_te.create_sheet('Prominence_Salience')
pr_headers = ['Theme', 'Segments', 'Speakers', 'Tables', 'Chars',
              'Questions', 'Composite', 'Salience', 'Explanation']
for col, h in enumerate(pr_headers, 1):
    cell = ws_pr.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, size=11, color='FFFFFF')
    cell.fill = blue_fill
for i, r in enumerate(prominence_rows, 2):
    ws_pr.cell(row=i, column=1, value=r['theme'])
    ws_pr.cell(row=i, column=2, value=r['participant_segments'])
    ws_pr.cell(row=i, column=3, value=r['unique_speakers'])
    ws_pr.cell(row=i, column=4, value=r['unique_tables'])
    ws_pr.cell(row=i, column=5, value=r['total_chars'])
    ws_pr.cell(row=i, column=6, value=r['questions_present'])
    ws_pr.cell(row=i, column=7, value=r['composite_score'])
    ws_pr.cell(row=i, column=8, value=r['salience'])
    ws_pr.cell(row=i, column=9, value=r['salience_explanation'])

wb_te.save(d1_out / 'CASE_D1_theme_evidence_workbook.xlsx')
print("Rebuilt CASE_D1_theme_evidence_workbook.xlsx (anonymized)")

# ═══════════════════════════════════════════════════════
# OUTPUT: Anonymization mapping for report edits
# ═══════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("ANONYMIZATION MAPPING (for report edits)")
print("=" * 60)
for ik in sorted(identity_key, key=lambda x: x['anonymized_code']):
    print(f"  {ik['anonymized_code']:8s} = {ik['real_name']:40s} [{ik['role']}]")

# Output actual counts for codebook fix
print("\n" + "=" * 60)
print("ACTUAL SEGMENT COUNTS (for codebook fix)")
print("=" * 60)
type_counts = Counter(s['speaker_type'] for s in segments)
total = len(segments)
print(f"  participant: {type_counts['participant']}")
print(f"  moderator: {type_counts['moderator']}")
print(f"  unclear: {type_counts['unclear']}")
print(f"  total: {total}")

# Unique codes count
all_codes = set()
for s in segments:
    for c in s['codes'].split(';'):
        all_codes.add(c)
print(f"  unique codes: {len(all_codes)}")

print("\n=== RECONCILIATION FIX COMPLETE ===")
print("Remaining manual edits needed:")
print("  1. CASE_D1_final_report.md — anonymize names, fix quotation labels, fix Theme 4 Q6/Q7 claims")
print("  2. CASE_D1_final_themes.md — anonymize names, fix Theme 4 key questions")
print("  3. CASE_D1_working_codebook.md — fix segment counts (line 80)")
print("  4. CASE_D1_crosscheck_report.md — fix file count, fix evidence rule, update all checks")
print("  5. CASE_D1_candidate_themes.md — anonymize if names present")
