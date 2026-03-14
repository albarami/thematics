"""CASE_D4 Gate 3 builder: segment extraction, coding, summary tables, workbook.

Reads d4_transcripts_extracted.json, extracts speaker-attributed question-bounded
segments, applies codebook-based codes, and writes:
  - CASE_D4_coded_segments.csv
  - CASE_D4_participant_summary.csv
  - CASE_D4_question_evidence_table.csv
  - CASE_D4_participant_workbook.xlsx
  - _working/CASE_D4_segment_notes.md
"""
from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Final

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
ROOT: Final = Path(r"c:\Users\baram\OneDrive\Desktop\themnatic")
CASE_DIR: Final = ROOT / "analysis" / "CASE_D4"
WORKING_DIR: Final = CASE_DIR / "_working"
TRANSCRIPTS_JSON: Final = WORKING_DIR / "d4_transcripts_extracted.json"

CODED_SEGMENTS_CSV: Final = CASE_DIR / "CASE_D4_coded_segments.csv"
PARTICIPANT_SUMMARY_CSV: Final = CASE_DIR / "CASE_D4_participant_summary.csv"
QUESTION_EVIDENCE_CSV: Final = CASE_DIR / "CASE_D4_question_evidence_table.csv"
PARTICIPANT_WORKBOOK: Final = CASE_DIR / "CASE_D4_participant_workbook.xlsx"
SEGMENT_NOTES_MD: Final = WORKING_DIR / "CASE_D4_segment_notes.md"

QUESTION_ORDER: Final = ["Q1", "Q2", "Q3", "Q4", "Q5", "Q6", "Q7"]
SEGMENT_CSV_FIELDS: Final = [
    "segment_id", "source_file", "table_id", "speaker_code", "speaker_type",
    "role_label", "attribution_status", "question_id", "segment_text",
    "codes", "language",
]
TOP_CODE_LIMIT: Final = 5

# ---------------------------------------------------------------------------
# Speaker roster — reuses the participant-register aliases
# ---------------------------------------------------------------------------
SPEAKERS: Final = [
    # (code, source, table, type, role_label, aliases)
    ("D4_M01", "HWEL1AR.docx", "1", "moderator", "Moderator, Table 1",
     ["Diana (moderator)"]),
    ("D4_M02", "HWEL3AR.docx", "3", "moderator", "Moderator, Table 3",
     ["أ.هالة", "أستاذة هالة", "ا.هالة"]),
    ("D4_M05", "HWEL7AR.docx", "7", "moderator", "Moderator, Table 7",
     ["مدير الجلسة", "مدير الجلسه"]),
    ("D4_M06", "HWEL9AR.docx", "9", "moderator", "Moderator, Table 9",
     ["Dr Abdellatif Moderator"]),
    ("D4_M07", "HWEL10AR.docx", "10", "moderator", "Moderator, Table 10",
     ["Moderator"]),
    ("D4_P01", "HWEL1AR.docx", "1", "participant",
     "Participant, role not stated, Table 1",
     ["د. درويش العمادي"]),
    ("D4_P02", "HWEL1AR.docx", "1", "participant",
     "Director of Pharmacy, Table 1",
     ["أنا د. أمينة اليزيدي ديريكتور الفارمسي(Director of Pharmacy)",
      "د. أمينة اليزيدي ديريكتور الفارمسي(Director of Pharmacy)"]),
    ("D4_P03", "HWEL1AR.docx", "1", "participant",
     "Home care participant, Hamad Medical Corporation, Table 1",
     ["مريم التميمي من الرعاية المنزلية مؤسسة حمد الطبية",
      "مريم التميمي من الرعاية المنزلية"]),
    ("D4_P04", "HWEL1AR.docx", "1", "participant",
     "Service recipient, Table 1",
     ["هند الجابر  :(patient)", "هند الجابر"]),
    ("D4_P05", "HWEL1AR.docx", "1", "participant",
     "Participant, Hamad Medical Corporation, Table 1",
     ["بشاير الراشد من مؤسسة حمد الطبية"]),
    ("D4_P06", "HWEL1AR.docx", "1", "participant",
     "Participant, role not stated, Table 1",
     ["الدكتور كاكل رسول", "د. كاكل رسول"]),
    ("D4_P07", "HWEL1AR.docx", "1", "participant",
     "Participant, role not stated, Table 1",
     ["د. احمد هاني"]),
    ("D4_P08", "HWEL3AR.docx", "3", "participant",
     "Participant, Hamad Medical Corporation, Table 3",
     ["د.سوسو (Hamad Medical Corporate)", "د.سوسو"]),
    ("D4_P09", "HWEL3AR.docx", "3", "participant",
     "Psychologist, Ihsan Center, Table 3",
     ["أ.مريم", "أستاذة مريم"]),
    ("D4_P10", "HWEL3AR.docx", "3", "participant",
     "Social-welfare / social-work participant, Table 3",
     ["دكتورة فيروز ( clinical social worker - HMC-Mental Health)",
      "دكتورة فيروز (Head of Social Welfare Section - Ihsan Center)",
      "د.فيروز"]),
    ("D4_P11", "HWEL3AR.docx", "3", "participant",
     "Participant, role not stated, Table 3",
     ["أ.ريما خليفة", "أ.ريم"]),
    ("D4_P12", "HWEL3AR.docx", "3", "participant",
     "Occupational therapy, elderly care, Hamad Medical Corporation, Table 3",
     ["أستاذ محمود (OCCUPATIONAL THERAPY-ELDERLY - Hamad Medical Corporate)",
      "أستاذ محمود", "أ.محمود", "أ,محمود"]),
    ("D4_P13", "HWEL7AR.docx", "7", "participant",
     "Rehabilitation-hospital participant, Table 7",
     ["ضحى محمود"]),
    ("D4_P14", "HWEL7AR.docx", "7", "participant",
     "Home care physician, Table 7",
     ["د. حنان اليافعى", "د.حنان اليافعي", "د. حنان اليافعي",
      "د.حنان اليافعى"]),
    ("D4_P15", "HWEL7AR.docx", "7", "participant",
     "Participant, role not stated, Table 7",
     ["د.علاء الدين الكيلاني"]),
    ("D4_P16", "HWEL7AR.docx", "7", "participant",
     "Participant, role not stated, Table 7",
     ["د. شاكيناز"]),
    ("D4_P17", "HWEL7AR.docx", "7", "participant",
     "Elderly-center participant, Table 7",
     ["عادل السليطي"]),
    ("D4_P18", "HWEL7AR.docx", "7", "participant",
     "Service recipient, Table 7",
     ["د. مصطفى"]),
    ("D4_P19", "HWEL7AR.docx", "7", "participant",
     "Participant, role not stated, Table 7",
     ["امل العبيدلى"]),
    ("D4_P20", "HWEL7AR.docx", "7", "participant",
     "Senior public-sector leader, Table 7",
     ["الوزير"]),
    ("D4_P21", "HWEL7AR.docx", "7", "participant",
     "Participant, role not stated, Table 7",
     ["د.محمد"]),
    ("D4_P22", "HWEL9AR.docx", "9", "participant",
     "Social worker, Ihsan Center, Table 9",
     ["Noor Ali Albadr Social Worker - ishan",
      "Noor Ali Albadr Social Worker \u2013 ishan"]),
    ("D4_P23", "HWEL9AR.docx", "9", "participant",
     "Acute geriatric and elderly care, Hamad Medical Corporation, Table 9",
     ["Esmat Swallmeh Acute Geriatric and Elderly care - Hamad Medical Corporate"]),
    ("D4_P24", "HWEL9AR.docx", "9", "participant",
     "Senior pharmacist, Naufar, Table 9",
     ["Randa Al Okka Sr. Pharmacist - Naufar"]),
    ("D4_P25", "HWEL9AR.docx", "9", "participant",
     "Service recipient / patient voice, Table 9",
     ["patient د. محمد عبد العليم"]),
    ("D4_P26", "HWEL10AR.docx", "10", "participant",
     "Consultant in family medicine, PHCC, Table 10",
     ["Dr. Amit"]),
    ("D4_P27", "HWEL10AR.docx", "10", "participant",
     "Student, HBKU, Table 10",
     ["Manal Sherif"]),
    ("D4_P28", "HWEL10AR.docx", "10", "participant",
     "Rumailah Hospital participant, Table 10",
     ["Al anood"]),
    ("D4_P29", "HWEL10AR.docx", "10", "participant",
     "Qatar University participant, Table 10",
     ["Dr. Khalood"]),
]

SOURCE_LANGUAGE: Final[dict[str, str]] = {
    "HWEL10AR.docx": "en",
}

# ---------------------------------------------------------------------------
# Question detection — content markers (forward-only like D3)
# ---------------------------------------------------------------------------
QUESTION_MARKERS: Final = [
    ("Q1", ["ما معنى الحياة الطيبة", "ما هي الحياة الطيبة", "what is well-being for you",
            "what is wellbeing for you", "defining well-being"]),
    ("Q2", ["كيف تعرف العافية", "كيف تعرّف العافية", "كيف نعرف العافية",
            "العافية من منظورك", "هذا الشخص بعافية",
            "describing someone as", "when we describe someone as well",
            "ماذا نقصد من كلمة عافية"]),
    ("Q3", ["كيف يمكن لهذه الركائز أن تسهم", "كيف يمكن لهذه الركائز أن تحسن",
            "the five pillars of well-being", "how can these pillars",
            "بناء على منظور الحياة الطيبة"]),
    ("Q4", ["هل تنعكس ركائز الحياة الطيبة", "هل هي ركائز الحياه الطيبه موجوده",
            "pillars in the current healthcare", "من واقع تجربتك",
            "based on your experience"]),
    ("Q5", ["ما التحديات الرئيسية", "التحديات", "ما هي الصعوبات", "التحدي",
            "the biggest challenges", "challenge:", "challenges:"]),
    ("Q6", ["ما الفرص المتاحة", "الفرص", "فرص متاحة", "ما هي الفرص",
            "opportunity:", "opportunities:", "available opportunities"]),
    ("Q7", ["اقتراحاتكم العملية", "الحلول", "المقترحات",
            "practical suggestions", "ننتقل للمقترحات",
            "(suggestion):", "suggestion:", "suggestions:"]),
]

# Source-specific paragraph breakpoints (1-indexed paragraph numbers)
SOURCE_BREAKPOINTS: Final[dict[str, list[tuple[int, str]]]] = {
    "HWEL10AR.docx": [
        (2, "Q1"),   # Part 1 moderator question / first substantive content
        (13, "Q2"),  # Part 2 first substantive response after header at paragraph 12
        (18, "Q3"),  # Part 3 first substantive response after header at paragraph 17
        (23, "Q4"),  # Part 4 first substantive response after header at paragraph 22
        (29, "Q5"),  # Part 5 first substantive response after header at paragraph 28
    ],
    "HWEL1AR.docx": [
        (2, "Q1"),   # Moderator opens with Q1
        (11, "Q2"),  # Diana moves to السؤال الثاني
        (29, "Q3"),  # Diana moves to السؤال الثالث
        (50, "Q4"),  # Participant shifts to Q4 (pillars in healthcare)
        (58, "Q5"),  # Diana "هذه من التحديات"
    ],
    "HWEL3AR.docx": [
        (2, "Q1"),   # د.سوسو opens Q1
        (8, "Q2"),   # Discussion shifts to العافية
        (26, "Q3"),  # فيروز starts Q3 about ركائز
        (38, "Q4"),  # هالة asks Q4 about pillars in healthcare
        (42, "Q5"),  # هالة says "تكلم في التحديات"
        (58, "Q6"),  # Discussion shifts to فرص/technology
        (71, "Q7"),  # Practical suggestions
    ],
    "HWEL7AR.docx": [
        (2, "Q1"),   # Moderator Q1
        (14, "Q2"),  # Moderator "ننتقل للسؤال الثاني"
        (30, "Q3"),  # Moderator Q3
        (42, "Q4"),  # Moderator "السؤال رقم اربعة"
        (88, "Q5"),  # Moderator "التحديات"
        (100, "Q6"), # Moderator "الفرص المتاحة"
        (110, "Q7"), # Practical suggestions near end
    ],
    "HWEL9AR.docx": [
        (2, "Q1"),    # Moderator opens Q1
        (65, "Q2"),   # Moderator shifts to العافية (~line 401)
        (82, "Q3"),   # Moderator asks about ركائز الخمسة (~line 418)
        (224, "Q4"),  # Patient asks about pillars reflected in services (~line 560)
        (347, "Q5"),  # Moderator says "نتكلم على التحديات" (~line 683)
        (466, "Q6"),  # Moderator asks "ما هي الفرص المتاحة" (~line 802)
        (556, "Q7"),  # Moderator says "اقتراحات عملية" (~line 892)
    ],
}

# ---------------------------------------------------------------------------
# Skip / header detection
# ---------------------------------------------------------------------------
SKIP_PREFIXES: Final = ("HWEL", "Part ", "Top of Form", "Bottom of Form")
HEADER_LINE_RE: Final = re.compile(
    r"^(Date:|Table #|Part \d|Transcript|التاريخ|طاولة رقم)", re.IGNORECASE
)
SEPARATOR_LINE: Final = "-" * 40


def _normalize(text: str) -> str:
    """Collapse whitespace and normalize dashes."""
    return re.sub(r"\s+", " ", text.replace("\u2013", "-").replace("\u2014", "-")).strip()


def _is_skip(text: str) -> bool:
    """Return True for headers, separators, and empty lines."""
    if not text:
        return True
    if text.startswith("-" * 20):
        return True
    if any(text.startswith(p) for p in SKIP_PREFIXES):
        return True
    if HEADER_LINE_RE.match(text):
        return True
    return False


def _is_substantive(text: str) -> bool:
    """Return True if text is long enough to be a meaningful segment."""
    if not text or len(text) < 12:
        return False
    if _is_skip(text):
        return False
    if text.endswith(":") and len(text.split()) <= 5:
        return False
    return True


# ---------------------------------------------------------------------------
# Alias lookup and label matching
# ---------------------------------------------------------------------------
def _build_alias_lookup() -> list[tuple[str, str, str, str, str]]:
    """Return (normalized_alias, code, table_id, speaker_type, role_label)
    sorted longest-first for greedy matching."""
    pairs: list[tuple[str, str, str, str, str]] = []
    for code, _src, tid, stype, rlabel, aliases in SPEAKERS:
        for alias in aliases:
            pairs.append((_normalize(alias), code, tid, stype, rlabel))
    return sorted(pairs, key=lambda p: len(p[0]), reverse=True)


def _match_label(
    text: str,
    alias_lookup: list[tuple[str, str, str, str, str]],
) -> tuple[dict[str, str] | None, str]:
    """Try to match a speaker label at the start of text.

    Returns (speaker_info_dict, remaining_content) or (None, '').
    """
    for alias, code, tid, stype, rlabel in alias_lookup:
        if text == alias or text == f"{alias}:" or text == f"{alias} :":
            return {"code": code, "table_id": tid, "speaker_type": stype,
                    "role_label": rlabel}, ""
        for sep in (":", " :", ": "):
            prefix = f"{alias}{sep}"
            if text.startswith(prefix):
                remainder = text[len(prefix):].strip()
                return {"code": code, "table_id": tid, "speaker_type": stype,
                        "role_label": rlabel}, remainder
    return None, ""


# ---------------------------------------------------------------------------
# Question detection
# ---------------------------------------------------------------------------
def _detect_question(text: str, current: str | None) -> str | None:
    """Forward-only question detection from content markers."""
    lowered = text.lower()
    current_idx = QUESTION_ORDER.index(current) if current in QUESTION_ORDER else -1
    for qid, markers in QUESTION_MARKERS:
        q_idx = QUESTION_ORDER.index(qid)
        if q_idx < current_idx:
            continue
        for marker in markers:
            if marker.lower() in lowered:
                return qid
    return current


def _apply_breakpoints(
    source: str, para_num: int, current: str | None,
) -> str | None:
    """Apply source-specific paragraph breakpoints."""
    breakpoints = SOURCE_BREAKPOINTS.get(source, [])
    result = current
    for start_para, qid in breakpoints:
        if para_num >= start_para:
            result = qid
        else:
            break
    return result


# ---------------------------------------------------------------------------
# Code assignment (retrieval-layer, not theme engine)
# ---------------------------------------------------------------------------
TEXT_CODE_PATTERNS: Final[dict[str, list[str]]] = {
    "contentment_acceptance": ["رضا", "قناعة", "acceptance", "contentment",
        "satisfied", "راحة البال", "peace of mind", "happiness is a decision"],
    "balance_multidimensional": ["توازن", "balance", "holistic", "integrated",
        "wholeness", "ركائز", "pillars", "all aspects", "شمولي"],
    "spiritual_moral_anchor": ["الله", "faith", "prayer", "الإيمان", "الروحية",
        "عبادة", "moral", "spiritual", "القرآن", "الدين", "بوصلة"],
    "dignity_autonomy": ["كرامة", "dignity", "استقلالية", "independence",
        "autonomy", "عالة", "burden", "حق القرار", "كبار القدر",
        "self-reliance", "معتمد على نفسه"],
    "inner_peace_stability": ["طمأنينة", "راحة نفسية", "استقرار",
        "peace of mind", "calm", "stability", "settled"],
    "safety_security": ["أمن", "أمان", "حماية", "safety", "secure", "آمن"],
    "person_centered_care": ["متمحورة حول الفرد", "person-centred",
        "person-centered", "PCCC", "holistic approach", "patient-centered"],
    "home_care_services": ["رعاية منزلية", "home care", "home visit",
        "زيارات منزلية", "هوم كير"],
    "rehabilitation_services": ["تأهيل", "rehabilitation", "physiotherapy",
        "مستشفى التأهيل", "rehab"],
    "multidisciplinary_team": ["فريق متعدد", "MDT", "interdisciplinary",
        "فريق طبي كامل", "team", "تعاون بين التخصصات", "collaboration"],
    "mental_health_elderly": ["صحة نفسية", "depression", "isolation",
        "cognitive", "اكتئاب", "عزلة", "mental health", "نفسي"],
    "community_support_programs": ["إحسان", "نادي كبار", "community centre",
        "برامج مجتمعية", "مجلس", "نادي", "مراكز"],
    "family_caregiver_ecology": ["أسرة", "أهل", "family support", "caregiver",
        "أبناء", "أولاد", "بنت", "زوج", "أم", "أب", "والد"],
    "containment_emotional_holding": ["احتواء", "containment", "holding",
        "تحتويه", "احتويت"],
    "service_adequacy_reflection": ["هل تنعكس", "are pillars reflected",
        "services currently provided", "موجودة", "متوفرة"],
    "institutional_example": ["I-COPE", "إحسان", "Rumailah", "رميلة",
        "PHCC", "حمد", "مستشفى التأهيل", "كتارا", "Qatar Biobank"],
    "care_fragmentation": ["فجوة", "fragmented", "silo", "coordination gap",
        "اندماج الخدمات", "disconnected", "تكاملية"],
    "workforce_constraint": ["نقص", "staffing", "shortage",
        "not enough time", "10 minutes", "15 minutes", "وقت محدود"],
    "training_deficit": ["تدريب", "training", "not trained", "skills",
        "تأهيل مهني", "مؤهل"],
    "cultural_barrier": ["ثقافة", "cultural", "language barrier",
        "اختلافات ثقافية", "لغوية", "جنسية"],
    "dependence_resistance": ["عالة", "burden", "resist help", "stubborn",
        "لا يتقبل", "عنيد", "ضغط عليه"],
    "post_retirement_gap": ["تقاعد", "retirement", "سن الستين",
        "في عز خبرته", "retired"],
    "intergenerational_shift": ["تنمر", "ageism", "تآكل",
        "عجوز", "عجوزة", "كم عمرك"],
    "awareness_deficit": ["ما بيعرفوا", "not aware", "لا يعرف الخدمات",
        "نقص الوعي", "what services exist"],
    "institutional_culture_gap": ["ليست مؤسساتية", "not institutional",
        "depends on the individual", "ثقافة المؤسسة", "غير مطلوب"],
    "awareness_education": ["توعية", "awareness", "education", "campaign",
        "تثقيف صحي", "workshops", "حملات"],
    "technology_integration": ["تكنولوجيا", "technology", "digital",
        "artificial intelligence", "apps", "chatbot", "ذكاء اصطناعي",
        "تطبيقات", "wearable", "sensor", "platform"],
    "training_development": ["برامج تأهيل", "training program",
        "specialisation", "تخصص أكاديمي"],
    "policy_recommendation": ["سياسة", "policy", "ministry", "institutional",
        "mandatory", "مؤسساتية", "وزارة"],
    "practical_suggestion": ["اقتراح", "suggest", "recommend",
        "لازم", "يجب", "ينبغي", "مقترح"],
    "interdisciplinary_model": ["فريق طبي كامل", "collaboration",
        "interdisciplinary", "بين التخصصات"],
    "epigenetics_precision": ["epigenetics", "precision medicine",
        "biomarkers", "Qatar Biobank", "precision health"],
    "pillar_spiritual": ["ركيزة روحية", "روحية", "spiritual pillar", "الروحية"],
    "pillar_emotional": ["ركيزة عاطفية", "عاطفية", "emotional pillar", "العاطفية"],
    "pillar_intellectual": ["ركيزة فكرية", "فكرية", "intellectual pillar", "الفكرية"],
    "pillar_physical": ["ركيزة بدنية", "بدنية", "physical pillar", "البدنية", "جسدية"],
    "pillar_social": ["ركيزة اجتماعية", "اجتماعية", "social pillar", "الاجتماعية"],
    "age_dignity_language": ["كبار القدر", "people of standing", "كبار السن",
        "ختيار", "شياب", "Gadal"],
}

ROLE_CODE_PATTERNS: Final[dict[str, list[str]]] = {
    "professional_identity": [
        "consultant", "physician", "pharmacist", "therapist", "social work",
        "occupational", "nurse", "psychologist", "geriatric", "rehabilitation",
        "director", "manager", "طبيب", "دكتور", "أخصائي", "صيدلي",
    ],
    "service_recipient_voice": [
        "service recipient", "patient", "service user", "مريض", "متلقي",
        "كمستخدم", "كمتلقي",
    ],
}

# Codes that count as semantic (not structural/residual)
SEMANTIC_CODES: Final = {
    code for code in TEXT_CODE_PATTERNS
    if code not in {
        "pillar_spiritual", "pillar_emotional", "pillar_intellectual",
        "pillar_physical", "pillar_social",
    }
}

TOP_CODE_EXCLUSIONS: Final = {
    "general_response", "moderator_context", "professional_identity",
    "service_recipient_voice", "pillar_spiritual", "pillar_emotional",
    "pillar_intellectual", "pillar_physical", "pillar_social",
}

PROPOSAL_MARKERS: Final = [
    "suggest", "recommend", "proposal", "should", "need to",
    "لازم", "يجب", "ينبغي", "مقترح", "اقتراح",
]


def _contains_any(text: str, markers: list[str]) -> bool:
    return any(m in text for m in markers)


def _derive_codes(row: dict[str, str]) -> set[str]:
    """Derive codes from text content and role label."""
    lowered = row["segment_text"].lower()
    role_low = row["role_label"].lower()
    codes: set[str] = set()

    for code, markers in TEXT_CODE_PATTERNS.items():
        if _contains_any(lowered, [m.lower() for m in markers]):
            codes.add(code)

    for code, markers in ROLE_CODE_PATTERNS.items():
        if _contains_any(role_low, markers):
            codes.add(code)

    if _contains_any(lowered, ["as a patient", "كمتلقي", "كمستخدم",
                                "my father", "my mother", "والدي", "أبوي"]):
        codes.add("service_recipient_voice")

    if row["question_id"] == "Q7" and _contains_any(lowered,
            [m.lower() for m in PROPOSAL_MARKERS]):
        codes.add("practical_suggestion")

    if row["question_id"] in {"Q6", "Q7"} and _contains_any(lowered,
            ["technology", "digital", "artificial intelligence",
             "apps", "chatbot", "تكنولوجيا", "ذكاء اصطناعي",
             "تطبيقات", "wearable", "sensor", "platform"]):
        codes.add("technology_integration")
        codes.add("practical_suggestion")

    if row["question_id"] in {"Q6", "Q7"} and _contains_any(lowered,
            ["policy", "ministry", "mandatory", "مؤسساتية", "وزارة"]):
        codes.add("policy_recommendation")
        codes.add("practical_suggestion")

    if row["question_id"] in {"Q5", "Q6", "Q7"} and _contains_any(lowered,
            ["team", "collaboration", "interdisciplinary", "فريق", "تعاون"]):
        codes.add("interdisciplinary_model")

    if not codes.intersection(SEMANTIC_CODES):
        codes.add("general_response")

    return codes


CODE_ORDER: Final = [
    "contentment_acceptance", "balance_multidimensional",
    "spiritual_moral_anchor", "dignity_autonomy",
    "inner_peace_stability", "safety_security",
    "person_centered_care", "home_care_services",
    "rehabilitation_services", "multidisciplinary_team",
    "mental_health_elderly", "community_support_programs",
    "family_caregiver_ecology", "containment_emotional_holding",
    "service_adequacy_reflection", "institutional_example",
    "care_fragmentation", "workforce_constraint", "training_deficit",
    "cultural_barrier", "dependence_resistance", "post_retirement_gap",
    "intergenerational_shift", "awareness_deficit",
    "institutional_culture_gap",
    "awareness_education", "technology_integration",
    "training_development", "policy_recommendation",
    "practical_suggestion", "interdisciplinary_model",
    "epigenetics_precision",
    "pillar_spiritual", "pillar_emotional", "pillar_intellectual",
    "pillar_physical", "pillar_social",
    "age_dignity_language", "professional_identity",
    "service_recipient_voice",
    "moderator_context", "general_response",
]


def _ordered_codes(codes: set[str]) -> list[str]:
    order_map = {c: i for i, c in enumerate(CODE_ORDER)}
    return sorted(codes, key=lambda c: (order_map.get(c, len(order_map)), c))


def _flush_active_segment(
    rows: list[dict[str, str]],
    seg_counter: int,
    source: str,
    table_id: str,
    speaker: dict[str, str] | None,
    question_id: str | None,
    active_parts: list[str],
    language: str,
) -> tuple[int, int]:
    if speaker is None or not active_parts or not question_id:
        return seg_counter, 0
    content = " ".join(active_parts)
    if not _is_substantive(content):
        return seg_counter, 0
    rows.append(
        _build_segment_row(
            seg_counter,
            source,
            table_id,
            speaker,
            question_id,
            content,
            language,
        )
    )
    return seg_counter + 1, 1


# ---------------------------------------------------------------------------
# Segment extraction
# ---------------------------------------------------------------------------
def _extract_segments() -> tuple[list[dict[str, str]], list[str]]:
    """Extract speaker-attributed, question-bounded segments from transcripts."""
    with TRANSCRIPTS_JSON.open(encoding="utf-8") as fh:
        records = json.load(fh)

    alias_lookup = _build_alias_lookup()
    rows: list[dict[str, str]] = []
    notes: list[str] = ["# CASE_D4 Segment Extraction Notes", ""]
    seg_counter = 1

    for record in records:
        source = record["filename"]
        paragraphs = record["paragraphs"]
        table_match = re.search(r"HWEL(\d+)", source)
        table_id = table_match.group(1) if table_match else ""
        lang = SOURCE_LANGUAGE.get(source, "ar")
        current_q: str | None = None
        active_speaker: dict[str, str] | None = None
        active_parts: list[str] = []
        extracted = 0
        unmatched_labels: set[str] = set()

        for idx, raw_para in enumerate(paragraphs):
            text = _normalize(raw_para)
            if _is_skip(text):
                continue

            para_num = idx + 1
            previous_q = current_q
            next_q = _apply_breakpoints(source, para_num, current_q)
            next_q = _detect_question(text, next_q)
            if previous_q and next_q != previous_q and active_speaker and active_parts:
                seg_counter, added = _flush_active_segment(
                    rows,
                    seg_counter,
                    source,
                    table_id,
                    active_speaker,
                    previous_q,
                    active_parts,
                    lang,
                )
                extracted += added
                active_parts = []
            current_q = next_q

            matched, remainder = _match_label(text, alias_lookup)

            if matched is not None:
                seg_counter, added = _flush_active_segment(
                    rows,
                    seg_counter,
                    source,
                    table_id,
                    active_speaker,
                    current_q,
                    active_parts,
                    lang,
                )
                extracted += added
                active_speaker = matched
                active_parts = [remainder] if remainder else []
                continue

            if active_speaker is not None:
                active_parts.append(text)
            elif current_q and _is_substantive(text):
                unmatched_labels.add(text[:60])

        seg_counter, added = _flush_active_segment(
            rows,
            seg_counter,
            source,
            table_id,
            active_speaker,
            current_q,
            active_parts,
            lang,
        )
        extracted += added

        notes.append(f"## {source}")
        notes.append(f"- Paragraphs: {len(paragraphs)}")
        notes.append(f"- Extracted segments: {extracted}")
        notes.append(f"- Unmatched text lines: {len(unmatched_labels)}")
        if unmatched_labels:
            for item in sorted(unmatched_labels):
                notes.append(f"  - {item}")
        notes.append("")

    return rows, notes


def _build_segment_row(
    seg_num: int,
    source: str,
    table_id: str,
    speaker: dict[str, str],
    question_id: str,
    content: str,
    language: str,
) -> dict[str, str]:
    stype = speaker["speaker_type"]
    attr = "identified" if stype != "unclear" else "indeterminate"
    return {
        "segment_id": f"D4_S{seg_num:04d}",
        "source_file": source,
        "table_id": table_id,
        "speaker_code": speaker["code"],
        "speaker_type": stype,
        "role_label": speaker["role_label"],
        "attribution_status": attr,
        "question_id": question_id,
        "segment_text": content.strip(),
        "codes": "moderator_context" if stype == "moderator" else "general_response",
        "language": language,
    }


# ---------------------------------------------------------------------------
# Recode segments with codebook
# ---------------------------------------------------------------------------
def _recode(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    recoded: list[dict[str, str]] = []
    for row in rows:
        updated = dict(row)
        if row["speaker_type"] == "moderator":
            updated["codes"] = "moderator_context"
            recoded.append(updated)
            continue
        codes = _derive_codes(row)
        updated["codes"] = ";".join(_ordered_codes(codes))
        recoded.append(updated)
    return recoded


# ---------------------------------------------------------------------------
# Summary tables
# ---------------------------------------------------------------------------
def _question_sort(qid: str) -> tuple[int, str]:
    try:
        return (QUESTION_ORDER.index(qid), qid)
    except ValueError:
        return (len(QUESTION_ORDER), qid)


def _normalise_codes(raw: str) -> list[str]:
    return [c.strip() for c in raw.split(";") if c.strip()]


def _format_top_codes(counter: Counter[str]) -> str:
    filtered = Counter({c: n for c, n in counter.items()
                        if c not in TOP_CODE_EXCLUSIONS})
    ranked_src = filtered if filtered else counter
    ranked = sorted(ranked_src.items(), key=lambda x: (-x[1], x[0]))
    return ";".join(c for c, _ in ranked[:TOP_CODE_LIMIT])


def _build_participant_summary(
    rows: list[dict[str, str]],
) -> list[dict[str, str]]:
    grouped: dict[tuple[str, str, str, str], list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        if row["speaker_type"] != "participant":
            continue
        key = (row["speaker_code"], row["source_file"],
               row["table_id"], row["speaker_type"])
        grouped[key].append(row)

    all_keys = {
        (s[0], s[1], s[2], s[3])
        for s in SPEAKERS if s[3] == "participant"
    }
    summary: list[dict[str, str]] = []
    for key in sorted(all_keys | set(grouped)):
        code, src, tid, stype = key
        group = grouped.get(key, [])
        if not group:
            summary.append({
                "anonymized_code": code, "source_file": src,
                "table_id": tid, "speaker_type": stype,
                "segment_count": "0", "total_chars": "0",
                "questions_covered": "none", "top_codes": "none",
            })
            continue
        questions = sorted(
            {r["question_id"] for r in group if r["question_id"]},
            key=_question_sort,
        )
        code_ctr: Counter[str] = Counter()
        for r in group:
            code_ctr.update(_normalise_codes(r["codes"]))
        total_chars = sum(len(r["segment_text"]) for r in group)
        summary.append({
            "anonymized_code": code, "source_file": src,
            "table_id": tid, "speaker_type": stype,
            "segment_count": str(len(group)),
            "total_chars": str(total_chars),
            "questions_covered": ";".join(questions),
            "top_codes": _format_top_codes(code_ctr),
        })
    return summary


def _build_question_evidence(
    rows: list[dict[str, str]],
) -> list[dict[str, str]]:
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        grouped[row["question_id"]].append(row)

    evidence: list[dict[str, str]] = []
    for qid in sorted(grouped, key=_question_sort):
        group = grouped[qid]
        code_ctr: Counter[str] = Counter()
        participants: set[str] = set()
        sources = sorted({r["source_file"] for r in group})
        p_seg = m_seg = u_seg = 0
        for r in group:
            code_ctr.update(_normalise_codes(r["codes"]))
            st = r["speaker_type"]
            if st == "participant":
                p_seg += 1
                participants.add(r["speaker_code"])
            elif st == "moderator":
                m_seg += 1
            else:
                u_seg += 1
        evidence.append({
            "question_id": qid,
            "participant_segments": str(p_seg),
            "moderator_segments": str(m_seg),
            "unclear_segments": str(u_seg),
            "unique_participant_speakers": str(len(participants)),
            "source_files": ";".join(sources),
            "top_codes": _format_top_codes(code_ctr),
        })
    return evidence


# ---------------------------------------------------------------------------
# Excel workbook
# ---------------------------------------------------------------------------
def _build_workbook(summary: list[dict[str, str]], target: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Participant_Summary"
    hfont = Font(bold=True, size=11, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    fill = PatternFill(start_color="2980B9", end_color="2980B9",
                       fill_type="solid")
    headers = ["Anonymized Code", "Source", "Table", "Speaker Type",
               "Segments", "Chars", "Questions", "Top Codes"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = fill
        cell.alignment = wrap
        cell.border = border
    for ri, row in enumerate(summary, 2):
        vals = [
            row["anonymized_code"], row["source_file"], row["table_id"],
            row["speaker_type"], int(row["segment_count"]),
            int(row["total_chars"]), row["questions_covered"],
            row["top_codes"],
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=col, value=v)
            cell.alignment = wrap
            cell.border = border
    wb.save(target)


# ---------------------------------------------------------------------------
# CSV writer
# ---------------------------------------------------------------------------
def _write_csv(
    path: Path, fields: list[str], rows: list[dict[str, str]],
) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    """Build the full CASE_D4 Gate 3 coding layer."""
    raw_rows, notes = _extract_segments()
    coded_rows = _recode(raw_rows)

    _write_csv(CODED_SEGMENTS_CSV, SEGMENT_CSV_FIELDS, coded_rows)

    summary = _build_participant_summary(coded_rows)
    _write_csv(PARTICIPANT_SUMMARY_CSV, [
        "anonymized_code", "source_file", "table_id", "speaker_type",
        "segment_count", "total_chars", "questions_covered", "top_codes",
    ], summary)

    evidence = _build_question_evidence(coded_rows)
    _write_csv(QUESTION_EVIDENCE_CSV, [
        "question_id", "participant_segments", "moderator_segments",
        "unclear_segments", "unique_participant_speakers",
        "source_files", "top_codes",
    ], evidence)

    _build_workbook(summary, PARTICIPANT_WORKBOOK)

    SEGMENT_NOTES_MD.write_text("\n".join(notes), encoding="utf-8")

    total = len(coded_rows)
    p_count = sum(1 for r in coded_rows if r["speaker_type"] == "participant")
    m_count = sum(1 for r in coded_rows if r["speaker_type"] == "moderator")
    u_count = total - p_count - m_count
    print(f"Total segments: {total}")
    print(f"  participant: {p_count}  moderator: {m_count}  unclear: {u_count}")
    print(f"Participant summary rows: {len(summary)}")
    print(f"Question evidence rows: {len(evidence)}")
    print(f"Files written:")
    print(f"  {CODED_SEGMENTS_CSV}")
    print(f"  {PARTICIPANT_SUMMARY_CSV}")
    print(f"  {QUESTION_EVIDENCE_CSV}")
    print(f"  {PARTICIPANT_WORKBOOK}")
    print(f"  {SEGMENT_NOTES_MD}")


if __name__ == "__main__":
    main()
