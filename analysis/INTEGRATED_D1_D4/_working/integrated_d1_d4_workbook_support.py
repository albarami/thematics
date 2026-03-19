from __future__ import annotations

import shutil
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from integrated_d1_d4_aggregation import extract_theme_number, matrix_metric, truncate
from integrated_d1_d4_constants import INTEGRATED_THEMES, INTERNAL_DIR, KNOWN_NAME_SCAN_TERMS, OUTWARD_DIR


def build_participant_master_rows(
    case_packages: dict[str, dict[str, Any]],
    quote_lookup: dict[tuple[str, str], dict[str, str]],
) -> list[dict[str, str]]:
    """Build integrated participant-workbook rows from outward-facing packages.

    Args:
        case_packages: Loaded outward-facing case package data.
        quote_lookup: Excerpt-bank lookup keyed by case and evidence ID.

    Returns:
        Participant-master rows using anonymized source-package fields only.
    """
    del quote_lookup
    rows: list[dict[str, str]] = []
    for case_id, package in case_packages.items():
        day_label = package["config"]["day_label"]
        excerpt_by_speaker: dict[str, list[dict[str, str]]] = defaultdict(list)
        for row in package["excerpt_bank"]:
            speaker_code = row.get("speaker_code", "")
            if speaker_code:
                excerpt_by_speaker[speaker_code].append(row)
        register_by_code = {row["anonymized_code"]: row for row in package["participant_register"]}
        for summary in package["participant_summary"]:
            if summary.get("speaker_type") == "moderator":
                continue
            speaker_code = summary["anonymized_code"]
            excerpt_rows = excerpt_by_speaker.get(speaker_code, [])
            linked_themes = sorted(
                {
                    f"IT{extract_theme_number(row.get('theme', ''))}"
                    for row in excerpt_rows
                    if extract_theme_number(row.get("theme", ""))
                }
            )
            notable_quote = truncate(excerpt_rows[0].get("excerpt_text", ""), 180) if excerpt_rows else ""
            language_note = ";".join(sorted({row.get("language", "") for row in excerpt_rows if row.get("language", "")}))
            register_row = register_by_code.get(speaker_code, {})
            rows.append(
                {
                    "participant_id": speaker_code,
                    "participant_name_or_role": register_row.get("role_label", summary.get("speaker_type", "participant")),
                    "case_id": case_id,
                    "day_label": day_label,
                    "source_file": summary.get("source_file", ""),
                    "question_id": summary.get("questions_covered", ""),
                    "response_summary": f"Contributed across {summary.get('questions_covered', '')} with top codes {summary.get('top_codes', '')}",
                    "linked_codes": summary.get("top_codes", ""),
                    "linked_themes": ";".join(linked_themes),
                    "notable_quote": notable_quote,
                    "language_note": language_note,
                    "source_type": summary.get("speaker_type", "participant"),
                    "notes": register_row.get("classification_basis", ""),
                }
            )
    return rows


def build_question_workbook_sheets(case_packages: dict[str, dict[str, Any]]) -> dict[str, list[dict[str, str]]]:
    """Build integrated question-matrix workbook sheets from excerpt-bank rows.

    Args:
        case_packages: Loaded outward-facing case package data.

    Returns:
        Sheet-name to row-list mapping for `Q1-Q7`.
    """
    sheets: dict[str, list[dict[str, str]]] = {f"Q{index}_Matrix": [] for index in range(1, 8)}
    for case_id, package in case_packages.items():
        for row in package["excerpt_bank"]:
            question_id = row.get("question_id", "")
            if not question_id:
                continue
            theme_number = extract_theme_number(row.get("theme", ""))
            sheet_name = f"{question_id}_Matrix"
            sheets[sheet_name].append(
                {
                    "question_id": question_id,
                    "case_id": case_id,
                    "participant_id_or_role": row.get("speaker_code", row.get("role_label", "")),
                    "source_file": row.get("source_file", ""),
                    "summary_of_response": truncate(row.get("excerpt_text", ""), 220),
                    "linked_code": row.get("codes", ""),
                    "linked_theme": f"IT{theme_number}" if theme_number else "",
                    "prominence_within_question": question_level_strength(package["question_matrix"], question_id, theme_number),
                    "quotation_reference": row.get("evidence_id", ""),
                    "divergence_note": divergence_note(row),
                    "notes": f"evidence_type={row.get('evidence_type', '')}; report_use={row.get('report_use', '')}",
                }
            )
    return sheets


def build_theme_workbook_sheets(case_packages: dict[str, dict[str, Any]]) -> dict[str, list[dict[str, str]]]:
    """Build integrated theme-evidence workbook sheets from approved excerpt banks.

    Args:
        case_packages: Loaded outward-facing case package data.

    Returns:
        Sheet-name to row-list mapping for the integrated theme workbook.
    """
    sheets: dict[str, list[dict[str, str]]] = {"Themes_Master": []}
    for theme in INTEGRATED_THEMES:
        sheets[f"Theme{theme['theme_number']}_Evidence"] = []
    for case_id, package in case_packages.items():
        for row in package["theme_summary"]:
            theme_number = row.get("theme_number", "")
            sheets["Themes_Master"].append(
                {
                    "theme_name": f"IT{theme_number}",
                    "case_id": case_id,
                    "question_id": row.get("questions_present", ""),
                    "participant_id_or_role": row.get("unique_speakers", ""),
                    "source_file": package["config"]["label"],
                    "source_type": "case_theme_summary",
                    "excerpt_reference": "",
                    "quotation": row.get("theme_name", ""),
                    "summary_of_point": row.get("theme_name", ""),
                    "prominence_note": row.get("salience", ""),
                    "divergence_or_tension": package["config"]["case_variation"],
                    "notes": f"segments={row.get('participant_segments', '')}; tables={row.get('unique_tables', '')}",
                }
            )
        for excerpt in package["excerpt_bank"]:
            theme_number = extract_theme_number(excerpt.get("theme", ""))
            if not theme_number:
                continue
            sheets[f"Theme{theme_number}_Evidence"].append(
                {
                    "theme_name": f"IT{theme_number}",
                    "case_id": case_id,
                    "question_id": excerpt.get("question_id", ""),
                    "participant_id_or_role": excerpt.get("speaker_code", ""),
                    "source_file": excerpt.get("source_file", ""),
                    "source_type": excerpt.get("evidence_type", ""),
                    "excerpt_reference": excerpt.get("evidence_id", ""),
                    "quotation": truncate(excerpt.get("excerpt_text", ""), 240),
                    "summary_of_point": truncate(excerpt.get("excerpt_text", ""), 120),
                    "prominence_note": excerpt.get("report_use", ""),
                    "divergence_or_tension": divergence_note(excerpt),
                    "notes": excerpt.get("codes", ""),
                }
            )
    return sheets


def write_workbook(path: Path, sheets: dict[str, list[dict[str, str]]]) -> None:
    """Write an ordered workbook from sheet-to-row mappings.

    Args:
        path: Destination workbook path.
        sheets: Mapping of sheet names to uniform row dictionaries.

    Returns:
        None. Writes the workbook to disk.
    """
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, rows in sheets.items():
        worksheet = workbook.create_sheet(title=sheet_name[:31])
        if not rows:
            continue
        headers = list(rows[0].keys())
        worksheet.append(headers)
        for row in rows:
            worksheet.append([row.get(header, "") for header in headers])
    workbook.save(path)


def prepare_output_dirs() -> None:
    """Create clean outward and internal package directories."""
    for directory in (OUTWARD_DIR, INTERNAL_DIR):
        directory.mkdir(parents=True, exist_ok=True)
        for child in directory.iterdir():
            if child.is_dir():
                shutil.rmtree(child)
            else:
                child.unlink()


def copy_to_outward(paths: list[Path]) -> None:
    """Copy integrated outward-facing output files into the outward package folder."""
    for path in paths:
        shutil.copy2(path, OUTWARD_DIR / path.name)


def write_internal_markdown(path: Path, content: str) -> None:
    """Write an internal markdown file to disk."""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def scan_output_paths(paths: list[Path]) -> dict[str, list[str]]:
    """Scan integrated outward-facing outputs for blocked terms and exclusions."""
    hits: dict[str, list[str]] = {}
    for path in paths:
        found = _scan_workbook(path) if path.suffix.lower() == ".xlsx" else _scan_text(path)
        if found:
            hits[path.name] = found
    return hits


def divergence_note(row: dict[str, str]) -> str:
    """Build a short evidence-handling note from an excerpt-bank row."""
    notes: list[str] = []
    speaker_type = row.get("speaker_type", "")
    evidence_type = row.get("evidence_type", "")
    report_use = row.get("report_use", "")
    if speaker_type == "unclear":
        notes.append("attribution-cautious `unclear` support")
    if evidence_type == "note_taker_summary":
        notes.append("note-based contextual support")
    if evidence_type == "note_style_transcript_summary":
        notes.append("note-style transcript summary")
    if report_use == "close_reading_theme_support":
        notes.append("close-reading support")
    if report_use == "contextual_theme_support":
        notes.append("contextual support")
    return "; ".join(notes)


def question_level_strength(
    matrix_rows: list[dict[str, str]],
    question_id: str,
    theme_number: str,
) -> str:
    """Label question-level theme strength from a case question-theme matrix."""
    if not theme_number:
        return "not_mapped"
    match = next((row for row in matrix_rows if row.get("question_id") == question_id), None)
    if match is None:
        return "not_mapped"
    segments = matrix_metric(match, theme_number, "segments")
    if segments >= 25:
        return "strong"
    if segments >= 10:
        return "moderate"
    if segments > 0:
        return "present"
    return "explicit_absence"


def _scan_text(path: Path) -> list[str]:
    text = path.read_text(encoding="utf-8")
    return [term for term in KNOWN_NAME_SCAN_TERMS if term in text]


def _scan_workbook(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    hits: set[str] = set()
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows(values_only=True):
            for value in row:
                if not isinstance(value, str):
                    continue
                for term in KNOWN_NAME_SCAN_TERMS:
                    if term in value:
                        hits.add(term)
    return sorted(hits)
